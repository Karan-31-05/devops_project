"""
Anna University CSE Department ERP System
HOD (Head of Department) Views
"""

import json
import csv
import io
import re
import requests
import uuid
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponseRedirect, get_object_or_404, redirect, render)
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.db import transaction, IntegrityError, router
from django.db.models.deletion import Collector
from datetime import datetime
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes

from .forms import (
    FacultyRegistrationForm, StudentRegistrationForm, NonTeachingStaffRegistrationForm,
    CourseForm, CourseAssignmentForm, RegulationForm, AcademicYearForm, SemesterForm,
    EventForm, LeaveApprovalForm, FeedbackReplyForm, AnnouncementForm,
    FacultyProfileEditForm, StudentProfileEditForm, AccountUserForm,
    QuestionPaperAssignmentForm, QuestionPaperReviewForm,
    TimetableForm, TimetableEntryForm, TimeSlotForm, ProgramForm,
    ExamScheduleForm, ExamScheduleEditForm
)
from .models import (
    Account_User, Faculty_Profile, Student_Profile, NonTeachingStaff_Profile,
    Course, Course_Assignment, Attendance, Regulation, CourseCategory, AcademicYear, Semester,
    Publication, Student_Achievement, Lab_Issue_Log, LeaveRequest, Feedback,
    Event, EventRegistration, Notification, Announcement, QuestionPaperAssignment,
    Timetable, TimetableEntry, TimeSlot, Program, ExamSchedule, StructuredQuestionPaper,
    RegulationCoursePlan, SemesterPromotion, ProgramRegulation,
    ProgramBatch, ElectiveVertical, ElectiveCourseOffering, ElectiveOfferingFacultyAssignment,
    LabRoom, LabRestriction, TimetableConfig, FixedSlotReservation, TimetableConfigLab,
    SameTimeConstraint, MELabAssistConstraint, FacultyTimeBlock, PECCourseCombinationRule,
    PECGroupConfig, ClubbedCourseGroup, ClubbedCourseMember, ProgramSemesterDate
)
from .utils.web_scrapper import fetch_acoe_updates
from .utils.cir_scrapper import fetch_cir_ticker_announcements
from .utils.qp_checklist_doc import append_checklist_to_document
def check_hod_permission(user):
    """
    Check if user has HOD privileges.
    HOD is identified via Faculty_Profile.designation == 'HOD',
    not by Account_User.role field.
    """
    if not user.is_authenticated:
        return False
    return user.is_hod


def _canonical_placeholder_title(placeholder_type):
    """Return a stable display title for a placeholder type."""
    if not placeholder_type:
        return ''

    roman_suffix_pattern = re.compile(r'\s*-\s*[IVX]+\s*$', re.IGNORECASE)
    generic_pattern = re.compile(r'^[A-Z0-9\s]+\s*-\s*[IVX]+\s*$', re.IGNORECASE)

    typed_courses = Course.objects.filter(
        is_placeholder=True,
        placeholder_type=placeholder_type
    ).order_by('slot_number', 'course_code')

    best_title = ''
    for item in typed_courses:
        title = (item.title or '').strip()
        if not title:
            continue
        if not generic_pattern.match(title):
            best_title = title
            break
        if not best_title:
            best_title = roman_suffix_pattern.sub('', title).strip()

    if best_title:
        return best_title

    return dict(Course.PLACEHOLDER_TYPE_CHOICES).get(placeholder_type, placeholder_type)


# =============================================================================
# HOD VIEW MODE TOGGLE
# =============================================================================

@login_required
def toggle_hod_view_mode(request):
    """Toggle HOD between Admin and Faculty view modes"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Toggle the view mode
    current_mode = request.session.get('hod_view_mode', 'admin')
    new_mode = 'faculty' if current_mode == 'admin' else 'admin'
    request.session['hod_view_mode'] = new_mode
    
    # Redirect to appropriate dashboard
    if new_mode == 'faculty':
        messages.success(request, "Switched to Faculty View")
        return redirect('staff_home')
    else:
        messages.success(request, "Switched to Admin View")
        return redirect('admin_home')


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def admin_home(request):
    """HOD Dashboard with statistics and overview"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Statistics
    total_faculty = Faculty_Profile.objects.filter(is_external=False).count()
    total_guest_faculty = Faculty_Profile.objects.filter(is_external=True).count()
    total_students = Student_Profile.objects.count()
    total_staff = Faculty_Profile.objects.filter(is_external=False).count()
    total_courses = Course.objects.count()
    total_course = total_courses  # Alias for template
    total_assignments = Course_Assignment.objects.count()
    
    # Students by branch
    students_by_branch = Student_Profile.objects.values('branch').annotate(count=Count('branch'))
    branch_labels = [item['branch'] for item in students_by_branch]
    branch_counts = [item['count'] for item in students_by_branch]
    
    # Students by batch
    students_by_batch = Student_Profile.objects.values('batch_label').annotate(count=Count('batch_label'))
    batch_labels = [item['batch_label'] for item in students_by_batch]
    batch_counts = [item['count'] for item in students_by_batch]
    
    # Course list for charts
    courses = Course.objects.all()
    course_list = [c.title for c in courses]
    course_name_list = course_list  # Alias for template
    
    # Students per course (via course assignments)
    student_count_list_in_course = []
    assignment_count_list = []
    attendance_list = []
    for course in courses:
        # Count assignments for this course
        assignments = Course_Assignment.objects.filter(course=course)
        assignment_count_list.append(assignments.count())
        # Attendance per course
        att_count = Attendance.objects.filter(assignment__course=course).count()
        attendance_list.append(att_count)
        # Student count is just a placeholder - we'd need to filter by batch
        student_count_list_in_course.append(Student_Profile.objects.count())
    
    # Student attendance and leave stats
    students = Student_Profile.objects.all()[:20]  # Limit for chart
    student_name_list = [s.user.full_name for s in students]
    student_attendance_present_list = []
    student_attendance_leave_list = []
    for student in students:
        present = Attendance.objects.filter(student=student, status='PRESENT').count()
        absent = Attendance.objects.filter(student=student, status='ABSENT').count()
        student_attendance_present_list.append(present)
        student_attendance_leave_list.append(absent)
    
    # Pending items
    pending_leaves = LeaveRequest.objects.filter(status='PENDING').count()
    pending_feedbacks = Feedback.objects.filter(status='PENDING').count()
    pending_lab_issues = Lab_Issue_Log.objects.filter(status='PENDING').count()
    unverified_publications = Publication.objects.filter(is_verified=False).count()
    
    # Recent activities
    recent_leaves = LeaveRequest.objects.order_by('-created_at')[:5]
    recent_feedbacks = Feedback.objects.order_by('-created_at')[:5]
    
    # Fetch external announcements
    announcements = []
    try:
        acoe_updates = fetch_acoe_updates()
        announcements.extend(acoe_updates)
    except:
        pass
    
    try:
        cir_announcements = fetch_cir_ticker_announcements(limit=5)
        announcements.extend(cir_announcements)
    except:
        pass
    
    # Department announcements
    dept_announcements = Announcement.objects.filter(is_active=True)[:5]
    
    # Current academic context (auto-detected from dates)
    current_year = AcademicYear.get_current()
    current_semester = Semester.get_current()

    context = {
        'page_title': "HOD Dashboard - CSE Department",
        'total_faculty': total_faculty,
        'total_guest_faculty': total_guest_faculty,
        'total_students': total_students,
        'total_staff': total_staff,
        'total_courses': total_courses,
        'total_course': total_course,
        'total_assignments': total_assignments,
        'branch_labels': json.dumps(branch_labels),
        'branch_counts': json.dumps(branch_counts),
        'batch_labels': json.dumps(batch_labels),
        'batch_counts': json.dumps(batch_counts),
        'pending_leaves': pending_leaves,
        'pending_feedbacks': pending_feedbacks,
        'pending_lab_issues': pending_lab_issues,
        'unverified_publications': unverified_publications,
        'recent_leaves': recent_leaves,
        'recent_feedbacks': recent_feedbacks,
        'announcements': announcements,
        'dept_announcements': dept_announcements,
        'current_year': current_year,
        'current_semester': current_semester,
        # Chart data
        'course_list': json.dumps(course_list),
        'course_name_list': json.dumps(course_name_list),
        'student_count_list_in_course': json.dumps(student_count_list_in_course),
        'assignment_count_list': json.dumps(assignment_count_list),
        'attendance_list': json.dumps(attendance_list),
        'student_name_list': json.dumps(student_name_list),
        'student_attendance_present_list': json.dumps(student_attendance_present_list),
        'student_attendance_leave_list': json.dumps(student_attendance_leave_list),
    }
    return render(request, 'hod_template/home_content.html', context)


# =============================================================================
# FACULTY MANAGEMENT
# =============================================================================

@login_required
def add_faculty(request):
    """Add new faculty member"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = FacultyRegistrationForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Faculty'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                is_external = form.cleaned_data.get('is_external', False)
                provided_email = (form.cleaned_data.get('email') or '').strip().lower()
                staff_id = form.cleaned_data['staff_id'].strip().upper()

                # External faculty may not provide an email; keep a deterministic unique placeholder email.
                if is_external and not provided_email:
                    email = f"{staff_id.lower()}@guest.local"
                else:
                    email = provided_email

                # Create user
                user = Account_User.objects.create(
                    email=email,
                    full_name=form.cleaned_data['full_name'],
                    role='GUEST' if is_external else 'FACULTY',
                    gender=form.cleaned_data.get('gender'),
                    phone=form.cleaned_data.get('phone') or None,
                    address=form.cleaned_data.get('address') or None,
                    is_active=True,
                )

                if is_external:
                    user.set_password(form.cleaned_data['password'])
                else:
                    user.set_unusable_password()
                user.save()
                
                # Handle profile pic
                if 'profile_pic' in request.FILES:
                    fs = FileSystemStorage()
                    filename = fs.save(request.FILES['profile_pic'].name, request.FILES['profile_pic'])
                    user.profile_pic = fs.url(filename)
                    user.save()
                
                # Update faculty profile
                faculty = user.faculty_profile
                faculty.staff_id = staff_id
                faculty.designation = form.cleaned_data['designation']
                faculty.is_external = is_external
                faculty.specialization = form.cleaned_data.get('specialization')
                faculty.qualification = form.cleaned_data.get('qualification')
                faculty.experience_years = form.cleaned_data.get('experience_years', 0)
                faculty.date_of_joining = form.cleaned_data.get('date_of_joining')
                faculty.contract_expiry = form.cleaned_data.get('contract_expiry')
                faculty.delete_guest_after_inactive_semesters = form.cleaned_data.get(
                    'delete_guest_after_inactive_semesters', 0
                ) or 0
                faculty.cabin_number = form.cleaned_data.get('cabin_number')
                faculty.save()

                if is_external:
                    if provided_email:
                        messages.success(request, "External faculty added successfully. Login with Email/Staff ID and assigned password.")
                    else:
                        messages.success(request, "External faculty added successfully. Login with Staff ID and assigned password.")
                else:
                    email_sent, email_error = send_password_setup_email(request, user)
                    if email_sent:
                        messages.success(request, "Faculty added successfully. Password setup email has been sent.")
                    else:
                        messages.warning(
                            request,
                            f"Faculty added, but password setup email could not be sent. Reason: {email_error}"
                        )
                
                return redirect(reverse('add_faculty'))
            except Exception as e:
                messages.error(request, f"Could not add faculty: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_staff_template.html', context)


@login_required
def manage_faculty(request):
    """List all faculty members"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    faculty_list = Faculty_Profile.objects.select_related('user').all().order_by('staff_id')
    current_hod = faculty_list.filter(user__role='HOD').first()
    hod_candidates = faculty_list.filter(
        user__is_active=True,
        is_external=False,
    ).exclude(user__role='HOD')

    context = {
        'faculty_list': faculty_list,
        'current_hod': current_hod,
        'hod_candidates': hod_candidates,
        'page_title': 'Manage Faculty'
    }
    return render(request, "hod_template/manage_staff.html", context)


@login_required
def change_hod(request, faculty_id=None):
    """Transfer HOD role to another faculty with optional pending workflow reassignment."""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')

    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect(reverse('manage_faculty'))

    selected_faculty_id = faculty_id or request.POST.get('faculty_id')
    if not selected_faculty_id:
        messages.error(request, "Select a faculty member to transfer HOD role.")
        return redirect(reverse('manage_faculty'))

    current_hod_password = request.POST.get('current_hod_password', '')
    if not current_hod_password or not request.user.check_password(current_hod_password):
        messages.error(request, "Current HOD password is incorrect. Transfer cancelled.")
        return redirect(reverse('manage_faculty'))

    target_faculty = get_object_or_404(
        Faculty_Profile.objects.select_related('user'),
        id=selected_faculty_id,
    )
    target_user = target_faculty.user

    if target_faculty.is_external or target_user.role == 'GUEST':
        messages.error(request, "Guest/External faculty cannot be assigned as HOD.")
        return redirect(reverse('manage_faculty'))

    if not target_user.is_active:
        messages.error(request, "Inactive users cannot be assigned as HOD.")
        return redirect(reverse('manage_faculty'))

    if target_user.pk == request.user.pk:
        messages.warning(request, "You are already the current HOD.")
        return redirect(reverse('manage_faculty'))

    reassign_pending = str(request.POST.get('reassign_pending', '0')).strip().lower() in {'1', 'true', 'yes', 'on'}

    def _ensure_profile_for_user(user_obj):
        """Create a minimal faculty profile if missing so FACULTY role users always resolve in staff views."""
        if Faculty_Profile.objects.filter(user=user_obj).exists():
            return False

        staff_id = f"TEMP_{uuid.uuid4().hex[:8].upper()}"
        while Faculty_Profile.objects.filter(staff_id=staff_id).exists():
            staff_id = f"TEMP_{uuid.uuid4().hex[:8].upper()}"

        Faculty_Profile.objects.create(
            user=user_obj,
            staff_id=staff_id,
            designation='AP',
            is_external=(user_obj.role == 'GUEST'),
        )
        return True

    try:
        with transaction.atomic():
            target_user = Account_User.objects.select_for_update().get(pk=target_user.pk)

            previous_hod_qs = Account_User.objects.select_for_update().filter(role='HOD').exclude(pk=target_user.pk)
            previous_hod_ids = list(previous_hod_qs.values_list('id', flat=True))
            created_profile_count = 0

            for previous_hod in previous_hod_qs:
                if _ensure_profile_for_user(previous_hod):
                    created_profile_count += 1

            if target_user.role != 'HOD':
                target_user.role = 'HOD'
                target_user.save(update_fields=['role'])

            if previous_hod_ids:
                Account_User.objects.filter(id__in=previous_hod_ids).update(role='FACULTY')

            reassignment_counts = {
                'assigned_by': 0,
                'reviewed_qp_assignment': 0,
                'reviewed_structured_qp': 0,
            }

            if reassign_pending and previous_hod_ids:
                pending_assignment_statuses = [
                    'ASSIGNED',
                    'IN_PROGRESS',
                    'SUBMITTED',
                    'UNDER_REVIEW',
                    'REVISION_REQUIRED',
                ]
                pending_structured_qp_statuses = ['SUBMITTED', 'UNDER_REVIEW']

                reassignment_counts['assigned_by'] = QuestionPaperAssignment.objects.filter(
                    assigned_by_id__in=previous_hod_ids,
                    status__in=pending_assignment_statuses,
                ).update(assigned_by=target_user)

                reassignment_counts['reviewed_qp_assignment'] = QuestionPaperAssignment.objects.filter(
                    reviewed_by_id__in=previous_hod_ids,
                    status__in=pending_assignment_statuses,
                ).update(reviewed_by=target_user)

                reassignment_counts['reviewed_structured_qp'] = StructuredQuestionPaper.objects.filter(
                    reviewed_by_id__in=previous_hod_ids,
                    status__in=pending_structured_qp_statuses,
                ).update(reviewed_by=target_user)

        reassigned_total = sum(reassignment_counts.values())
        if reassign_pending and reassigned_total:
            transfer_msg = f"HOD transferred to {target_user.full_name}. Reassigned {reassigned_total} pending workflow record(s)."
            if created_profile_count:
                transfer_msg += f" Auto-created {created_profile_count} missing faculty profile(s) for demoted users."
            messages.success(
                request,
                transfer_msg,
            )
        elif reassign_pending:
            transfer_msg = f"HOD transferred to {target_user.full_name}. No pending workflow records required reassignment."
            if created_profile_count:
                transfer_msg += f" Auto-created {created_profile_count} missing faculty profile(s) for demoted users."
            messages.success(
                request,
                transfer_msg,
            )
        else:
            transfer_msg = f"HOD transferred to {target_user.full_name}."
            if created_profile_count:
                transfer_msg += f" Auto-created {created_profile_count} missing faculty profile(s) for demoted users."
            messages.success(request, transfer_msg)
    except Exception as exc:
        messages.error(request, f"Could not transfer HOD role: {exc}")

    return redirect(reverse('manage_faculty'))


@login_required
def edit_faculty(request, faculty_id):
    """Edit faculty details"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    faculty = get_object_or_404(Faculty_Profile, id=faculty_id)
    user_form = AccountUserForm(request.POST or None, request.FILES or None, instance=faculty.user)
    user_form.fields['role'].disabled = True
    user_form.fields['role'].help_text = "Use the Set as HOD action in Manage Faculty for role transfer."
    profile_form = FacultyProfileEditForm(request.POST or None, instance=faculty)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'faculty': faculty,
        'page_title': 'Edit Faculty'
    }
    
    if request.method == 'POST':
        if user_form.is_valid() and profile_form.is_valid():
            try:
                user = user_form.save(commit=False)
                # Role transfer is handled by change_hod to keep one authoritative handover path.
                user.role = faculty.user.role
                password = user_form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                
                if 'profile_pic' in request.FILES:
                    fs = FileSystemStorage()
                    filename = fs.save(request.FILES['profile_pic'].name, request.FILES['profile_pic'])
                    user.profile_pic = fs.url(filename)
                
                user.save()
                profile_form.save()
                
                messages.success(request, "Faculty updated successfully!")
                return redirect(reverse('manage_faculty'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please fill the form correctly")
    
    return render(request, "hod_template/edit_staff_template.html", context)


@login_required
def delete_faculty(request, faculty_id):
    """Delete faculty"""
    if not check_hod_permission(request.user):
        return redirect('/')

    if request.method != 'POST':
        messages.error(request, "Invalid request method for delete.")
        return redirect(reverse('manage_faculty'))
    
    faculty = get_object_or_404(Faculty_Profile.objects.select_related('user'), id=faculty_id)
    force_delete = str(request.POST.get('force', '0')).strip().lower() in {'1', 'true', 'yes', 'on'}
    confirm_staff_id = (request.POST.get('confirm_staff_id') or '').strip().upper()

    if faculty.user.role == 'HOD':
        error_msg = "Cannot delete current HOD. Transfer HOD role first."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect(reverse('manage_faculty'))

    if faculty.user_id == request.user.id:
        error_msg = "You cannot delete your own account while logged in."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect(reverse('manage_faculty'))

    collector = Collector(using=router.db_for_write(Account_User))
    collector.collect([faculty.user])

    dependencies = []
    for model, objects in collector.data.items():
        if model in (Account_User, Faculty_Profile):
            continue
        count = len(objects)
        if count > 0:
            dependencies.append({'model': model.__name__, 'count': count})

    dependencies.sort(key=lambda item: (-item['count'], item['model']))

    if dependencies and not force_delete:
        warning_msg = "Faculty has linked records. Deletion is blocked by default."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {
                    'success': False,
                    'error': warning_msg,
                    'requires_force': True,
                    'staff_id': faculty.staff_id,
                    'dependencies': dependencies[:8],
                },
                status=400,
            )

        dependency_preview = ", ".join(
            [f"{item['model']} ({item['count']})" for item in dependencies[:5]]
        )
        messages.error(
            request,
            f"{warning_msg} Linked: {dependency_preview}. Submit force delete with confirmation to continue.",
        )
        return redirect(reverse('manage_faculty'))

    if dependencies and force_delete and confirm_staff_id != faculty.staff_id:
        error_msg = f"Confirmation failed. Type exact staff ID {faculty.staff_id} to force delete."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect(reverse('manage_faculty'))

    try:
        faculty.user.delete()  # This will cascade delete the profile
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'Faculty deleted successfully.'})
        messages.success(request, "Faculty deleted successfully!")
    except Exception as e:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_faculty'))


# =============================================================================
# STUDENT MANAGEMENT
# =============================================================================

@login_required
def add_student(request):
    """Add new student - password is set by student via OTP first-time login
    Fields match bulk upload for consistency.
    """
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = StudentRegistrationForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Student'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                # Create user without password (student will set via OTP)
                user = Account_User.objects.create(
                    email=form.cleaned_data['email'],
                    full_name=form.cleaned_data['full_name'],
                    role='STUDENT',
                    gender=form.cleaned_data['gender'],
                    phone=form.cleaned_data.get('phone') or None,
                    address=form.cleaned_data.get('address') or None,
                    is_active=True
                )
                # Mark password as unusable until student sets it
                user.set_unusable_password()
                user.save()
                
                # Determine current_sem and admission_year based on entry_type
                entry_type = form.cleaned_data['entry_type']
                current_sem = 1 if entry_type == 'REGULAR' else 3
                from datetime import datetime
                admission_year = datetime.now().year
                
                # Update student profile (auto-created by signal)
                student = user.student_profile
                student.register_no = form.cleaned_data['register_no']
                student.batch_label = form.cleaned_data['batch_label']
                student.branch = form.cleaned_data['branch']
                student.program_type = form.cleaned_data['program_type']
                student.entry_type = entry_type
                student.admission_year = admission_year
                student.current_sem = current_sem
                student.parent_name = form.cleaned_data.get('parent_name') or None
                student.parent_phone = form.cleaned_data.get('parent_phone') or None
                
                # Auto-assign regulation using ProgramRegulation mapping
                # REGULAR: joins as 1st year → follows regulation of their admission year
                # LATERAL: joins as 2nd year → follows regulation of batch that started 1 year earlier
                regulation_lookup_year = admission_year if entry_type == 'REGULAR' else (admission_year - 1)
                
                # Use ProgramRegulation to find the correct regulation
                regulation = ProgramRegulation.get_regulation_for_student(
                    program_code=student.branch,
                    program_level=student.program_type,
                    admission_year=regulation_lookup_year
                )
                
                # Fallback to old method if no ProgramRegulation mapping exists
                if not regulation:
                    regulation = Regulation.objects.filter(
                        year__lte=regulation_lookup_year
                    ).order_by('-year').first()
                
                if regulation:
                    student.regulation = regulation
                
                student.save()
                
                # Send first-time login notification to college email
                try:
                    student_data = {
                        'name': user.full_name,
                        'register_no': student.register_no,
                        'email': user.email,
                        'college_email': student.college_email,
                    }
                    send_first_login_notification(student_data)
                    messages.success(request, f"Student added successfully! Login instructions sent to {student.college_email}")
                except Exception as e:
                    messages.warning(request, f"Student added but email notification failed: {str(e)}")
                
                return redirect(reverse('add_student'))
            except Exception as e:
                messages.error(request, f"Could not add student: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_student_template.html', context)


@login_required
def manage_student(request):
    """List all students with filters"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    students = Student_Profile.objects.select_related('user', 'advisor', 'regulation').filter(status='ACTIVE')
    
    # Apply filters
    branch = request.GET.get('branch')
    batch = request.GET.get('batch')
    year_of_study = request.GET.get('year')
    program_level = request.GET.get('level')  # UG or PG
    
    if program_level:
        students = students.filter(program_type=program_level)
    if branch:
        students = students.filter(branch=branch)
    if year_of_study:
        # Year 1 = Sem 1,2 | Year 2 = Sem 3,4 | Year 3 = Sem 5,6 | Year 4 = Sem 7,8
        year_int = int(year_of_study)
        sem_start = (year_int - 1) * 2 + 1
        sem_end = year_int * 2
        students = students.filter(current_sem__gte=sem_start, current_sem__lte=sem_end)
    if batch:
        students = students.filter(batch_label=batch)
    
    # Order by register number
    students = students.order_by('register_no')
    
    # Get unique program codes for branch filtering (avoid duplicates like multiple CSE entries)
    # Filter by level if selected
    programs_qs = Program.objects.all()
    if program_level:
        programs_qs = programs_qs.filter(level=program_level)
    
    branch_choices = list(programs_qs.values('code', 'name', 'level').distinct().order_by('level', 'code'))
    # Deduplicate by code - keep first occurrence
    seen_codes = set()
    unique_branches = []
    for prog in branch_choices:
        if prog['code'] not in seen_codes:
            seen_codes.add(prog['code'])
            unique_branches.append(prog)
    
    # Get batch choices from database - filter by level if selected
    current_year = AcademicYear.get_current()
    if current_year:
        batch_qs = ProgramBatch.objects.filter(
            academic_year=current_year,
            is_active=True
        )
        if program_level:
            batch_qs = batch_qs.filter(program__level=program_level)
        batch_choices = list(batch_qs.values_list('batch_name', flat=True).distinct().order_by('batch_name'))
    else:
        batch_choices = []
    
    # Get counts by year of study for quick stats
    year_counts = {}
    base_qs = Student_Profile.objects.filter(status='ACTIVE')
    if program_level:
        base_qs = base_qs.filter(program_type=program_level)
    if branch:
        base_qs = base_qs.filter(branch=branch)
    if batch:
        base_qs = base_qs.filter(batch_label=batch)
    
    for yr in range(1, 5):
        sem_start = (yr - 1) * 2 + 1
        sem_end = yr * 2
        year_counts[yr] = base_qs.filter(
            current_sem__gte=sem_start,
            current_sem__lte=sem_end
        ).count()
    
    total_all_years = sum(year_counts.values())
    
    context = {
        'students': students,
        'page_title': 'Manage Students',
        'branch_choices': unique_branches,
        'batch_choices': batch_choices,
        'year_counts': year_counts,
        'year_1_count': year_counts.get(1, 0),
        'year_2_count': year_counts.get(2, 0),
        'year_3_count': year_counts.get(3, 0),
        'year_4_count': year_counts.get(4, 0),
        'total_all_years': total_all_years,
        'selected_branch': branch or '',
        'selected_batch': batch or '',
        'selected_year': year_of_study or '',
        'selected_level': program_level or '',
        'total_students': students.count(),
    }
    return render(request, "hod_template/manage_student.html", context)


@login_required
def edit_student(request, student_id):
    """Edit student details"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    student = get_object_or_404(Student_Profile, id=student_id)
    user_form = AccountUserForm(request.POST or None, request.FILES or None, instance=student.user)
    profile_form = StudentProfileEditForm(request.POST or None, instance=student)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'student': student,
        'page_title': 'Edit Student'
    }
    
    if request.method == 'POST':
        if user_form.is_valid() and profile_form.is_valid():
            try:
                user = user_form.save(commit=False)
                password = user_form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                
                if 'profile_pic' in request.FILES:
                    fs = FileSystemStorage()
                    filename = fs.save(request.FILES['profile_pic'].name, request.FILES['profile_pic'])
                    user.profile_pic = fs.url(filename)
                
                user.save()
                profile_form.save()
                
                messages.success(request, "Student updated successfully!")
                return redirect(reverse('manage_student'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please fill the form correctly")
    
    return render(request, "hod_template/edit_student_template.html", context)


@login_required
def delete_student(request, student_id):
    """Delete student"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    student = get_object_or_404(Student_Profile, id=student_id)
    try:
        student.user.delete()
        messages.success(request, "Student deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_student'))


@login_required
def download_students_excel(request):
    """Download student list as Excel with user-selected fields, filters, and sheet layout."""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect(reverse('manage_student'))
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # ----- Field definitions (key → header, extractor) -----
    FIELD_MAP = {
        'sno':            ('S.No',            None),  # handled separately
        'register_no':    ('Register No',     lambda s: s.register_no),
        'name':           ('Student Name',    lambda s: s.user.full_name),
        'email':          ('Email',           lambda s: s.user.email),
        'college_email':  ('College Email',   lambda s: s.college_email),
        'gender':         ('Gender',          lambda s: s.user.get_gender_display() if s.user.gender else ''),
        'phone':          ('Phone',           lambda s: s.user.phone or ''),
        'branch':         ('Branch',          lambda s: s.branch),
        'program_type':   ('Program Type',    lambda s: s.program_type),
        'year_of_study':  ('Year of Study',   lambda s: s.year_of_study),
        'current_sem':    ('Current Semester', lambda s: s.current_sem),
        'section':        ('Section',         lambda s: s.batch_label),
        'entry_type':     ('Entry Type',      lambda s: (s.entry_type or '')),
        'admission_year': ('Admission Year',  lambda s: s.admission_year or ''),
        'regulation':     ('Regulation',      lambda s: str(s.regulation) if s.regulation else ''),
        'parent_name':    ('Parent Name',     lambda s: s.parent_name or ''),
        'parent_phone':   ('Parent Phone',    lambda s: s.parent_phone or ''),
    }
    
    # ----- Read POST params -----
    selected_fields = request.POST.getlist('fields')
    years = request.POST.getlist('years')
    levels = request.POST.getlist('levels')
    entry_types = request.POST.getlist('entry_types')
    branches = request.POST.getlist('branches')
    sections = request.POST.getlist('sections')
    statuses = request.POST.getlist('statuses')
    sheet_mode = request.POST.get('sheet_mode', 'single')
    
    # Default: at least some fields
    if not selected_fields:
        selected_fields = ['sno', 'register_no', 'name', 'email', 'branch', 'year_of_study', 'section']
    
    # Default status if none selected
    if not statuses:
        statuses = ['ACTIVE']
    
    # ----- Build queryset -----
    students = Student_Profile.objects.select_related('user', 'regulation').filter(
        status__in=statuses
    )
    
    if levels:
        students = students.filter(program_type__in=levels)
    if entry_types:
        students = students.filter(entry_type__in=entry_types)
    if branches:
        students = students.filter(branch__in=branches)
    if sections:
        students = students.filter(batch_label__in=sections)
    if years:
        from django.db.models import Q as Q_filter
        year_q = Q_filter()
        for y in years:
            yr = int(y)
            sem_start = (yr - 1) * 2 + 1
            sem_end = yr * 2
            year_q |= Q_filter(current_sem__gte=sem_start, current_sem__lte=sem_end)
        students = students.filter(year_q)
    
    students = students.order_by('register_no')
    
    # ----- Prepare ordered field keys/headers -----
    # Preserve order from FIELD_MAP keys (which preserves insertion order in Python 3.7+)
    ordered_fields = [f for f in FIELD_MAP.keys() if f in selected_fields]
    headers = [FIELD_MAP[f][0] for f in ordered_fields]
    
    # ----- Helper: extract a row for a student -----
    def make_row(student, sno):
        row = []
        for field_key in ordered_fields:
            if field_key == 'sno':
                row.append(sno)
            else:
                extractor = FIELD_MAP[field_key][1]
                try:
                    row.append(extractor(student))
                except Exception:
                    row.append('')
        return row
    
    # ----- Helper: style a worksheet -----
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def write_sheet(ws, student_list):
        """Write headers and student data to a worksheet."""
        # Write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, student in enumerate(student_list, 1):
            row_data = make_row(student, row_idx)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
        
        # Auto-size columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 40)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    # ----- Create workbook -----
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    if sheet_mode == 'by_year':
        year_labels = {1: '1st Year', 2: '2nd Year', 3: '3rd Year', 4: '4th Year'}
        for yr in range(1, 5):
            sem_start = (yr - 1) * 2 + 1
            sem_end = yr * 2
            year_students = [s for s in students if sem_start <= s.current_sem <= sem_end]
            if year_students or not years:  # Always create sheet if no year filter
                ws = wb.create_sheet(title=year_labels[yr])
                write_sheet(ws, year_students)
    
    elif sheet_mode == 'by_branch':
        # Group students by branch
        branch_groups = {}
        for student in students:
            br = student.branch or 'Unknown'
            if br not in branch_groups:
                branch_groups[br] = []
            branch_groups[br].append(student)
        
        if branch_groups:
            for br in sorted(branch_groups.keys()):
                # Sheet titles max 31 chars
                sheet_title = br[:31]
                ws = wb.create_sheet(title=sheet_title)
                write_sheet(ws, branch_groups[br])
        else:
            ws = wb.create_sheet(title='No Data')
            write_sheet(ws, [])
    
    else:  # single sheet
        ws = wb.create_sheet(title='All Students')
        write_sheet(ws, list(students))
    
    # Ensure at least one sheet exists
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title='No Data')
        write_sheet(ws, [])
    
    # ----- Build filename -----
    from django.utils import timezone
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"students_{timestamp}.xlsx"
    
    # ----- Return response -----
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# COURSE MANAGEMENT
# =============================================================================

@login_required
def add_course(request):
    """Add new course"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = CourseForm(request.POST or None, request.FILES or None)
    context = {
        'form': form,
        'page_title': 'Add Course',
        'placeholder_type_choices': Course.PLACEHOLDER_TYPE_CHOICES,
    }
    
    if request.method == 'POST':
        if request.POST.get('action') == 'create_placeholder':
            placeholder_name = (request.POST.get('placeholder_name') or '').strip()
            placeholder_alias = (request.POST.get('placeholder_alias') or '').strip().upper()

            if not placeholder_name:
                messages.error(request, "Placeholder name is required")
                return render(request, 'hod_template/add_course_template.html', context)

            if not placeholder_alias:
                messages.error(request, "Placeholder alias is required")
                return render(request, 'hod_template/add_course_template.html', context)

            if not re.match(r'^[A-Z0-9]{2,5}$', placeholder_alias):
                messages.error(request, "Alias must be 2-5 uppercase letters/numbers (example: SDC, IOC)")
                return render(request, 'hod_template/add_course_template.html', context)

            existing = Course.objects.filter(
                is_placeholder=True,
                placeholder_type=placeholder_alias
            ).order_by('slot_number', 'course_code').first()

            if existing:
                messages.info(
                    request,
                    f"Placeholder course already available for alias {placeholder_alias} ({existing.course_code})"
                )
                return redirect(reverse('add_course'))

            # Register placeholder alias once from Add Course page.
            # Additional slots are created only during regulation plan assignment.
            course, _ = Course.get_or_create_placeholder(
                placeholder_alias,
                1,
                credits=3,
                update_existing=False,
            )
            course.title = placeholder_name
            course.save(update_fields=['title', 'updated_at'])
            messages.success(request, f"Placeholder {course.course_code} created successfully")
            return redirect(reverse('add_course'))

        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Course added successfully!")
                return redirect(reverse('add_course'))
            except Exception as e:
                messages.error(request, f"Could not add course: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_course_template.html', context)


@login_required
def manage_course(request):
    """List all courses"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    include_placeholders = request.GET.get('include_placeholders') == '1'
    courses = Course.objects.filter(is_placeholder=False)
    
    # Apply filters
    course_type = request.GET.get('course_type')
    search = request.GET.get('search', '').strip()
    
    if course_type:
        courses = courses.filter(course_type=course_type)
    if search:
        courses = courses.filter(
            Q(course_code__icontains=search) | Q(title__icontains=search)
        )

    placeholder_rows = []
    if include_placeholders:
        placeholder_qs = Course.objects.filter(is_placeholder=True).order_by('placeholder_type', 'slot_number')
        by_type = {}
        for c in placeholder_qs:
            ptype = c.placeholder_type or 'UNK'
            if ptype not in by_type:
                by_type[ptype] = {
                    'placeholder_type': ptype,
                    'title': _canonical_placeholder_title(ptype),
                    'slot_count': 0,
                    'sample_code': c.course_code,
                }
            by_type[ptype]['slot_count'] += 1
        placeholder_rows = [by_type[k] for k in sorted(by_type.keys())]
    
    context = {
        'courses': courses,
        'include_placeholders': include_placeholders,
        'placeholder_rows': placeholder_rows,
        'course_type_choices': Course.COURSE_TYPE_CHOICES,
        'page_title': 'Manage Courses'
    }
    return render(request, "hod_template/manage_course.html", context)


@login_required
def delete_placeholder_course_type(request, placeholder_type):
    """Delete all placeholder slots for a placeholder type (e.g., SDC)."""
    if not check_hod_permission(request.user):
        return redirect('/')

    ptype = (placeholder_type or '').strip().upper()
    qs = Course.objects.filter(is_placeholder=True, placeholder_type=ptype)
    if not qs.exists():
        messages.warning(request, f"No placeholder courses found for type {ptype}")
        return redirect(f"{reverse('manage_course')}?include_placeholders=1")

    try:
        deleted_count, _ = qs.delete()
        messages.success(request, f"Deleted placeholder type {ptype} ({deleted_count} records)")
    except Exception as e:
        messages.error(request, f"Could not delete placeholder type {ptype}: {str(e)}")

    return redirect(f"{reverse('manage_course')}?include_placeholders=1")


@login_required
def edit_course(request, course_code):
    """Edit course"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    course = get_object_or_404(Course, course_code=course_code)
    form = CourseForm(request.POST or None, request.FILES or None, instance=course)
    
    # Disable course_code field (primary key)
    form.fields['course_code'].widget.attrs['readonly'] = True
    
    context = {
        'form': form,
        'course': course,
        'page_title': 'Edit Course'
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Course updated successfully!")
                return redirect(reverse('manage_course'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please fill the form correctly")
    
    return render(request, 'hod_template/edit_course_template.html', context)


@login_required
def delete_course(request, course_code):
    """Delete course"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    course = get_object_or_404(Course, course_code=course_code)
    try:
        course.delete()
        messages.success(request, "Course deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_course'))


# =============================================================================
# COURSE ASSIGNMENT MANAGEMENT
# =============================================================================

@login_required
def manage_course_assignment(request):
    """List all course assignments"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    assignments = Course_Assignment.objects.select_related(
        'course', 'faculty', 'faculty__user', 'academic_year', 'semester'
    ).all()
    
    context = {
        'assignments': assignments,
        'page_title': 'Manage Course Assignments'
    }
    return render(request, 'hod_template/manage_course_allocation.html', context)


@login_required
def delete_course_assignment(request, assignment_id):
    """Delete course assignment"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    assignment = get_object_or_404(Course_Assignment, id=assignment_id)
    try:
        assignment.delete()
        messages.success(request, "Assignment deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_course_assignment'))


@login_required
@csrf_exempt
def api_delete_course_assignment(request):
    """API endpoint to delete course assignment (with automatic refresh)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    assignment_id = request.POST.get('assignment_id')
    
    try:
        assignment = get_object_or_404(Course_Assignment, id=assignment_id)
        course_code = assignment.course.course_code

        # Also remove legacy duplicate rows (batch FK missing, batch_label used)
        # so deleted assignments do not continue to appear in timetable wizard.
        if assignment.batch_id and assignment.batch:
            legacy_qs = Course_Assignment.objects.filter(
                course=assignment.course,
                academic_year=assignment.academic_year,
                semester=assignment.semester,
                batch__isnull=True,
                batch_label=assignment.batch.batch_name,
            )
            legacy_qs.delete()

        assignment.delete()
        return JsonResponse({
            'success': True,
            'message': f'Assignment for {course_code} deleted successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Could not delete: {str(e)}'
        }, status=400)


@login_required
@csrf_exempt
def api_assign_catalog_course_as_ioc(request):
    """
    API endpoint to quickly assign a catalog course as IOC for a specific semester
    without adding it to the universal course plan.
    
    POST params:
      - semester_id
      - academic_year_id
      - program_type (UG/PG)
      - branch (program code)
      - course_code (real catalog course)
      - faculty_id
      - credits (optional)
    
    Returns: { success, message, course_code, faculty_name }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        semester_id = request.POST.get('semester_id')
        academic_year_id = request.POST.get('academic_year_id')
        program_type = request.POST.get('program_type', 'UG')
        branch = request.POST.get('branch', '')
        course_code = request.POST.get('course_code', '').strip()
        faculty_id = request.POST.get('faculty_id')

        if not all([semester_id, academic_year_id, course_code, faculty_id]):
            return JsonResponse({'error': 'Missing required parameters'}, status=400)

        semester = get_object_or_404(Semester, id=semester_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        course = get_object_or_404(Course, course_code=course_code)
        faculty = get_object_or_404(Faculty_Profile, id=faculty_id)

        # Get the program and batches for this program/branch
        program = get_object_or_404(Program, code=branch, level=program_type)

        # Get all active batches for this program/year/semester
        batches = ProgramBatch.objects.filter(
            program=program,
            academic_year=academic_year,
            year_of_study=semester.year_of_study,
            is_active=True
        )

        if not batches.exists():
            return JsonResponse({
                'error': f'No active batches found for {program_type} {branch}'
            }, status=400)

        # Enforce one quick IOC course per semester/program/branch context.
        existing_quick_ioc = Course_Assignment.objects.filter(
            semester=semester,
            academic_year=academic_year,
            batch__program__code=branch,
            batch__program__level=program_type,
            batch__year_of_study=semester.year_of_study,
            special_note__icontains='Quick IOC',
            is_active=True,
        ).exclude(course__course_code=course_code).select_related('course').first()

        if existing_quick_ioc:
            return JsonResponse({
                'error': (
                    f'Only one IOC course is allowed for this semester. '
                    f'Already assigned: {existing_quick_ioc.course.course_code} '
                    f'- {existing_quick_ioc.course.title}'
                )
            }, status=400)

        # Resolve IOC/EEC placeholder slot (if configured in course plan) so faculty stays unified.
        ioc_placeholder_plan = RegulationCoursePlan.objects.filter(
            semester=semester.semester_number,
            branch=branch,
            program_type=program_type,
            course__is_placeholder=True,
        ).filter(
            Q(category__code__in=['IOC', 'EEC']) |
            Q(course__placeholder_type__in=['IOC', 'EEC'])
        ).select_related('course').order_by('course__slot_number', 'course__course_code').first()
        ioc_placeholder_course = ioc_placeholder_plan.course if ioc_placeholder_plan else None

        # Create/update Course_Assignment for this IOC across all batches.
        created_count = 0
        updated_count = 0
        for batch in batches:
            # Upsert quick IOC assignment for the selected real course.
            existing = Course_Assignment.objects.filter(
                course=course,
                batch=batch,
                semester=semester,
                academic_year=academic_year,
            ).first()

            if existing:
                changed = False
                if existing.faculty_id != faculty.id:
                    existing.faculty = faculty
                    changed = True
                if not existing.is_active:
                    existing.is_active = True
                    changed = True
                note = existing.special_note or ''
                if 'Quick IOC' not in note:
                    existing.special_note = 'Quick IOC Assignment (Semester-only)'
                    changed = True
                if changed:
                    existing.save(update_fields=['faculty', 'is_active', 'special_note', 'updated_at'])
                    updated_count += 1
            else:
                Course_Assignment.objects.create(
                    course=course,
                    faculty=faculty,
                    batch=batch,
                    semester=semester,
                    academic_year=academic_year,
                    is_active=True,
                    # Store a note that this is a quick IOC assignment
                    special_note='Quick IOC Assignment (Semester-only)'
                )
                created_count += 1

            # Keep IOC placeholder assignment in sync with quick IOC faculty.
            if ioc_placeholder_course:
                Course_Assignment.objects.update_or_create(
                    course=ioc_placeholder_course,
                    batch=batch,
                    semester=semester,
                    academic_year=academic_year,
                    defaults={
                        'faculty': faculty,
                        'batch_label': batch.batch_name,
                        'is_active': True,
                    }
                )

        return JsonResponse({
            'success': True,
            'message': f'IOC assigned/updated for {created_count + updated_count} batch(es)',
            'course_code': course_code,
            'faculty_name': faculty.user.full_name
        })

    except Course.DoesNotExist:
        return JsonResponse({'error': f'Course "{course_code}" not found'}, status=404)
    except Faculty_Profile.DoesNotExist:
        return JsonResponse({'error': 'Faculty not found'}, status=404)
    except Program.DoesNotExist:
        return JsonResponse({'error': f'Program {branch} ({program_type}) not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_get_quick_ioc_assignments(request):
    """
    GET: Retrieve all quick IOC assignments for a semester.
    Query params:
      - semester_id
    
    Returns: { assignments: [{id, course_code, course_title, faculty_name, faculty_id, credits}] }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    semester_id = request.GET.get('semester_id')
    academic_year_id = request.GET.get('academic_year_id')
    program_type = request.GET.get('program_type')
    branch = request.GET.get('branch')
    batch_name = request.GET.get('batch')
    if not semester_id:
        return JsonResponse({'assignments': []})
    
    try:
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Find assignments marked as Quick IOC.
        assignments = Course_Assignment.objects.filter(
            semester=semester,
            special_note__contains='Quick IOC',
            is_active=True
        ).select_related('course', 'faculty', 'faculty__user', 'batch', 'batch__program')

        # Scope to the active page filters so output matches the selected context.
        if academic_year_id:
            assignments = assignments.filter(academic_year_id=academic_year_id)
        if program_type:
            assignments = assignments.filter(batch__program__level=program_type)
        if branch:
            assignments = assignments.filter(batch__program__code=branch)
        if batch_name:
            assignments = assignments.filter(batch__batch_name=batch_name)

        # Backend-agnostic dedupe (SQLite/MySQL/PostgreSQL): one row per course.
        deduped_assignments = []
        seen_course_ids = set()
        for assignment in assignments.order_by('course__course_code', '-updated_at', '-id'):
            if assignment.course_id in seen_course_ids:
                continue
            seen_course_ids.add(assignment.course_id)
            deduped_assignments.append(assignment)
        
        result = []
        for assignment in deduped_assignments:
            result.append({
                'id': assignment.id,
                'course_code': assignment.course.course_code,
                'course_title': assignment.course.title,
                'faculty_name': assignment.faculty.user.full_name if assignment.faculty else 'N/A',
                'faculty_id': assignment.faculty_id,
                'credits': assignment.course.credits
            })
        
        return JsonResponse({'assignments': result})
    except Exception as e:
        return JsonResponse({'assignments': []})


@login_required
@csrf_exempt
def api_delete_quick_ioc_assignment(request):
    """
    POST: Delete a quick IOC assignment.
    Body: { assignment_id }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        assignment_id = request.POST.get('assignment_id')
        assignment = get_object_or_404(Course_Assignment, id=assignment_id)
        
        # Check if this is a quick IOC assignment
        if 'Quick IOC' not in (assignment.special_note or ''):
            return JsonResponse({'error': 'Not a quick IOC assignment'}, status=400)
        
        # Soft delete or actually delete
        assignment.is_active = False
        assignment.save()
        
        return JsonResponse({'success': True, 'message': 'IOC assignment removed'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def update_placeholder_title(request):
    """Update the title (and optionally code) of an AC/IOC placeholder course."""
    if not check_hod_permission(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        data = request.POST

    course_code = data.get('course_code', '').strip()
    new_title = data.get('new_title', '').strip()

    if not course_code or not new_title:
        return JsonResponse({'success': False, 'error': 'course_code and new_title are required'})

    try:
        course = Course.objects.get(course_code=course_code)
        if not course.is_placeholder or course.placeholder_type not in ('AC', 'IOC'):
            return JsonResponse({'success': False, 'error': 'Only AC/IOC placeholder titles can be updated here'})
        course.title = new_title
        course.save()
        return JsonResponse({'success': True, 'new_title': course.title})
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'Course {course_code} not found'})


# =============================================================================
# ACADEMIC YEAR & SEMESTER MANAGEMENT
# =============================================================================

@login_required
def add_academic_year(request):
    """Add academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = AcademicYearForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Academic Year'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Academic Year added successfully!")
                return redirect(reverse('manage_academic_year'))
            except Exception as e:
                messages.error(request, f"Could not add: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, "hod_template/add_session_template.html", context)


@login_required
def manage_academic_year(request):
    """Manage academic years with their semesters - unified view"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    academic_years = AcademicYear.objects.prefetch_related('semesters').all()
    semester_form = SemesterForm()

    # IDs of semesters that have at least one program-specific date override
    semesters_with_overrides = set(
        ProgramSemesterDate.objects.values_list('semester_id', flat=True).distinct()
    )

    context = {
        'academic_years': academic_years,
        'semester_form': semester_form,
        'semesters_with_overrides': semesters_with_overrides,
        'page_title': 'Manage Academic Years'
    }
    return render(request, "hod_template/manage_session.html", context)


@login_required
def edit_academic_year(request, year_id):
    """Edit academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    year = get_object_or_404(AcademicYear, id=year_id)
    form = AcademicYearForm(request.POST or None, instance=year)
    context = {'form': form, 'year': year, 'page_title': 'Edit Academic Year'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Academic year updated successfully!")
                return redirect(reverse('manage_academic_year'))
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            messages.error(request, "Please fix the errors")
    
    return render(request, "hod_template/add_session_template.html", context)


@login_required
def delete_academic_year(request, year_id):
    """Delete academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    year = get_object_or_404(AcademicYear, id=year_id)
    try:
        year.delete()
        messages.success(request, "Academic year deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_academic_year'))


# =============================================================================
# SEMESTER MANAGEMENT
# =============================================================================

@login_required
def add_semester(request, year_id=None):
    """Add multiple semesters at once for an academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    academic_year = None
    
    # If year_id provided, pre-select that academic year
    if year_id:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
    
    # Get existing semesters for this academic year to show which are already added
    existing_semesters = []
    if academic_year:
        existing_semesters = list(Semester.objects.filter(academic_year=academic_year).values_list('semester_number', flat=True))
    
    context = {
        'academic_year': academic_year,
        'academic_years': AcademicYear.objects.all().order_by('-year'),
        'existing_semesters': existing_semesters,
        'page_title': f'Add Semesters for {academic_year.year}' if academic_year else 'Add Semesters'
    }
    
    if request.method == 'POST':
        academic_year_id = request.POST.get('academic_year') or (academic_year.id if academic_year else None)
        
        if not academic_year_id:
            messages.error(request, "Please select an academic year")
            return render(request, "hod_template/add_semester.html", context)
        
        academic_year_obj = get_object_or_404(AcademicYear, id=academic_year_id)
        
        # Get all semester data from POST
        semester_numbers = request.POST.getlist('semester_number[]')
        start_dates = request.POST.getlist('start_date[]')
        end_dates = request.POST.getlist('end_date[]')
        
        created_count = 0
        errors = []
        
        for i in range(len(semester_numbers)):
            sem_num = semester_numbers[i]
            start_date = start_dates[i] if i < len(start_dates) else ''
            end_date = end_dates[i] if i < len(end_dates) else ''
            
            if sem_num and start_date and end_date:
                try:
                    # Check if semester already exists
                    if Semester.objects.filter(academic_year=academic_year_obj, semester_number=sem_num).exists():
                        errors.append(f"Semester {sem_num} already exists for {academic_year_obj.year}")
                        continue
                    
                    Semester.objects.create(
                        academic_year=academic_year_obj,
                        semester_number=int(sem_num),
                        start_date=start_date,
                        end_date=end_date
                    )
                    created_count += 1
                except Exception as e:
                    errors.append(f"Error creating Semester {sem_num}: {str(e)}")
        
        if created_count > 0:
            messages.success(request, f"Successfully created {created_count} semester(s)!")
        
        for error in errors:
            messages.warning(request, error)
        
        if created_count > 0:
            return redirect(reverse('manage_academic_year'))
    
    return render(request, "hod_template/add_semester.html", context)


@login_required
def manage_semester(request):
    """Manage semesters"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    semesters = Semester.objects.select_related('academic_year').all()
    context = {'semesters': semesters, 'page_title': 'Manage Semesters'}
    return render(request, "hod_template/manage_semester.html", context)


@login_required
def delete_semester(request, semester_id):
    """Delete semester"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    semester = get_object_or_404(Semester, id=semester_id)
    try:
        semester.delete()
        messages.success(request, "Semester deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_academic_year'))


@login_required
def edit_semester(request, semester_id):
    """Edit an existing semester, including per-program date overrides"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    semester = get_object_or_404(Semester, id=semester_id)
    academic_year = semester.academic_year
    
    if request.method == 'POST':
        semester_number = request.POST.get('semester_number')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        try:
            # Check if another semester with this number exists (excluding current)
            if Semester.objects.filter(
                academic_year=academic_year, 
                semester_number=semester_number
            ).exclude(id=semester_id).exists():
                messages.error(request, f"Semester {semester_number} already exists for {academic_year.year}")
            else:
                semester.semester_number = int(semester_number)
                semester.start_date = start_date
                semester.end_date = end_date
                semester.save()

                # Save per-program date overrides
                for program in Program.objects.all():
                    prog_start = request.POST.get(f'prog_start_{program.id}', '').strip()
                    prog_end = request.POST.get(f'prog_end_{program.id}', '').strip()
                    if prog_start or prog_end:
                        ProgramSemesterDate.objects.update_or_create(
                            semester=semester, program=program,
                            defaults={
                                'start_date': prog_start or None,
                                'end_date': prog_end or None,
                            }
                        )
                    else:
                        # No override — delete any existing override for this program
                        ProgramSemesterDate.objects.filter(semester=semester, program=program).delete()

                messages.success(request, f"Semester {semester_number} updated successfully!")
                return redirect(reverse('manage_academic_year'))
        except Exception as e:
            messages.error(request, f"Error updating semester: {str(e)}")
    
    # Get existing semester numbers for this year (excluding current)
    existing_semesters = list(
        Semester.objects.filter(academic_year=academic_year)
        .exclude(id=semester_id)
        .values_list('semester_number', flat=True)
    )

    # Build per-program override dict {program.id: {'start': 'YYYY-MM-DD', 'end': 'YYYY-MM-DD'}}
    program_overrides = {
        psd.program_id: {
            'start': psd.start_date.isoformat() if psd.start_date else '',
            'end': psd.end_date.isoformat() if psd.end_date else '',
        }
        for psd in ProgramSemesterDate.objects.filter(semester=semester).select_related('program')
    }

    programs_with_overrides = [
        {
            'program': p,
            'start': program_overrides.get(p.id, {}).get('start', ''),
            'end': program_overrides.get(p.id, {}).get('end', ''),
        }
        for p in Program.objects.all().order_by('level', 'name')
    ]
    
    context = {
        'semester': semester,
        'academic_year': academic_year,
        'existing_semesters': existing_semesters,
        'programs_with_overrides': programs_with_overrides,
        'page_title': f'Edit Semester {semester.semester_number} for {academic_year.year}'
    }
    return render(request, "hod_template/edit_semester.html", context)


@login_required
def add_regulation(request):
    """Add regulation with course categories"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = RegulationForm(request.POST or None)
    course_category_choices = CourseCategory.CATEGORY_CHOICES
    
    # Default: select all categories for new regulation
    all_category_codes = [code for code, label in course_category_choices]
    
    if request.method == 'POST':
        selected_categories = request.POST.getlist('course_categories', [])
        selected_program_ids = request.POST.getlist('applicable_programs', [])
    else:
        # Pre-select all categories by default
        selected_categories = all_category_codes
        selected_program_ids = []

    all_programs = Program.objects.all().order_by('level', 'code')
    
    context = {
        'form': form, 
        'page_title': 'Add Regulation',
        'course_category_choices': course_category_choices,
        'selected_categories': selected_categories,
        'all_programs': all_programs,
        'selected_program_ids': selected_program_ids,
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    regulation = form.save()

                    # Save selected predefined course categories
                    for cat_code in selected_categories:
                        CourseCategory.objects.create(
                            regulation=regulation,
                            code=cat_code,
                            is_active=True
                        )

                    # Save custom categories
                    custom_codes = request.POST.getlist('custom_cat_codes', [])
                    custom_descs = request.POST.getlist('custom_cat_descs', [])
                    for code, desc in zip(custom_codes, custom_descs):
                        if code and desc:
                            CourseCategory.objects.create(
                                regulation=regulation,
                                code=code.upper(),
                                description=desc,
                                is_active=True
                            )

                    # Map regulation only to selected programs.
                    # This applies to new incoming batches from regulation.year onward.
                    selected_programs = Program.objects.filter(id__in=selected_program_ids)
                    for program in selected_programs:
                        # Close any currently open mapping before activating the new regulation.
                        ProgramRegulation.objects.filter(
                            program=program,
                            is_active=True,
                            effective_to_year__isnull=True,
                        ).exclude(regulation=regulation).update(
                            effective_to_year=regulation.year - 1
                        )

                        ProgramRegulation.objects.update_or_create(
                            program=program,
                            regulation=regulation,
                            defaults={
                                'effective_from_year': regulation.year,
                                'effective_to_year': None,
                                'is_active': True,
                                'notes': 'Configured while creating regulation.',
                            }
                        )
                
                messages.success(request, "Regulation added successfully!")
                return redirect(reverse('manage_regulation'))
            except Exception as e:
                messages.error(request, f"Could not add: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, "hod_template/add_regulation.html", context)


@login_required
def manage_regulation(request):
    """Manage regulations"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    regulations = Regulation.objects.prefetch_related('course_categories').all()
    context = {'regulations': regulations, 'page_title': 'Manage Regulations'}
    return render(request, "hod_template/manage_regulation.html", context)


@login_required
def edit_regulation(request, regulation_id):
    """Edit regulation with course categories"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    form = RegulationForm(request.POST or None, instance=regulation)
    course_category_choices = CourseCategory.CATEGORY_CHOICES
    predefined_codes = [code for code, label in course_category_choices]
    
    # Get currently selected predefined categories
    existing_predefined = list(
        regulation.course_categories.filter(code__in=predefined_codes, is_active=True).values_list('code', flat=True)
    )
    # Get existing custom categories (not in predefined list)
    existing_custom_categories = regulation.course_categories.exclude(code__in=predefined_codes)
    
    if request.method == 'POST':
        selected_categories = request.POST.getlist('course_categories', [])
    else:
        selected_categories = existing_predefined
    
    context = {
        'form': form, 
        'page_title': 'Edit Regulation',
        'regulation': regulation,
        'course_category_choices': course_category_choices,
        'selected_categories': selected_categories,
        'existing_custom_categories': existing_custom_categories,
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    selected_set = set(request.POST.getlist('course_categories', []))

                    # Never delete predefined categories because course plans keep FK links to them.
                    # Toggle active/inactive instead so historical mappings remain intact.
                    predefined_desc_map = dict(CourseCategory.CATEGORY_CHOICES)
                    for cat_code in predefined_codes:
                        defaults = {
                            'description': predefined_desc_map.get(cat_code, cat_code),
                            'is_active': cat_code in selected_set,
                        }
                        CourseCategory.objects.update_or_create(
                            regulation=regulation,
                            code=cat_code,
                            defaults=defaults,
                        )

                    # Handle existing custom categories: keep checked active, mark unchecked inactive.
                    kept_custom_ids = set(request.POST.getlist('existing_custom_cats', []))
                    custom_qs = regulation.course_categories.exclude(code__in=predefined_codes)
                    for custom in custom_qs:
                        should_be_active = str(custom.id) in kept_custom_ids
                        if custom.is_active != should_be_active:
                            custom.is_active = should_be_active
                            custom.save(update_fields=['is_active'])

                    # Add/update new custom categories as active.
                    custom_codes = request.POST.getlist('custom_cat_codes', [])
                    custom_descs = request.POST.getlist('custom_cat_descs', [])
                    for code, desc in zip(custom_codes, custom_descs):
                        code = (code or '').strip().upper()
                        desc = (desc or '').strip()
                        if code and desc:
                            CourseCategory.objects.update_or_create(
                                regulation=regulation,
                                code=code,
                                defaults={'description': desc, 'is_active': True}
                            )

                messages.success(request, "Regulation updated successfully!")
                return redirect(reverse('manage_regulation'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, "hod_template/edit_regulation.html", context)


@login_required
def delete_regulation(request, regulation_id):
    """Delete regulation"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    try:
        regulation = get_object_or_404(Regulation, id=regulation_id)
        regulation.delete()
        messages.success(request, "Regulation deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_regulation'))


# =============================================================================
# REGULATION COURSE PLAN MANAGEMENT
# =============================================================================

@login_required
def manage_regulation_courses(request, regulation_id):
    """Manage course plan for a specific regulation"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    # Get all course plans for this regulation, grouped by semester and branch
    course_plans = RegulationCoursePlan.objects.filter(
        regulation=regulation
    ).select_related('course', 'category').order_by('semester', 'branch', 'course__course_code')
    
    # Group by semester, then by branch
    plans_by_semester = {}
    for plan in course_plans:
        if plan.semester not in plans_by_semester:
            plans_by_semester[plan.semester] = {}
        if plan.branch not in plans_by_semester[plan.semester]:
            plans_by_semester[plan.semester][plan.branch] = []
        plans_by_semester[plan.semester][plan.branch].append(plan)
    
    # Get all available courses (universal, not tied to regulation)
    available_courses = Course.objects.all().order_by('course_code')
    
    # Get programs from database, grouped by level
    all_programs = Program.objects.all().order_by('level', 'code')
    program_levels = Program.PROGRAM_LEVEL_CHOICES
    
    active_categories = regulation.course_categories.filter(is_active=True).order_by('code')

    context = {
        'page_title': f'Course Plan - {regulation}',
        'regulation': regulation,
        'active_categories': active_categories,
        'placeholder_type_choices': Course.PLACEHOLDER_TYPE_CHOICES,
        'plans_by_semester': dict(sorted(plans_by_semester.items())),
        'available_courses': available_courses,
        'all_programs': all_programs,
        'program_levels': program_levels,
        'semesters': range(1, 13),  # Support up to 12 semesters (max allowed by Program model)
    }
    return render(request, 'hod_template/manage_regulation_courses.html', context)


@login_required
def add_regulation_course(request, regulation_id):
    """Add a course to regulation course plan"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        category_id = request.POST.get('category')
        semester = request.POST.get('semester')
        branch = request.POST.get('branch')
        program_type = request.POST.get('program_type', 'UG')
        is_elective = request.POST.get('is_elective') == 'on'
        
        try:
            course = Course.objects.get(course_code=course_code)
            category = CourseCategory.objects.get(id=category_id) if category_id else None
            
            # Check if already exists
            if RegulationCoursePlan.objects.filter(
                regulation=regulation,
                course=course,
                branch=branch,
                program_type=program_type
            ).exists():
                messages.warning(request, f"Course {course_code} already exists in the plan for {branch} {program_type}")
            else:
                RegulationCoursePlan.objects.create(
                    regulation=regulation,
                    course=course,
                    category=category,
                    semester=int(semester),
                    branch=branch,
                    program_type=program_type,
                    is_elective=is_elective
                )
                messages.success(request, f"Course {course_code} added to Semester {semester} for {branch}")
        except Course.DoesNotExist:
            messages.error(request, f"Course {course_code} not found")
        except Exception as e:
            messages.error(request, f"Error adding course: {str(e)}")
    
    return redirect('manage_regulation_courses', regulation_id=regulation_id)


@login_required
def remove_regulation_course(request, plan_id):
    """Remove a course from regulation course plan"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    plan = get_object_or_404(RegulationCoursePlan, id=plan_id)
    regulation_id = plan.regulation.id
    course_code = plan.course.course_code
    
    try:
        plan.delete()
        messages.success(request, f"Course {course_code} removed from course plan")
    except Exception as e:
        messages.error(request, f"Error removing course: {str(e)}")
    
    return redirect('manage_regulation_courses', regulation_id=regulation_id)


@login_required
def bulk_add_regulation_courses(request, regulation_id):
    """Bulk add courses to a semester in regulation course plan"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    if request.method == 'POST':
        semester = int(request.POST.get('semester'))
        branch = request.POST.get('branch')
        program_type = request.POST.get('program_type', 'UG')
        
        # Get all courses for this regulation that are defined for this semester
        courses = Course.objects.filter(regulation=regulation, semester=semester, branch=branch)
        
        added_count = 0
        for course in courses:
            if not RegulationCoursePlan.objects.filter(
                regulation=regulation,
                course=course,
                branch=branch,
                program_type=program_type
            ).exists():
                RegulationCoursePlan.objects.create(
                    regulation=regulation,
                    course=course,
                    semester=semester,
                    branch=branch,
                    program_type=program_type,
                    is_elective=False
                )
                added_count += 1
        
        if added_count > 0:
            messages.success(request, f"Added {added_count} courses to Semester {semester} for {branch}")
        else:
            messages.info(request, f"No new courses to add for Semester {semester} {branch}")
    
    return redirect('manage_regulation_courses', regulation_id=regulation_id)


@login_required
def api_get_programs_by_level(request):
    """API endpoint to get programs filtered by level (UG/PG/PHD)"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    level = request.GET.get('level', '').strip().upper()
    regulation_id = request.GET.get('regulation_id', '').strip()
    
    programs = Program.objects.all()
    
    if level:
        programs = programs.filter(level=level)
    
    if regulation_id:
        # Prefer the normalized ProgramRegulation mapping table.
        # Keep deprecated Program.regulation as fallback for legacy records.
        programs = programs.filter(
            Q(regulation_mappings__regulation_id=regulation_id, regulation_mappings__is_active=True)
            | Q(regulation_id=regulation_id)
        ).distinct()
    
    programs = programs.order_by('code')
    
    data = [{
        'code': p.code,
        'name': p.name,
        'full_name': p.full_name,
        'level': p.level,
        'degree': p.degree,
        'specialization': p.specialization or '',
        'total_semesters': p.total_semesters
    } for p in programs]
    
    return JsonResponse({'programs': data})


@login_required
def api_get_semester_courses(request, regulation_id):
    """API endpoint to get courses for a specific semester in a regulation"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    semester = request.GET.get('semester')
    program_type = request.GET.get('program_type', 'UG')
    branch = request.GET.get('branch', '')
    
    plans = RegulationCoursePlan.objects.filter(
        regulation=regulation,
        semester=semester,
        program_type=program_type,
        branch=branch
    ).select_related('course', 'category', 'elective_vertical')

    # Build semester-ordered numbering for placeholders across this regulation/program/branch.
    all_placeholder_plans = RegulationCoursePlan.objects.filter(
        regulation=regulation,
        program_type=program_type,
        branch=branch,
        course__is_placeholder=True,
    ).select_related('course').order_by('semester', 'id')

    normalized_index_by_plan_id = {}
    per_type_counts = {}
    for rp in all_placeholder_plans:
        ptype = rp.course.placeholder_type or 'UNK'
        per_type_counts[ptype] = per_type_counts.get(ptype, 0) + 1
        normalized_index_by_plan_id[rp.id] = per_type_counts[ptype]

    placeholder_types = {
        p.course.placeholder_type
        for p in plans
        if p.course and p.course.is_placeholder and p.course.placeholder_type
    }
    placeholder_title_map = {
        ptype: _canonical_placeholder_title(ptype)
        for ptype in placeholder_types
    }

    roman_numerals = {
        1: 'I', 2: 'II', 3: 'III', 4: 'IV', 5: 'V', 6: 'VI', 7: 'VII', 8: 'VIII', 9: 'IX', 10: 'X',
        11: 'XI', 12: 'XII', 13: 'XIII', 14: 'XIV', 15: 'XV', 16: 'XVI', 17: 'XVII', 18: 'XVIII', 19: 'XIX', 20: 'XX'
    }

    def _placeholder_display_title(course, normalized_slot=None):
        base = placeholder_title_map.get(course.placeholder_type, course.title)
        slot_to_use = normalized_slot or course.slot_number
        if not slot_to_use:
            return base
        suffix = roman_numerals.get(slot_to_use, str(slot_to_use))
        return f"{base} - {suffix}"

    data = []
    for p in plans:
        normalized_slot = normalized_index_by_plan_id.get(p.id)
        if p.course.is_placeholder and p.course.placeholder_type:
            display_code = f"{p.course.placeholder_type}-{normalized_slot:02d}" if normalized_slot else p.course.course_code
            display_title = _placeholder_display_title(p.course, normalized_slot)
        else:
            display_code = p.course.course_code
            display_title = p.course.title

        data.append({
            'plan_id': p.id,
            'course_code': p.course.course_code,
            'display_code': display_code,
            'title': display_title,
            'credits': p.course.credits,
            'ltp': p.course.ltp_display if not p.course.is_placeholder else '-',
            'category': p.category.code if p.category else None,
            'is_elective': p.is_elective,
            'elective_vertical': p.elective_vertical.name if p.elective_vertical else None,
            'elective_vertical_id': p.elective_vertical.id if p.elective_vertical else None,
            'is_placeholder': p.course.is_placeholder,
            'placeholder_type': p.course.placeholder_type if p.course.is_placeholder else None,
            'normalized_slot': normalized_slot if p.course.is_placeholder else None,
        })
    
    return JsonResponse({'courses': data})


@login_required
@csrf_exempt
def api_add_regulation_course(request, regulation_id):
    """API endpoint to add a course to regulation plan"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    course_code = request.POST.get('course_code', '').strip()
    semester = request.POST.get('semester')
    category_id = request.POST.get('category')
    program_type = request.POST.get('program_type', 'UG')
    branch = request.POST.get('branch', '')
    is_elective = request.POST.get('is_elective') == '1'
    elective_vertical_id = request.POST.get('elective_vertical', '').strip() or None
    auto_placeholder = request.POST.get('auto_placeholder') == '1'  # New flag for auto-assign
    try:
        credits = Decimal(str(request.POST.get('credits', 3)).strip())
    except (TypeError, ValueError, InvalidOperation):
        return JsonResponse({'success': False, 'error': 'Invalid credits value'})

    if credits < 0 or credits > 20:
        return JsonResponse({'success': False, 'error': 'Credits must be between 0 and 20'})
    placeholder_type = (request.POST.get('placeholder_type') or '').strip().upper()
    placeholder_selection = (request.POST.get('placeholder_selection') or '').strip()

    def _parse_int_field(name, default=0):
        raw = (request.POST.get(name, '') or '').strip()
        if raw == '':
            return default
        try:
            value = int(raw)
        except (TypeError, ValueError):
            raise ValueError(f'Invalid value for {name}')
        if value < 0:
            raise ValueError(f'{name} cannot be negative')
        return value

    try:
        lecture_hours = _parse_int_field('lecture_hours', 0)
        tutorial_hours = _parse_int_field('tutorial_hours', 0)
        practical_hours = _parse_int_field('practical_hours', 0)
    except ValueError as ex:
        return JsonResponse({'success': False, 'error': str(ex)})
    
    # All placeholder-capable category codes (including UC which maps to AC placeholder)
    elective_categories = ['PEC', 'OEC', 'ETC', 'SDC', 'SLC', 'IOC', 'EEC', 'AC', 'NCC', 'HON', 'MIN', 'UC']

    # Map category code → placeholder_type (when they differ)
    CATEGORY_TO_PLACEHOLDER_TYPE = {
        'UC': 'AC',   # University Course category → Audit Course placeholder
        'EEC': 'IOC',  # Empty EEC slot behaves like an IOC placeholder
    }

    try:
        category = CourseCategory.objects.get(id=category_id) if category_id else None

        def _allocate_safe_placeholder_course(placeholder_type, requested_credits, req_l, req_t, req_p):
            # Allocate next slot strictly from this regulation+program+branch context.
            existing_count = RegulationCoursePlan.objects.filter(
                regulation=regulation,
                program_type=program_type,
                branch=branch,
                course__is_placeholder=True,
                course__placeholder_type=placeholder_type
            ).count()
            next_slot = existing_count + 1

            return Course.get_or_create_placeholder(
                placeholder_type,
                next_slot,
                credits=requested_credits,
                lecture_hours=req_l,
                tutorial_hours=req_t,
                practical_hours=req_p,
                update_existing=False,
            )

        valid_placeholder_types = {code for code, _ in Course.PLACEHOLDER_TYPE_CHOICES}
        explicit_placeholder_request = bool(placeholder_type and placeholder_selection)

        if explicit_placeholder_request and placeholder_type not in valid_placeholder_types:
            return JsonResponse({'success': False, 'error': 'Invalid placeholder type'})
        
        # Auto-detect elective based on category
        # PEC/OEC/ETC/HON/MIN/AC/UC: always elective (PEC=choice-based, OEC/AC/UC=blocked slot).
        # SDC/SLC/NCC: elective only when adding a placeholder slot.
        # IOC/EEC placeholder slots: NEVER elective.
        always_elective = ['PEC', 'OEC', 'ETC', 'HON', 'MIN', 'AC', 'UC']
        mixed_elective = ['SDC', 'SLC', 'NCC']
        never_elective_placeholder = ['IOC', 'EEC']
        if category and category.code in always_elective:
            is_elective = True
        elif category and category.code in mixed_elective and auto_placeholder:
            is_elective = True
        elif category and category.code in never_elective_placeholder:
            is_elective = False
        
        # New explicit placeholder flow from UI: choose existing placeholder or create a new one.
        if explicit_placeholder_request:
            if placeholder_selection == '__NEW__':
                source_course = None
                if course_code:
                    try:
                        source_course = Course.objects.get(course_code=course_code)
                    except Course.DoesNotExist:
                        return JsonResponse({'success': False, 'error': f'Course "{course_code}" not found'})

                requested_l = source_course.lecture_hours if source_course else None
                requested_t = source_course.tutorial_hours if source_course else None
                requested_p = source_course.practical_hours if source_course else None

                course, created = _allocate_safe_placeholder_course(
                    placeholder_type,
                    credits,
                    requested_l,
                    requested_t,
                    requested_p,
                )
                if not course:
                    return JsonResponse({'success': False, 'error': f'Could not create placeholder for {placeholder_type}'})
            else:
                try:
                    course = Course.objects.get(
                        course_code=placeholder_selection,
                        is_placeholder=True,
                        placeholder_type=placeholder_type,
                    )
                except Course.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Selected placeholder course not found'})

        # Backward-compatible IOC/EEC handling for old clients.
        # Only used when course is intentionally left empty.
        elif category and category.code in ['IOC', 'EEC'] and not course_code:
            placeholder_type = CATEGORY_TO_PLACEHOLDER_TYPE.get(category.code, category.code)

            source_course = None
            if course_code:
                try:
                    source_course = Course.objects.get(course_code=course_code)
                except Course.DoesNotExist:
                    return JsonResponse({'success': False, 'error': f'Course "{course_code}" not found'})

            # If source selected, use its LTP as template; credits still come from plan input.
            if source_course:
                lecture_hours = source_course.lecture_hours
                tutorial_hours = source_course.tutorial_hours
                practical_hours = source_course.practical_hours
            else:
                lecture_hours = None
                tutorial_hours = None
                practical_hours = None

            course, created = _allocate_safe_placeholder_course(
                placeholder_type,
                credits,
                lecture_hours,
                tutorial_hours,
                practical_hours,
            )
            if not course:
                return JsonResponse({
                    'success': False,
                    'error': f'Could not create placeholder for {category.code} slot'
                })

        # Auto-assign placeholder course if requested (non IOC/EEC)
        elif auto_placeholder and category and category.code in elective_categories:
            # Resolve the actual placeholder_type (UC category → AC placeholder)
            placeholder_type = CATEGORY_TO_PLACEHOLDER_TYPE.get(category.code, category.code)

            # Find how many of this type are already in the regulation (across all semesters)
            existing_count = RegulationCoursePlan.objects.filter(
                regulation=regulation,
                program_type=program_type,
                branch=branch,
                course__is_placeholder=True,
                course__placeholder_type=placeholder_type
            ).count()
            
            next_slot = existing_count + 1
            
            # Get or create the placeholder course for this slot with user-specified credits
            course, created = Course.get_or_create_placeholder(
                placeholder_type,
                next_slot,
                credits=credits,
                lecture_hours=lecture_hours,
                tutorial_hours=tutorial_hours,
                practical_hours=practical_hours,
                update_existing=False,
            )
            if not course:
                return JsonResponse({
                    'success': False, 
                    'error': f'Could not create placeholder for {category.code} slot {next_slot}'
                })
        else:
            # Regular course lookup
            if not course_code:
                return JsonResponse({'success': False, 'error': 'Please select a course'})
            try:
                selected_course = Course.objects.get(course_code=course_code)
            except Course.DoesNotExist:
                return JsonResponse({'success': False, 'error': f'Course "{course_code}" not found'})

            # If a placeholder course is selected, allocate a regulation-safe next slot
            # for that placeholder type and apply the entered credits for this plan slot.
            if selected_course.is_placeholder and selected_course.placeholder_type:
                course, created = _allocate_safe_placeholder_course(
                    selected_course.placeholder_type,
                    credits,
                    selected_course.lecture_hours,
                    selected_course.tutorial_hours,
                    selected_course.practical_hours,
                )
                if not course:
                    return JsonResponse({'success': False, 'error': 'Could not prepare placeholder for this regulation'})
            else:
                course = selected_course
        
        # Get elective vertical object if provided
        elective_vertical = None
        if is_elective and elective_vertical_id:
            try:
                elective_vertical = ElectiveVertical.objects.get(id=elective_vertical_id, regulation=regulation)
            except ElectiveVertical.DoesNotExist:
                pass  # Vertical not found, leave as None
        
        # Check if this course already exists anywhere in this regulation+branch+program_type
        # (unique constraint is on regulation, course, branch, program_type — not per semester)
        existing_plan = RegulationCoursePlan.objects.filter(
            regulation=regulation,
            course=course,
            program_type=program_type,
            branch=branch
        ).first()
        
        if existing_plan:
            return JsonResponse({
                'success': False,
                'error': f'{course.course_code} - "{course.title}" is already assigned to Semester {existing_plan.semester} for {branch} ({program_type}). A course can only appear once per regulation.'
            })
        
        RegulationCoursePlan.objects.create(
            regulation=regulation,
            course=course,
            semester=semester,
            category=category,
            program_type=program_type,
            branch=branch,
            is_elective=is_elective,
            elective_vertical=elective_vertical
        )
        
        return JsonResponse({
            'success': True,
            'course_code': course.course_code,
            'course_title': course.title
        })
    except CourseCategory.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid category selected'})
    except IntegrityError:
        return JsonResponse({'success': False, 'error': f'{course_code} is already assigned in this regulation. A course can only appear once per branch/program.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
def api_remove_regulation_course(request):
    """API endpoint to remove a course from regulation plan.

    GET-like POST with check=1 → returns impact counts without deleting.
    Normal POST                → blocks deletion if attendance records exist.
    POST with force=1          → deletes plan AND all related assignments/entries.
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    plan_id = request.POST.get('plan_id')
    check_only = request.POST.get('check') == '1'
    force = request.POST.get('force') == '1'

    try:
        plan = RegulationCoursePlan.objects.select_related('course').get(id=plan_id)
        course = plan.course

        # Count live records linked to this course that would become strays
        assignment_qs = Course_Assignment.objects.filter(course=course)
        assignment_count = assignment_qs.count()
        attendance_count = Attendance.objects.filter(assignment__course=course).count()
        timetable_count = TimetableEntry.objects.filter(course=course).count()

        if check_only:
            return JsonResponse({
                'success': True,
                'course_code': course.course_code,
                'assignments': assignment_count,
                'attendance': attendance_count,
                'timetable_entries': timetable_count,
            })

        # Block deletion if attendance records exist and force is not set
        if attendance_count > 0 and not force:
            return JsonResponse({
                'success': False,
                'blocked': True,
                'error': (
                    f"{course.course_code} has {attendance_count} attendance record(s). "
                    f"Use force delete to remove everything."
                ),
                'assignments': assignment_count,
                'attendance': attendance_count,
                'timetable_entries': timetable_count,
            })

        with transaction.atomic():
            if force:
                # Clean up stray data explicitly before deleting the plan
                TimetableEntry.objects.filter(course=course).delete()
                assignment_qs.delete()  # also cascades Attendance via its FK
            plan.delete()  # cascades ElectiveCourseOffering → ElectiveOfferingFacultyAssignment

        return JsonResponse({'success': True})

    except RegulationCoursePlan.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Course plan not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_search_courses(request):
    """API endpoint to search courses for dropdown"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    query = request.GET.get('q', '').strip()
    limit = int(request.GET.get('limit', 20))
    include_placeholders = request.GET.get('include_placeholders', 'true').lower() == 'true'
    placeholders_only = request.GET.get('placeholders_only', 'false').lower() == 'true'
    placeholder_type = request.GET.get('placeholder_type', '')
    
    courses = Course.objects.all()
    
    # Filter by placeholder status
    if placeholders_only:
        courses = courses.filter(is_placeholder=True)
        if placeholder_type:
            courses = courses.filter(placeholder_type=placeholder_type)
    elif not include_placeholders:
        courses = courses.filter(is_placeholder=False)
    
    if query:
        courses = courses.filter(
            Q(course_code__icontains=query) | Q(title__icontains=query)
        )
    
    # Order placeholders by slot number, others by course code
    courses = courses.order_by('-is_placeholder', 'placeholder_type', 'slot_number', 'course_code')[:limit]

    placeholder_types = {c.placeholder_type for c in courses if c.is_placeholder and c.placeholder_type}
    placeholder_title_map = {
        ptype: _canonical_placeholder_title(ptype)
        for ptype in placeholder_types
    }
    
    data = [{
        'course_code': c.course_code,
        'title': c.title,
        'placeholder_base_title': placeholder_title_map.get(c.placeholder_type, c.title) if c.is_placeholder else c.title,
        'credits': c.credits,
        'course_type': c.course_type,
        'ltp': c.ltp_display,
        'is_placeholder': c.is_placeholder,
        'placeholder_type': c.placeholder_type,
        'slot_number': c.slot_number,
        'display': f"{c.course_code} - {c.title}" + (" (Placeholder)" if c.is_placeholder else f" ({c.ltp_display}, {c.credits} cr)")
    } for c in courses]
    
    return JsonResponse({'courses': data})


@login_required
def api_get_placeholder_courses(request):
    """API endpoint to get placeholder courses by type"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    placeholder_type = request.GET.get('type', '')
    
    courses = Course.objects.filter(is_placeholder=True)
    if placeholder_type:
        courses = courses.filter(placeholder_type=placeholder_type)
    
    courses = courses.order_by('placeholder_type', 'slot_number')
    
    data = [{
        'course_code': c.course_code,
        'title': c.title,
        'credits': c.credits,
        'placeholder_type': c.placeholder_type,
        'slot_number': c.slot_number,
    } for c in courses]
    
    return JsonResponse({'placeholders': data})


# =============================================================================
# ELECTIVE COURSE OFFERINGS APIs
# =============================================================================

@login_required
def api_get_elective_offerings(request):
    """API endpoint to get elective offerings for a regulation course plan"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    plan_id = request.GET.get('plan_id')
    semester_id = request.GET.get('semester_id')
    
    if not plan_id:
        return JsonResponse({'error': 'plan_id required'}, status=400)
    
    filters = {'regulation_course_plan_id': plan_id}
    if semester_id:
        filters['semester_id'] = semester_id
    
    offerings = ElectiveCourseOffering.objects.filter(
        **filters
    ).select_related('actual_course', 'elective_vertical').prefetch_related(
        'faculty_assignments__faculty__user',
        'faculty_assignments__lab_assistant__user'
    )
    
    data = []
    for o in offerings:
        # Get existing faculty assignments for this offering
        assignments = {}
        for fa in o.faculty_assignments.all():
            assignments[fa.batch_number] = {
                'faculty_id': fa.faculty_id,
                'faculty_name': fa.faculty.user.full_name if fa.faculty else None,
                'lab_assistant_id': fa.lab_assistant_id,
                'lab_assistant_name': fa.lab_assistant.user.full_name if fa.lab_assistant else None,
            }
        
        # Check if course needs lab assistant
        course = o.actual_course
        needs_lab = course.course_type in ['L', 'LIT'] or course.practical_hours > 0
        
        data.append({
            'id': o.id,
            'course_code': course.course_code,
            'course_title': course.title,
            'course_type': course.course_type,
            'batch_count': o.batch_count,
            'capacity_per_batch': o.capacity_per_batch,
            'vertical': o.elective_vertical.name if o.elective_vertical else None,
            'needs_lab_assistant': needs_lab,
            'assignments': assignments,
            'regulation_course_plan_id': o.regulation_course_plan_id,
        })
    
    return JsonResponse({'offerings': data})


@login_required
@csrf_exempt
def api_get_or_create_matrix_offerings(request):
    """Return existing elective offerings for given course codes, creating missing
    offerings under the first provided PEC plan id.

    Expects JSON POST: { 'course_codes': [], 'plan_ids': [], 'semester_id': 24 }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    import json as _json
    try:
        data = _json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    course_codes = data.get('course_codes') or []
    plan_ids = data.get('plan_ids') or []
    semester_id = data.get('semester_id')

    if not course_codes:
        return JsonResponse({'offerings': []})

    semester = None
    if semester_id:
        try:
            semester = Semester.objects.get(id=semester_id)
        except Semester.DoesNotExist:
            return JsonResponse({'error': 'Semester not found'}, status=404)

    # Resolve plans
    plans = RegulationCoursePlan.objects.filter(id__in=plan_ids).select_related('course') if plan_ids else RegulationCoursePlan.objects.none()
    plan_map = {p.id: p for p in plans}
    default_plan = plans.first() if plans.exists() else None

    offerings = []
    for code in course_codes:
        try:
            course = Course.objects.get(course_code=code)
        except Course.DoesNotExist:
            continue

        # Look for existing offering for this course across provided plans (and semester if provided)
        q = ElectiveCourseOffering.objects.filter(actual_course=course)
        if plan_ids:
            q = q.filter(regulation_course_plan_id__in=plan_ids)
        if semester:
            q = q.filter(semester=semester)

        existing = q.select_related('actual_course', 'elective_vertical', 'regulation_course_plan').first()
        if existing:
            offerings.append(existing)
            continue

        # Create under default_plan if available
        if default_plan:
            off = ElectiveCourseOffering.objects.create(
                regulation_course_plan=default_plan,
                semester=semester,
                actual_course=course,
                batch_count=1,
                capacity_per_batch=course.credits or 30,
                elective_vertical=default_plan.elective_vertical
            )
            offerings.append(off)

    # Serialize offerings similar to api_get_elective_offerings
    out = []
    for o in offerings:
        assignments = {}
        for fa in o.faculty_assignments.all():
            assignments[fa.batch_number] = {
                'faculty_id': fa.faculty_id,
                'faculty_name': fa.faculty.user.full_name if fa.faculty else None,
                'lab_assistant_id': fa.lab_assistant_id,
                'lab_assistant_name': fa.lab_assistant.user.full_name if fa.lab_assistant else None,
            }

        course = o.actual_course
        needs_lab = course.course_type in ['L', 'LIT'] or course.practical_hours > 0

        out.append({
            'id': o.id,
            'course_code': course.course_code,
            'course_title': course.title,
            'course_type': course.course_type,
            'batch_count': o.batch_count,
            'capacity_per_batch': o.capacity_per_batch,
            'vertical': o.elective_vertical.name if o.elective_vertical else None,
            'needs_lab_assistant': needs_lab,
            'assignments': assignments,
            'regulation_course_plan_id': o.regulation_course_plan_id,
        })

    return JsonResponse({'offerings': out})


@login_required
@csrf_exempt
def api_add_elective_offering(request):
    """API endpoint to add an elective course offering"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    plan_id = request.POST.get('plan_id')
    course_code = request.POST.get('course_code')
    batch_count = request.POST.get('batch_count', 1)
    capacity_per_batch = request.POST.get('capacity_per_batch', 30)
    faculty_id = request.POST.get('faculty_id', '').strip() or None
    semester_id = request.POST.get('semester_id')
    
    try:
        plan = RegulationCoursePlan.objects.get(id=plan_id)
        course = Course.objects.get(course_code=course_code)
        semester = Semester.objects.get(id=semester_id) if semester_id else None
        
        # Check if this course is already offered for this plan
        existing = ElectiveCourseOffering.objects.filter(
            regulation_course_plan=plan,
            actual_course=course,
            semester=semester
        ).exists()
        
        if existing:
            return JsonResponse({
                'success': False,
                'error': f'{course.course_code} is already offered for this slot'
            })
        
        ElectiveCourseOffering.objects.create(
            regulation_course_plan=plan,
            semester=semester,
            actual_course=course,
            batch_count=int(batch_count),
            capacity_per_batch=int(capacity_per_batch),
            elective_vertical=plan.elective_vertical
        )
        
        return JsonResponse({'success': True})
    except RegulationCoursePlan.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid plan ID'})
    except Course.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Course not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
def api_remove_elective_offering(request):
    """API endpoint to remove an elective course offering"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    offering_id = request.POST.get('offering_id')
    
    try:
        offering = ElectiveCourseOffering.objects.get(id=offering_id)
        offering.delete()
        return JsonResponse({'success': True})
    except ElectiveCourseOffering.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Offering not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
def api_save_elective_offering_assignment(request):
    """
    API endpoint to save faculty assignments for an elective course offering.
    Handles multiple batch assignments at once.
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    import json
    
    try:
        # Parse JSON body
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        offering_id = data.get('offering_id')
        assignments = data.get('assignments', [])  # [{batch_number, faculty_id, lab_assistant_id}, ...]
        
        if not offering_id:
            return JsonResponse({'success': False, 'error': 'offering_id required'})
        
        offering = ElectiveCourseOffering.objects.get(id=offering_id)
        
        # Import the model here to avoid circular imports
        from main_app.models import ElectiveOfferingFacultyAssignment, Course_Assignment, ProgramBatch

        # Process each assignment
        saved_count = 0
        saved_fas = []  # track saved ElectiveOfferingFacultyAssignment objects
        for a in assignments:
            batch_number = int(a.get('batch_number', 0))
            faculty_id = a.get('faculty_id')
            lab_assistant_id = a.get('lab_assistant_id')

            if batch_number < 1 or batch_number > offering.batch_count:
                continue

            if faculty_id:
                # Create or update offering-specific assignment
                fa, created = ElectiveOfferingFacultyAssignment.objects.update_or_create(
                    offering=offering,
                    batch_number=batch_number,
                    defaults={
                        'faculty_id': faculty_id,
                        'lab_assistant_id': lab_assistant_id if lab_assistant_id else None,
                        'is_active': True
                    }
                )
                saved_count += 1
                saved_fas.append(fa)
            else:
                # Remove offering assignment if no faculty specified
                ElectiveOfferingFacultyAssignment.objects.filter(
                    offering=offering,
                    batch_number=batch_number
                ).delete()

        # Also mirror offering assignments to Course_Assignment for timetable engine
        try:
            plan = offering.regulation_course_plan
            sem = offering.semester
            # Find program batches for this plan's branch and semester year
            program_batches = ProgramBatch.objects.filter(
                academic_year=sem.academic_year,
                program__code=plan.branch,
                year_of_study=sem.year_of_study,
                is_active=True
            ).order_by('batch_name')

            if program_batches.exists():
                batches = list(program_batches)
                nb = len(batches)
                for fa in saved_fas:
                    # Map offering batch_number -> program batch index (1-based)
                    idx = (fa.batch_number - 1) % nb
                    batch = batches[idx]

                    # Create or update Course_Assignment for this program batch
                    Course_Assignment.objects.update_or_create(
                        course=offering.actual_course,
                        batch=batch,
                        academic_year=sem.academic_year,
                        semester=sem,
                        defaults={
                            'faculty': fa.faculty,
                            'lab_assistant': fa.lab_assistant if getattr(fa, 'lab_assistant', None) else None,
                            'batch_label': batch.batch_name,
                            'is_active': True,
                        }
                    )
        except Exception:
            # Don't block saving assignments if mirroring fails; collect warning server-side
            pass

        return JsonResponse({
            'success': True,
            'message': f'Saved {saved_count} assignment(s)'
        })
    
    except ElectiveCourseOffering.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Offering not found'})
    except Faculty_Profile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Faculty not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# ELECTIVE VERTICAL MANAGEMENT APIs
# =============================================================================

@login_required
def api_get_elective_verticals(request, regulation_id):
    """API endpoint to get all elective verticals for a regulation"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    verticals = ElectiveVertical.objects.filter(
        regulation=regulation,
        is_active=True
    ).order_by('name')
    
    data = [{
        'id': v.id,
        'name': v.name,
        'description': v.description or '',
        'course_count': v.course_plans.count()
    } for v in verticals]
    
    return JsonResponse({'verticals': data})


@login_required
@csrf_exempt
def api_add_elective_vertical(request, regulation_id):
    """API endpoint to add a new elective vertical to a regulation"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    regulation = get_object_or_404(Regulation, id=regulation_id)
    
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip() or None
    
    if not name:
        return JsonResponse({'success': False, 'error': 'Vertical name is required'})
    
    # Check if already exists
    if ElectiveVertical.objects.filter(regulation=regulation, name__iexact=name).exists():
        return JsonResponse({'success': False, 'error': 'A vertical with this name already exists'})
    
    try:
        vertical = ElectiveVertical.objects.create(
            regulation=regulation,
            name=name,
            description=description
        )
        return JsonResponse({
            'success': True,
            'vertical': {
                'id': vertical.id,
                'name': vertical.name,
                'description': vertical.description or ''
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
def api_edit_elective_vertical(request, vertical_id):
    """API endpoint to edit an elective vertical"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    vertical = get_object_or_404(ElectiveVertical, id=vertical_id)
    
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip() or None
    
    if not name:
        return JsonResponse({'success': False, 'error': 'Vertical name is required'})
    
    # Check if name already exists (excluding current vertical)
    if ElectiveVertical.objects.filter(
        regulation=vertical.regulation, 
        name__iexact=name
    ).exclude(id=vertical_id).exists():
        return JsonResponse({'success': False, 'error': 'A vertical with this name already exists'})
    
    try:
        vertical.name = name
        vertical.description = description
        vertical.save()
        return JsonResponse({
            'success': True,
            'vertical': {
                'id': vertical.id,
                'name': vertical.name,
                'description': vertical.description or ''
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
def api_delete_elective_vertical(request, vertical_id):
    """API endpoint to delete an elective vertical"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    vertical = get_object_or_404(ElectiveVertical, id=vertical_id)
    
    # Check if any courses are using this vertical
    course_count = vertical.course_plans.count()
    if course_count > 0:
        return JsonResponse({
            'success': False, 
            'error': f'Cannot delete: {course_count} course(s) are assigned to this vertical'
        })
    
    try:
        vertical.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# SEMESTER COURSE ASSIGNMENT (for specific academic semester)
# =============================================================================

@login_required
def semester_course_assignment(request):
    """
    Assign courses for a specific semester with pre-fill from regulation course plan.
    Shows courses based on student's regulation.
    """
    if not check_hod_permission(request.user):
        return redirect('/')
    
    from django.utils import timezone
    
    # Get filter parameters first to determine program type
    selected_semester = request.GET.get('semester_id')
    selected_branch = request.GET.get('branch', '')
    selected_batch = request.GET.get('batch')
    selected_program = request.GET.get('program_type', 'UG')
    
    # Get current academic year
    current_academic_year = AcademicYear.get_current()
    
    # Get semesters filtered by:
    # 1. Only CURRENT or UPCOMING (not COMPLETED)
    # 2. If PG selected, only sem 1-4; if UG, sem 1-8
    today = timezone.now().date()
    max_semester = 4 if selected_program == 'PG' else 8
    
    if current_academic_year:
        # Get semesters that are current or upcoming (end_date >= today)
        semesters = Semester.objects.filter(
            academic_year=current_academic_year,
            end_date__gte=today,  # Not completed
            semester_number__lte=max_semester  # Respect program type
        ).order_by('semester_number')
    else:
        semesters = Semester.objects.none()
    
    # Get programs from database for dynamic branch selection
    all_programs = Program.objects.all().order_by('level', 'code')
    program_levels = Program.PROGRAM_LEVEL_CHOICES
    
    # Get branches filtered by selected program type
    filtered_programs = all_programs.filter(level=selected_program) if selected_program else all_programs
    
    # Ensure selected branch is valid for selected program level.
    # Example: when switching UG->PG, prior branch like "CSE" should not persist.
    if filtered_programs.exists():
        if not selected_branch or not filtered_programs.filter(code=selected_branch).exists():
            selected_branch = filtered_programs.first().code
    
    # Get batch choices - will be populated after semester is selected
    batch_choices = []
    
    # Get regulations for manual selection (prefer program-specific mappings)
    if selected_branch and selected_program:
        mapped_reg_ids = ProgramRegulation.objects.filter(
            program__code=selected_branch,
            program__level=selected_program,
            is_active=True,
        ).values_list('regulation_id', flat=True).distinct()
        all_regulations = Regulation.objects.filter(id__in=mapped_reg_ids).order_by('-year', 'name')
        if not all_regulations.exists():
            all_regulations = Regulation.objects.all().order_by('-year', 'name')
    else:
        all_regulations = Regulation.objects.all().order_by('-year', 'name')
    selected_regulation_id = request.GET.get('regulation_id', '')
    
    context = {
        'page_title': 'Semester Course Assignment',
        'semesters': semesters,
        'all_programs': all_programs,
        'program_levels': program_levels,
        'filtered_programs': filtered_programs,
        'batch_choices': batch_choices,
        'selected_semester': selected_semester,
        'selected_branch': selected_branch,
        'selected_batch': selected_batch,
        'selected_program': selected_program,
        'academic_year': current_academic_year,
        'course_plans': [],
        'existing_assignments': [],
        'faculty_list': Faculty_Profile.objects.select_related('user').filter(user__is_active=True),
        'all_regulations': all_regulations,
        'selected_regulation_id': selected_regulation_id,
        'show_quick_ioc': False,
    }
    
    if selected_semester and selected_branch:
        semester_obj = get_object_or_404(Semester, id=selected_semester)
        context['semester_obj'] = semester_obj

        selected_program_obj = Program.objects.filter(
            code=selected_branch,
            level=selected_program,
        ).first()
        
        # Calculate year of study from semester number
        import math
        year_of_study = math.ceil(semester_obj.semester_number / 2)
        
        # Get batch choices for this specific program/year from ProgramBatch config
        if current_academic_year:
            # Resolve batches from ProgramBatch only:
            # 1) Prefer explicit config for current AY + current year_of_study
            # 2) Otherwise derive cohort admission AY and read Year 1 config from that AY
            source_year = current_academic_year
            source_year_of_study = year_of_study

            program_batches = ProgramBatch.objects.filter(
                academic_year=source_year,
                program__code=selected_branch,
                program__level=selected_program,
                year_of_study=source_year_of_study,
                is_active=True
            ).order_by('batch_name')

            if not program_batches.exists() and year_of_study > 1:
                try:
                    start_year = int(current_academic_year.year.split('-')[0])
                    admission_start_year = start_year - (year_of_study - 1)
                    admission_year_label = f"{admission_start_year}-{str(admission_start_year + 1)[-2:]}"
                    source_year = AcademicYear.objects.filter(year=admission_year_label).first()
                    source_year_of_study = 1
                except (ValueError, IndexError):
                    source_year = None

                if source_year:
                    program_batches = ProgramBatch.objects.filter(
                        academic_year=source_year,
                        program__code=selected_branch,
                        program__level=selected_program,
                        year_of_study=source_year_of_study,
                        is_active=True
                    ).order_by('batch_name')

            batch_choices = [(b.batch_name, b.batch_display) for b in program_batches]
            context['batch_choices'] = batch_choices
        
        # Determine which regulation applies to students in this semester
        # Find students in this semester number with this branch
        students_in_sem = Student_Profile.objects.filter(
            current_sem=semester_obj.semester_number,
            branch=selected_branch,
            status='ACTIVE'
        )
        if selected_batch:
            students_in_sem = students_in_sem.filter(batch_label=selected_batch)
        if selected_program:
            students_in_sem = students_in_sem.filter(program_type=selected_program)
        
        # Get the most common regulation among these students
        regulation = None
        student_count = students_in_sem.count()
        
        if students_in_sem.exists():
            from django.db.models import Count
            reg_counts = students_in_sem.values('regulation').annotate(
                count=Count('regulation')
            ).order_by('-count')
            if reg_counts and reg_counts[0]['regulation']:
                regulation = Regulation.objects.get(id=reg_counts[0]['regulation'])
        
        # If no students or no regulation detected, auto-detect from program mapping.
        if not regulation:
            # Calculate admission year from semester and current academic year
            # year_of_study = ceil(semester / 2): sem 1-2 = year 1, sem 3-4 = year 2, etc.
            import math
            year_of_study = math.ceil(semester_obj.semester_number / 2)
            
            # Get start year from current academic year (e.g., "2024-25" -> 2024)
            try:
                start_year = int(current_academic_year.year.split('-')[0])
                admission_year = start_year - (year_of_study - 1)

                auto_regulation = ProgramRegulation.get_regulation_for_student(
                    selected_branch,
                    selected_program,
                    admission_year,
                )
                if auto_regulation:
                    regulation = auto_regulation
                    context['regulation_auto_detected'] = True
                    context['detected_admission_year'] = admission_year
            except (ValueError, IndexError):
                pass

        # Final safe fallback: use program's direct regulation if configured.
        if not regulation and selected_program_obj and selected_program_obj.regulation:
            regulation = selected_program_obj.regulation
            context['regulation_from_program_default'] = True
        
        # Allow manual override if explicitly selected
        if selected_regulation_id:
            try:
                regulation = Regulation.objects.get(id=selected_regulation_id)
                context['regulation_manual_override'] = True
            except Regulation.DoesNotExist:
                pass
        
        context['regulation'] = regulation
        context['student_count'] = student_count
        context['needs_regulation_selection'] = (not regulation)
        
        # Get course plan from regulation
        if regulation:
            course_plans = RegulationCoursePlan.objects.filter(
                regulation=regulation,
                semester=semester_obj.semester_number,
                branch=selected_branch,
                program_type=selected_program
            ).select_related('course', 'category', 'elective_vertical')
            
            # Separate core and elective courses for better display
            core_courses = [p for p in course_plans if not p.is_elective]
            elective_courses = [p for p in course_plans if p.is_elective]
            
            context['course_plans'] = course_plans
            context['core_courses'] = core_courses
            context['elective_courses'] = elective_courses
            
            # Get elective offerings for placeholder courses
            elective_offerings = ElectiveCourseOffering.objects.filter(
                regulation_course_plan__in=course_plans,
                semester=semester_obj
            ).select_related('actual_course', 'elective_vertical')
            
            # Create a map of plan_id to offerings
            elective_offerings_map = {}
            for offering in elective_offerings:
                plan_id = offering.regulation_course_plan_id
                if plan_id not in elective_offerings_map:
                    elective_offerings_map[plan_id] = []
                elective_offerings_map[plan_id].append(offering)
            context['elective_offerings_map'] = elective_offerings_map
            
            # Build PEC Course Arrangement Matrix data
            # Gather all PEC placeholder courses for this semester
            pec_slots = [
                p for p in elective_courses
                if p.course and p.course.is_placeholder
                and p.course.placeholder_type == 'PEC'
            ]
            if len(pec_slots) >= 1:
                import json as _json
                
                pec_plan_ids = [p.id for p in pec_slots]
                min_groups = max(len(pec_slots), 1)
                
                # Check if a saved PECGroupConfig exists
                try:
                    group_config = PECGroupConfig.objects.get(
                        semester=semester_obj,
                        branch=selected_branch,
                        program_type=selected_program,
                    )
                    saved_groups = group_config.groups or []
                except PECGroupConfig.DoesNotExist:
                    saved_groups = []
                
                # If no saved config, build groups from existing ElectiveCourseOfferings
                if not saved_groups:
                    # Build default groups (one per PEC slot)
                    for plan in sorted(pec_slots, key=lambda p: p.course.course_code):
                        group = []
                        offerings = elective_offerings_map.get(plan.id, [])
                        for off in offerings:
                            group.append({
                                'code': off.actual_course.course_code,
                                'title': off.actual_course.title,
                                'credits': off.actual_course.credits,
                                'type': off.actual_course.course_type or '',
                                'batch_count': off.batch_count,
                                'capacity': off.capacity_per_batch,
                            })
                        saved_groups.append(group)
                
                # Ensure minimum group count
                while len(saved_groups) < min_groups:
                    saved_groups.append([])
                
                context['pec_min_groups'] = min_groups
                context['pec_plan_ids_json'] = _json.dumps(pec_plan_ids)
                context['pec_groups_json'] = _json.dumps(saved_groups)
                context['pec_group_count'] = len(saved_groups)
                context['pec_has_matrix'] = True
                
                # For merged PEC row in the elective table
                context['pec_plan_ids_set'] = set(pec_plan_ids)
                pec_codes = sorted([p.course.course_code for p in pec_slots])
                context['pec_merged_label'] = ' & '.join(pec_codes)
                context['pec_merged_credits'] = pec_slots[0].course.credits if pec_slots else 3
                # Collect ALL offerings across all PEC slots
                pec_all_offerings = []
                for plan in pec_slots:
                    for off in elective_offerings_map.get(plan.id, []):
                        pec_all_offerings.append(off)
                context['pec_all_offerings'] = pec_all_offerings
                context['pec_slot_count'] = len(pec_slots)
                # Provide a small JSON map of PEC plan ids -> course codes for the template
                try:
                    context['pec_slots_json'] = _json.dumps([
                        {'id': p.id, 'code': p.course.course_code} for p in pec_slots
                    ])
                except Exception:
                    context['pec_slots_json'] = '[]'
        
        # Get existing course assignments for this semester
        existing_filter = {
            'semester': semester_obj,
            'academic_year': current_academic_year,
            'batch__program__code': selected_branch,
            'batch__program__level': selected_program,
        }
        if selected_batch:
            existing_filter['batch__batch_name'] = selected_batch

        existing_assignments = Course_Assignment.objects.filter(
            **existing_filter
        ).select_related('course', 'faculty', 'faculty__user', 'lab_assistant', 'lab_assistant__user', 'lab_main_faculty', 'lab_main_faculty__user')

        # Show Quick IOC panel only when IOC/EEC is part of this semester plan,
        # or when quick IOC rows already exist in the selected context.
        has_ioc_in_plan = False
        try:
            if 'course_plans' in context and context['course_plans']:
                has_ioc_in_plan = any(
                    (
                        (p.category and p.category.code in ['IOC', 'EEC'])
                        or (
                            p.course
                            and p.course.is_placeholder
                            and p.course.placeholder_type in ['IOC', 'EEC']
                        )
                    )
                    for p in context['course_plans']
                )
        except Exception:
            has_ioc_in_plan = False

        has_quick_ioc_rows = Course_Assignment.objects.filter(
            **existing_filter,
            special_note__icontains='Quick IOC',
            is_active=True,
        ).exists()

        context['show_quick_ioc'] = bool(has_ioc_in_plan or has_quick_ioc_rows)

        # Only show assignments that belong to the selected regulation's course plan.
        # Keep Quick IOC rows visible even when they are not part of the regulation plan.
        try:
            if 'course_plans' in context and context['course_plans']:
                allowed_codes = [p.course.course_code for p in context['course_plans']]
                existing_assignments = existing_assignments.filter(
                    Q(course__course_code__in=allowed_codes) |
                    Q(special_note__icontains='Quick IOC')
                )
        except Exception:
            pass

        context['existing_assignments'] = existing_assignments

        # Create a dict for quick lookup in template
        assignments_by_course = {}
        for assign in existing_assignments:
            code = assign.course.course_code
            if code not in assignments_by_course:
                assignments_by_course[code] = []
            assignments_by_course[code].append(assign)
        context['assignments_by_course'] = assignments_by_course
    
    return render(request, 'hod_template/semester_course_assignment.html', context)


def _sync_pg_course_clubbing(academic_year, semester, course, faculty, enable=True):
    """Create/update or disable PG clubbing for a course+faculty pair."""
    existing_group = ClubbedCourseGroup.objects.filter(
        academic_year=academic_year,
        semester=semester,
        course=course,
        faculty=faculty,
    ).first()

    if not enable:
        if existing_group:
            existing_group.is_active = False
            existing_group.save(update_fields=['is_active'])
            existing_group.members.all().delete()
        return False, 0

    matching_assignments = list(
        Course_Assignment.objects.filter(
            academic_year=academic_year,
            semester=semester,
            course=course,
            faculty=faculty,
            batch__program__level='PG',
            batch__is_active=True,
            is_active=True,
        ).select_related('batch__program')
    )

    distinct_program_ids = {a.batch.program_id for a in matching_assignments if a.batch_id and a.batch and a.batch.program_id}
    batch_ids = sorted({a.batch_id for a in matching_assignments if a.batch_id})

    if len(distinct_program_ids) < 2:
        if existing_group:
            existing_group.is_active = False
            existing_group.save(update_fields=['is_active'])
            existing_group.members.all().delete()
        return False, len(distinct_program_ids)

    group, _ = ClubbedCourseGroup.objects.get_or_create(
        academic_year=academic_year,
        semester=semester,
        course=course,
        faculty=faculty,
        defaults={'is_active': True},
    )
    if not group.is_active:
        group.is_active = True
        group.save(update_fields=['is_active'])

    group.members.exclude(program_batch_id__in=batch_ids).delete()
    for batch_id in batch_ids:
        ClubbedCourseMember.objects.get_or_create(group=group, program_batch_id=batch_id)

    return True, len(distinct_program_ids)


@login_required
def create_course_assignments(request):
    """Create course assignments from semester course assignment page"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method == 'POST':
        semester_id = request.POST.get('semester_id')
        academic_year_id = request.POST.get('academic_year_id')
        selected_branch = request.POST.get('branch', '').strip()
        selected_program_type = request.POST.get('program_type', 'UG').strip()
        
        semester_obj = get_object_or_404(Semester, id=semester_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

        selected_program_obj = Program.objects.filter(
            code=selected_branch,
            level=selected_program_type,
        ).first()
        if not selected_program_obj and selected_branch:
            # Fallback for composite branch tokens (e.g. "CSE-DCS")
            # to avoid losing assignments due to strict code matching.
            base_branch = selected_branch.split('-')[0].strip()
            selected_program_obj = Program.objects.filter(
                level=selected_program_type
            ).filter(
                Q(code=base_branch) |
                Q(code__iexact=selected_branch) |
                Q(code__istartswith=selected_branch) |
                Q(code__istartswith=base_branch)
            ).first()
        if not selected_program_obj:
            messages.error(request, "Selected program not found for assignment.")
            return redirect(
                f"{reverse('semester_course_assignment')}?semester_id={semester_id}&branch={selected_branch}&batch={request.POST.get('batch', '')}&program_type={selected_program_type}"
            )
        
        # Process each course assignment
        course_codes = request.POST.getlist('course_code[]')
        faculty_ids = request.POST.getlist('faculty_id[]')
        batch_labels = request.POST.getlist('batch_label[]')
        lab_main_faculty_ids = request.POST.getlist('lab_main_faculty_id[]')
        lab_assistant_ids = request.POST.getlist('lab_assistant_id[]')
        separate_lab_theory_flags = request.POST.getlist('separate_lab_theory_staff[]')
        club_same_time = request.POST.get('club_same_time') == '1'
        practical_in_classroom = str(request.POST.get('practical_in_classroom', '')).strip().lower() in ['1', 'true', 'yes', 'on']

        # Fallback path: if dynamic JS arrays are missing, use static modal fields.
        if not course_codes:
            fallback_course_code = (request.POST.get('course_code_fallback', '') or '').strip()
            fallback_faculty_ids = request.POST.getlist('faculty_id_fallback[]')
            fallback_batch_labels = request.POST.getlist('batch_label_fallback[]')
            fallback_lab_main_faculty_ids = request.POST.getlist('lab_main_faculty_id_fallback[]')
            fallback_lab_assistant_ids = request.POST.getlist('lab_assistant_id_fallback[]')
            fallback_separate_flags = request.POST.getlist('separate_lab_theory_staff_fallback[]')

            if fallback_course_code and fallback_faculty_ids and fallback_batch_labels:
                for idx, fallback_faculty_id in enumerate(fallback_faculty_ids):
                    if not str(fallback_faculty_id or '').strip():
                        continue
                    if idx >= len(fallback_batch_labels):
                        continue
                    course_codes.append(fallback_course_code)
                    faculty_ids.append(fallback_faculty_id)
                    batch_labels.append(fallback_batch_labels[idx])
                    lab_main_faculty_ids.append(
                        fallback_lab_main_faculty_ids[idx] if idx < len(fallback_lab_main_faculty_ids) else ''
                    )
                    lab_assistant_ids.append(
                        fallback_lab_assistant_ids[idx] if idx < len(fallback_lab_assistant_ids) else ''
                    )
                    separate_lab_theory_flags.append(
                        fallback_separate_flags[idx] if idx < len(fallback_separate_flags) else '0'
                    )

        def _parse_non_negative(name, default=0):
            raw = (request.POST.get(name, '') or '').strip()
            if raw == '':
                return default
            try:
                value = int(raw)
            except (TypeError, ValueError):
                return default
            return max(0, value)

        ioc_lecture_hours = _parse_non_negative('ioc_lecture_hours', 0)
        ioc_tutorial_hours = _parse_non_negative('ioc_tutorial_hours', 0)
        ioc_practical_hours = _parse_non_negative('ioc_practical_hours', 0)

        # Parse IOC-specific fields (lists indexed by assignment index)
        ioc_course_codes = request.POST.getlist('ioc_course_code[]')
        ioc_lecture_hours_list = request.POST.getlist('ioc_lecture_hours[]')
        ioc_tutorial_hours_list = request.POST.getlist('ioc_tutorial_hours[]')
        ioc_practical_hours_list = request.POST.getlist('ioc_practical_hours[]')
        ioc_schedule_timetable_list = request.POST.getlist('ioc_schedule_timetable[]')
        
        # Parse SDC-specific fields (lists indexed by assignment index)
        sdc_course_codes = request.POST.getlist('sdc_course_code[]')
        sdc_lecture_hours_list = request.POST.getlist('sdc_lecture_hours[]')
        sdc_tutorial_hours_list = request.POST.getlist('sdc_tutorial_hours[]')
        sdc_practical_hours_list = request.POST.getlist('sdc_practical_hours[]')
        sdc_lab_assistant_ids = request.POST.getlist('sdc_lab_assistant_id[]')
        sdc_schedule_timetable_list = request.POST.getlist('sdc_schedule_timetable[]')
        
        created_count = 0
        errors = []
        current_pairs = set()
        cleanup_pairs = set()
        
        try:
            with transaction.atomic():
                for i, course_code in enumerate(course_codes):
                    if not course_code:
                        continue

                    faculty_id = faculty_ids[i] if i < len(faculty_ids) else None
                    batch_label = batch_labels[i] if i < len(batch_labels) else None
                    lab_main_faculty_id = lab_main_faculty_ids[i] if i < len(lab_main_faculty_ids) else None
                    lab_assistant_id = lab_assistant_ids[i] if i < len(lab_assistant_ids) else None
                    separate_lab_theory_raw = separate_lab_theory_flags[i] if i < len(separate_lab_theory_flags) else '0'
                    separate_lab_theory_staff = str(separate_lab_theory_raw).strip().lower() in ['1', 'true', 'yes', 'on']

                    # IOC-specific data for this assignment
                    ioc_course_code = ioc_course_codes[i] if i < len(ioc_course_codes) else None
                    try:
                        ioc_lec = max(0, int(ioc_lecture_hours_list[i])) if i < len(ioc_lecture_hours_list) else ioc_lecture_hours
                    except (TypeError, ValueError):
                        ioc_lec = ioc_lecture_hours
                    try:
                        ioc_tut = max(0, int(ioc_tutorial_hours_list[i])) if i < len(ioc_tutorial_hours_list) else ioc_tutorial_hours
                    except (TypeError, ValueError):
                        ioc_tut = ioc_tutorial_hours
                    try:
                        ioc_prac = max(0, int(ioc_practical_hours_list[i])) if i < len(ioc_practical_hours_list) else ioc_practical_hours
                    except (TypeError, ValueError):
                        ioc_prac = ioc_practical_hours
                    ioc_schedule = ioc_schedule_timetable_list[i] if i < len(ioc_schedule_timetable_list) else '1'
                    
                    # SDC-specific data for this assignment
                    sdc_course_code = sdc_course_codes[i] if i < len(sdc_course_codes) else None
                    try:
                        sdc_lec = max(0, int(sdc_lecture_hours_list[i])) if i < len(sdc_lecture_hours_list) else 0
                    except (TypeError, ValueError):
                        sdc_lec = 0
                    try:
                        sdc_tut = max(0, int(sdc_tutorial_hours_list[i])) if i < len(sdc_tutorial_hours_list) else 0
                    except (TypeError, ValueError):
                        sdc_tut = 0
                    try:
                        sdc_prac = max(0, int(sdc_practical_hours_list[i])) if i < len(sdc_practical_hours_list) else 0
                    except (TypeError, ValueError):
                        sdc_prac = 0
                    sdc_lab_asst_id = sdc_lab_assistant_ids[i] if i < len(sdc_lab_assistant_ids) else None
                    sdc_schedule = sdc_schedule_timetable_list[i] if i < len(sdc_schedule_timetable_list) else '1'

                    if not faculty_id or not batch_label:
                        continue

                    try:
                        course = Course.objects.get(course_code=course_code)
                        is_ioc_flow = bool(course.is_placeholder and course.placeholder_type in ['IOC', 'EEC'])
                        is_sdc_flow = bool(course.is_placeholder and course.placeholder_type in ['SDC', 'SLC'])
                        mapped_ioc_course = None
                        mapped_sdc_course = None
                        if is_ioc_flow and ioc_course_code:
                            mapped_ioc_course = Course.objects.filter(course_code=ioc_course_code).first()
                        if is_sdc_flow and sdc_course_code:
                            mapped_sdc_course = Course.objects.filter(course_code=sdc_course_code).first()
                        faculty = Faculty_Profile.objects.get(id=faculty_id)
                        lab_main_faculty = None
                        if lab_main_faculty_id:
                            lab_main_faculty = Faculty_Profile.objects.get(id=lab_main_faculty_id)
                        lab_assistant = None
                        if lab_assistant_id:
                            lab_assistant = Faculty_Profile.objects.get(id=lab_assistant_id)
                        
                        sdc_lab_assistant = None
                        if sdc_lab_asst_id:
                            sdc_lab_assistant = Faculty_Profile.objects.get(id=sdc_lab_asst_id)

                        normalized_batch_label = (batch_label or '').strip()
                        batch_candidates = []
                        for candidate in [
                            normalized_batch_label,
                            normalized_batch_label.replace(' Section', '').replace(' section', '').strip(),
                            (normalized_batch_label.split()[0] if normalized_batch_label else ''),
                        ]:
                            if candidate and candidate not in batch_candidates:
                                batch_candidates.append(candidate)

                        source_year = academic_year
                        source_year_of_study = semester_obj.year_of_study

                        target_batch = ProgramBatch.objects.filter(
                            academic_year=source_year,
                            program=selected_program_obj,
                            year_of_study=source_year_of_study,
                            batch_name__in=batch_candidates,
                            is_active=True,
                        ).first()

                        if not target_batch:
                            # Resolve by program code+level if direct FK match fails.
                            target_batch = ProgramBatch.objects.filter(
                                academic_year=source_year,
                                program__code=selected_branch,
                                program__level=selected_program_type,
                                year_of_study=source_year_of_study,
                                batch_name__in=batch_candidates,
                                is_active=True,
                            ).first()

                        # Cohort mapping for higher years:
                        # If current AY doesn't have year-wise rows, use admission AY Year 1 rows.
                        if not target_batch and semester_obj.year_of_study > 1:
                            try:
                                start_year = int(academic_year.year.split('-')[0])
                                admission_start_year = start_year - (semester_obj.year_of_study - 1)
                                admission_year_label = f"{admission_start_year}-{str(admission_start_year + 1)[-2:]}"
                                source_year = AcademicYear.objects.filter(year=admission_year_label).first()
                                source_year_of_study = 1
                            except (ValueError, IndexError):
                                source_year = None

                            if source_year:
                                target_batch = ProgramBatch.objects.filter(
                                    academic_year=source_year,
                                    program__code=selected_branch,
                                    program__level=selected_program_type,
                                    year_of_study=source_year_of_study,
                                    batch_name__in=batch_candidates,
                                    is_active=True,
                                ).first()

                        if not target_batch:
                            errors.append(
                                f"Batch {batch_label} not found in {selected_program_type} {selected_branch} for year {semester_obj.year_of_study}"
                            )
                            continue

                        # IOC/EEC placeholder without LTP should receive LTP at assignment time.
                        if is_ioc_flow:
                            # If a mapped IOC course is selected, use its LTP for this slot.
                            if mapped_ioc_course:
                                ioc_lec = max(0, int(mapped_ioc_course.lecture_hours or 0))
                                ioc_tut = max(0, int(mapped_ioc_course.tutorial_hours or 0))
                                ioc_prac = max(0, int(mapped_ioc_course.practical_hours or 0))

                            current_total_ltp = (course.lecture_hours or 0) + (course.tutorial_hours or 0) + (course.practical_hours or 0)
                            entered_total = ioc_lec + ioc_tut + ioc_prac
                            if current_total_ltp <= 0:
                                if entered_total <= 0:
                                    errors.append(
                                        f"{course.course_code}: Enter L-T-P while assigning faculty (minimum one period)."
                                    )
                                    continue

                            # Apply entered/mapped LTP to IOC slot when slot has no LTP yet,
                            # when mapped course is selected, or when user changes L-T-P.
                            has_ioc_ltp_change = (
                                (course.lecture_hours or 0) != ioc_lec
                                or (course.tutorial_hours or 0) != ioc_tut
                                or (course.practical_hours or 0) != ioc_prac
                            )
                            if entered_total > 0 and (current_total_ltp <= 0 or mapped_ioc_course or has_ioc_ltp_change):

                                if ioc_prac > 0 and (ioc_lec + ioc_tut) > 0:
                                    resolved_type = 'LIT'
                                elif ioc_prac > 0:
                                    resolved_type = 'L'
                                else:
                                    resolved_type = 'T'

                                course.lecture_hours = ioc_lec
                                course.tutorial_hours = ioc_tut
                                course.practical_hours = ioc_prac
                                course.course_type = resolved_type
                                course.save(update_fields=['lecture_hours', 'tutorial_hours', 'practical_hours', 'course_type', 'updated_at'])
                        
                        # SDC/SLC placeholder without LTP should receive LTP at assignment time.
                        if is_sdc_flow:
                            # If a mapped SDC/SLC course is selected, use its LTP for this slot.
                            if mapped_sdc_course:
                                sdc_lec = max(0, int(mapped_sdc_course.lecture_hours or 0))
                                sdc_tut = max(0, int(mapped_sdc_course.tutorial_hours or 0))
                                sdc_prac = max(0, int(mapped_sdc_course.practical_hours or 0))

                            current_total_ltp = (course.lecture_hours or 0) + (course.tutorial_hours or 0) + (course.practical_hours or 0)
                            entered_total = sdc_lec + sdc_tut + sdc_prac
                            if current_total_ltp <= 0:
                                if entered_total <= 0:
                                    errors.append(
                                        f"{course.course_code}: Enter L-T-P while assigning faculty (minimum one period)."
                                    )
                                    continue

                            # Keep assignments on slot code, but allow selected mapped course
                            # or explicit user changes to update slot LTP.
                            has_sdc_ltp_change = (
                                (course.lecture_hours or 0) != sdc_lec
                                or (course.tutorial_hours or 0) != sdc_tut
                                or (course.practical_hours or 0) != sdc_prac
                            )
                            if entered_total > 0 and (current_total_ltp <= 0 or mapped_sdc_course or has_sdc_ltp_change):

                                if sdc_prac > 0 and (sdc_lec + sdc_tut) > 0:
                                    resolved_type = 'LIT'
                                elif sdc_prac > 0:
                                    resolved_type = 'L'
                                else:
                                    resolved_type = 'T'

                                course.lecture_hours = sdc_lec
                                course.tutorial_hours = sdc_tut
                                course.practical_hours = sdc_prac
                                course.course_type = resolved_type
                                course.save(update_fields=['lecture_hours', 'tutorial_hours', 'practical_hours', 'course_type', 'updated_at'])

                        existing = Course_Assignment.objects.filter(
                            course=course,
                            batch=target_batch,
                            academic_year=academic_year,
                            semester=semester_obj
                        ).first()
                        old_faculty = existing.faculty if existing else None

                        assignment_defaults = {
                            'faculty': faculty,
                            'separate_lab_theory_staff': False,
                            'lab_main_faculty': None,
                            'lab_assistant': lab_assistant,
                            'batch_label': batch_label,
                            'practical_in_classroom': practical_in_classroom,
                            'is_active': True,
                        }

                        # For LIT courses, allow separate theory and lab main faculty.
                        if course.course_type == 'LIT':
                            if separate_lab_theory_staff and not lab_main_faculty:
                                errors.append(
                                    f"{course.course_code}: Lab main faculty is required when separate lab/theory staff is enabled."
                                )
                                continue
                            assignment_defaults['separate_lab_theory_staff'] = bool(separate_lab_theory_staff)
                            assignment_defaults['lab_main_faculty'] = lab_main_faculty if separate_lab_theory_staff else None

                        # For SDC/SLC assignment flow, use SDC-specific lab assistant and timetable note.
                        if is_sdc_flow:
                            assignment_defaults['lab_assistant'] = sdc_lab_assistant or lab_assistant
                            assignment_defaults['separate_lab_theory_staff'] = False
                            assignment_defaults['lab_main_faculty'] = None
                            sdc_notes = []
                            if str(sdc_schedule) == '0':
                                sdc_notes.append('Skip Timetable')
                            if mapped_sdc_course:
                                sdc_notes.append(f"Mapped Course: {mapped_sdc_course.course_code}")
                            assignment_defaults['special_note'] = ' | '.join(sdc_notes)

                        # For IOC/EEC assignment flow, support timetable skip selection.
                        if is_ioc_flow:
                            ioc_notes = []
                            if str(ioc_schedule) == '0':
                                ioc_notes.append('Skip Timetable')
                            if mapped_ioc_course:
                                ioc_notes.append(f"Mapped Course: {mapped_ioc_course.course_code}")
                            assignment_defaults['special_note'] = ' | '.join(ioc_notes)

                        Course_Assignment.objects.update_or_create(
                            course=course,
                            batch=target_batch,
                            academic_year=academic_year,
                            semester=semester_obj,
                            defaults=assignment_defaults
                        )

                        # If this is IOC/EEC placeholder assignment, mirror faculty to quick IOC for the same batch.
                        if is_ioc_flow:
                            quick_ioc_qs = Course_Assignment.objects.filter(
                                semester=semester_obj,
                                academic_year=academic_year,
                                batch=target_batch,
                                special_note__icontains='Quick IOC',
                                is_active=True,
                            )
                            for quick_ioc in quick_ioc_qs:
                                if quick_ioc.faculty_id != faculty.id:
                                    quick_ioc.faculty = faculty
                                    quick_ioc.save(update_fields=['faculty', 'updated_at'])

                        created_count += 1

                        if selected_program_type == 'PG':
                            current_pairs.add((course.course_code, faculty.id))
                            if old_faculty and old_faculty.id != faculty.id:
                                cleanup_pairs.add((course.course_code, old_faculty.id))
                    except Exception as e:
                        errors.append(f"Error for {course_code}: {str(e)}")

                clubbing_updates = []
                if selected_program_type == 'PG':
                    for course_code, faculty_id in cleanup_pairs:
                        _sync_pg_course_clubbing(
                            academic_year=academic_year,
                            semester=semester_obj,
                            course=Course.objects.get(course_code=course_code),
                            faculty=Faculty_Profile.objects.get(id=faculty_id),
                            enable=True,
                        )

                    for course_code, faculty_id in current_pairs:
                        course = Course.objects.get(course_code=course_code)
                        faculty = Faculty_Profile.objects.get(id=faculty_id)
                        is_clubbed, program_count = _sync_pg_course_clubbing(
                            academic_year=academic_year,
                            semester=semester_obj,
                            course=course,
                            faculty=faculty,
                            enable=club_same_time,
                        )
                        if club_same_time and is_clubbed:
                            clubbing_updates.append(
                                f"{course_code} clubbed across {program_count} PG programs"
                            )
                        elif club_same_time and not is_clubbed and program_count < 2:
                            clubbing_updates.append(
                                f"{course_code} saved; clubbing will activate once the same faculty is assigned in another PG program"
                            )

        except Exception as e:
            errors.append(str(e))
        
        if created_count > 0:
            messages.success(request, f"Created {created_count} course assignment(s)")
        elif errors:
            messages.error(request, f"No assignments were created. {errors[0]}")
        if selected_program_type == 'PG' and club_same_time and 'clubbing_updates' in locals() and clubbing_updates:
            messages.info(request, '; '.join(clubbing_updates[:3]))
        if errors:
            messages.warning(request, f"Some errors occurred: {'; '.join(errors[:3])}")
        
        # Redirect back with same filters
        return redirect(f"{reverse('semester_course_assignment')}?semester_id={semester_id}&branch={request.POST.get('branch', 'CSE')}&batch={request.POST.get('batch', '')}&program_type={request.POST.get('program_type', 'UG')}")
    
    return redirect('semester_course_assignment')


# =============================================================================
# PROGRAM BATCH MANAGEMENT
# =============================================================================

@login_required
def manage_program_batches(request, year_id=None):
    """Manage classroom batches for programs in an academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    # Select academic year
    if year_id:
        selected_year = get_object_or_404(AcademicYear, id=year_id)
    else:
        selected_year = AcademicYear.get_current() or academic_years.first()
    
    # Get all programs with regulation
    programs = Program.objects.all().select_related('regulation').order_by('level', 'code')
    
    # Get Year 1 batches for the selected year, grouped by program
    year1_batches = []
    if selected_year:
        batches = ProgramBatch.objects.filter(
            academic_year=selected_year,
            year_of_study=1  # Only Year 1 batches
        ).select_related('program', 'program__regulation').order_by('program__level', 'program__code', 'batch_name')
        
        # Group batches by program (using ID to differentiate same code under different regulations)
        batches_by_program = {}
        for batch in batches:
            if batch.program.id not in batches_by_program:
                # Check if students exist for this program's Year 1
                has_students = ProgramBatch.has_students(selected_year, batch.program, 1)
                batches_by_program[batch.program.id] = {
                    'program': batch.program,
                    'batches': [],
                    'has_students': has_students,
                }
            batches_by_program[batch.program.id]['batches'].append(batch)
        
        year1_batches = list(batches_by_program.values())
    
    # Get previous year for copy option
    previous_year = None
    if selected_year:
        try:
            start_year = int(selected_year.year.split('-')[0])
            prev_year_str = f"{start_year - 1}-{str(start_year)[-2:]}"
            previous_year = AcademicYear.objects.filter(year=prev_year_str).first()
        except:
            pass
    
    # Check which programs don't have Year 1 batches configured
    programs_without_batches = []
    for program in programs:
        has_year1 = ProgramBatch.objects.filter(
            academic_year=selected_year,
            program=program,
            year_of_study=1,
            is_active=True
        ).exists() if selected_year else False
        if not has_year1:
            programs_without_batches.append(program)
    
    context = {
        'page_title': 'Manage Program Batches',
        'academic_years': academic_years,
        'selected_year': selected_year,
        'programs': programs,
        'year1_batches': year1_batches,
        'previous_year': previous_year,
        'programs_without_batches': programs_without_batches,
    }
    return render(request, 'hod_template/manage_program_batches.html', context)


@login_required
def add_program_batch(request):
    """Add a new batch for a program"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method == 'POST':
        year_id = request.POST.get('academic_year')
        program_id = request.POST.get('program')
        year_of_study = request.POST.get('year_of_study')
        batch_names = request.POST.get('batch_names', '').strip()  # Comma-separated
        capacity = request.POST.get('capacity', 60)
        
        if not all([year_id, program_id, year_of_study, batch_names]):
            messages.error(request, "All fields are required")
            return redirect('manage_program_batches', year_id=year_id)
        
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        program = get_object_or_404(Program, id=program_id)
        
        # Parse batch names (comma or space separated)
        import re
        batch_list = re.split(r'[,\s]+', batch_names.upper())
        batch_list = [b.strip() for b in batch_list if b.strip()]
        
        created_count = 0
        skipped_count = 0
        
        for batch_name in batch_list:
            _, was_created = ProgramBatch.objects.get_or_create(
                academic_year=academic_year,
                program=program,
                year_of_study=int(year_of_study),
                batch_name=batch_name,
                defaults={
                    'batch_display': f"{batch_name} Section",
                    'capacity': int(capacity),
                    'is_active': True
                }
            )
            if was_created:
                created_count += 1
            else:
                skipped_count += 1
        
        if created_count:
            messages.success(request, f"Added {created_count} batch(es) for {program.code} Year {year_of_study}")
        if skipped_count:
            messages.info(request, f"{skipped_count} batch(es) already existed")
        
        return redirect('manage_program_batches_year', year_id=year_id)
    
    return redirect('manage_program_batches')


@login_required
def copy_batches_from_previous_year(request):
    """Copy batch configuration from previous academic year"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method == 'POST':
        source_year_id = request.POST.get('source_year')
        target_year_id = request.POST.get('target_year')
        program_id = request.POST.get('program')  # Optional - if None, copy all
        
        if not all([source_year_id, target_year_id]):
            messages.error(request, "Source and target years are required")
            return redirect('manage_program_batches')
        
        source_year = get_object_or_404(AcademicYear, id=source_year_id)
        target_year = get_object_or_404(AcademicYear, id=target_year_id)
        program = Program.objects.filter(id=program_id).first() if program_id else None
        
        created, skipped = ProgramBatch.copy_from_previous_year(source_year, target_year, program)
        
        if created:
            messages.success(request, f"Copied {created} batch(es) from {source_year} to {target_year}")
        if skipped:
            messages.info(request, f"{skipped} batch(es) already existed")
        if not created and not skipped:
            messages.warning(request, f"No batches found to copy from {source_year}")
        
        return redirect('manage_program_batches_year', year_id=target_year_id)
    
    return redirect('manage_program_batches')


@login_required
def delete_program_batch(request, batch_id):
    """Delete a program batch"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    batch = get_object_or_404(ProgramBatch, id=batch_id)
    year_id = batch.academic_year.id
    
    # Check if any students are assigned to this batch
    student_count = Student_Profile.objects.filter(
        branch=batch.program.code,
        batch_label=batch.batch_name
    ).count()
    
    if student_count > 0:
        messages.error(request, f"Cannot delete batch {batch.batch_name} - {student_count} student(s) assigned")
    else:
        batch.delete()
        messages.success(request, f"Batch {batch.batch_name} deleted successfully")
    
    return redirect('manage_program_batches_year', year_id=year_id)


@login_required
def initialize_default_batches(request, year_id, program_id):
    """Initialize default batches for a program from its default settings"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    academic_year = get_object_or_404(AcademicYear, id=year_id)
    program = get_object_or_404(Program, id=program_id)
    
    # Only allow for Year 1 and if no students exist
    has_students = ProgramBatch.has_students(academic_year, program, 1)
    if has_students:
        messages.error(request, f"Cannot initialize batches - students already assigned to {program.code}")
        return redirect('manage_program_batches_year', year_id=year_id)
    
    # Create default batches
    created_count, created_names = ProgramBatch.create_default_batches(
        academic_year=academic_year,
        program=program,
        year_of_study=1,
        capacity=60
    )
    
    if created_count:
        messages.success(request, f"Created {created_count} batch(es) for {program.code}: {', '.join(created_names)}")
    else:
        messages.info(request, f"Batches already exist for {program.code}")
    
    return redirect('manage_program_batches_year', year_id=year_id)


@login_required
def api_get_batches(request):
    """API endpoint to get batches filtered by program, year, and academic year"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    academic_year_id = request.GET.get('academic_year')
    program_code = request.GET.get('program')
    year_of_study = request.GET.get('year_of_study')
    
    # Get academic year
    if academic_year_id:
        academic_year = AcademicYear.objects.filter(id=academic_year_id).first()
    else:
        academic_year = AcademicYear.get_current()
    
    if not academic_year:
        return JsonResponse({'batches': []})
    
    # Build query
    qs = ProgramBatch.objects.filter(academic_year=academic_year, is_active=True)
    
    if program_code:
        qs = qs.filter(program__code=program_code)
    if year_of_study:
        qs = qs.filter(year_of_study=int(year_of_study))
    
    batches = list(qs.values('id', 'batch_name', 'batch_display', 'capacity', 'year_of_study').order_by('batch_name'))
    
    return JsonResponse({'batches': batches})


# =============================================================================
# PROGRAM MANAGEMENT
# =============================================================================

@login_required
def add_program(request):
    """Add academic program"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = ProgramForm(request.POST or None)
    context = {'form': form, 'page_title': 'Add Academic Program'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    program = form.save()

                    if program.regulation:
                        ProgramRegulation.objects.get_or_create(
                            program=program,
                            regulation=program.regulation,
                            defaults={
                                'effective_from_year': program.regulation.year,
                                'effective_to_year': None,
                                'is_active': True,
                                'notes': 'Auto-created when program was added.',
                            }
                        )

                messages.success(request, "Program added successfully!")
                return redirect(reverse('manage_programs'))
            except Exception as e:
                messages.error(request, f"Could not add: {str(e)}")
        else:
            # Show specific field errors
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        field_label = form.fields[field].label or field.replace('_', ' ').title()
                        messages.error(request, f"{field_label}: {error}")
    
    return render(request, "hod_template/add_program.html", context)


@login_required
def manage_programs(request):
    """Manage academic programs"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    programs = Program.objects.all().select_related('regulation')
    context = {'programs': programs, 'page_title': 'Manage Programs'}
    return render(request, "hod_template/manage_programs.html", context)


@login_required
def edit_program(request, program_id):
    """Edit academic program"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    program = get_object_or_404(Program, id=program_id)
    form = ProgramForm(request.POST or None, instance=program)
    
    # Get actual student count using branch field (stores program code)
    student_count = Student_Profile.objects.filter(branch=program.code).count()
    
    context = {
        'form': form, 
        'program': program, 
        'student_count': student_count,
        'page_title': f'Edit Program - {program.code}'
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Program updated successfully!")
                return redirect(reverse('manage_programs'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please correct the errors")
    
    return render(request, "hod_template/edit_program.html", context)


@login_required
def delete_program(request, program_id):
    """Delete academic program"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    try:
        program = get_object_or_404(Program, id=program_id)
        program.delete()
        messages.success(request, "Program deleted successfully")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_programs'))


# =============================================================================
# LEAVE MANAGEMENT
# =============================================================================

@login_required
@csrf_exempt
def view_leave_requests(request):
    """View and manage all leave requests"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method == 'POST':
        leave_id = request.POST.get('id')
        status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        
        try:
            leave = get_object_or_404(LeaveRequest, id=leave_id)
            leave.status = status
            leave.admin_remarks = remarks
            leave.approved_by = request.user
            leave.save()
            
            # Send notification
            Notification.objects.create(
                recipient=leave.user,
                sender=request.user,
                title=f"Leave Request {status}",
                message=f"Your leave request from {leave.start_date} to {leave.end_date} has been {status.lower()}.",
                notification_type='INFO'
            )
            
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)
    
    leaves = LeaveRequest.objects.select_related('user').order_by('-created_at')
    pending_leaves = leaves.filter(status='PENDING')
    processed_leaves = leaves.exclude(status='PENDING')
    
    context = {
        'pending_leaves': pending_leaves,
        'processed_leaves': processed_leaves,
        'page_title': 'Leave Requests'
    }
    return render(request, "hod_template/staff_leave_view.html", context)


# =============================================================================
# FEEDBACK MANAGEMENT
# =============================================================================

@login_required
@csrf_exempt
def view_feedbacks(request):
    """View and respond to feedbacks"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    if request.method == 'POST':
        feedback_id = request.POST.get('id')
        reply = request.POST.get('reply')
        
        try:
            feedback = get_object_or_404(Feedback, id=feedback_id)
            feedback.reply = reply
            feedback.status = 'REVIEWED'
            feedback.replied_by = request.user
            feedback.save()
            
            # Send notification
            if not feedback.is_anonymous:
                Notification.objects.create(
                    recipient=feedback.user,
                    sender=request.user,
                    title="Feedback Response",
                    message=f"Your feedback '{feedback.subject[:30]}...' has been responded to.",
                    notification_type='INFO'
                )
            
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)
    
    feedbacks = Feedback.objects.select_related('user').order_by('-created_at')
    
    context = {
        'feedbacks': feedbacks,
        'page_title': 'Feedback Messages'
    }
    return render(request, 'hod_template/student_feedback_template.html', context)


# =============================================================================
# EVENT MANAGEMENT
# =============================================================================

@login_required
def add_event(request):
    """Add new event"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = EventForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Event'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Event added successfully!")
                return redirect(reverse('manage_event'))
            except Exception as e:
                messages.error(request, f"Could not add event: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_event_template.html', context)


@login_required
def manage_event(request):
    """Manage events"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    events = Event.objects.all()
    context = {
        'events': events,
        'page_title': 'Manage Events'
    }
    return render(request, 'hod_template/manage_event.html', context)


@login_required
def edit_event(request, event_id):
    """Edit event"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    event = get_object_or_404(Event, id=event_id)
    form = EventForm(request.POST or None, request.FILES or None, instance=event)
    context = {'form': form, 'event': event, 'page_title': 'Edit Event'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Event updated successfully!")
                return redirect(reverse('manage_event'))
            except Exception as e:
                messages.error(request, f"Could not update event: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_event_template.html', context)


@login_required
def delete_event(request, event_id):
    """Delete event"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    event = get_object_or_404(Event, id=event_id)
    try:
        event.delete()
        messages.success(request, "Event deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_event'))


# =============================================================================
# PUBLICATION VERIFICATION
# =============================================================================

@login_required
def verify_publications(request):
    """View and verify faculty publications"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    publications = Publication.objects.select_related('faculty', 'faculty__user').order_by('-created_at')
    unverified = publications.filter(is_verified=False)
    verified = publications.filter(is_verified=True)
    
    context = {
        'unverified_publications': unverified,
        'verified_publications': verified,
        'page_title': 'Verify Publications'
    }
    return render(request, 'hod_template/verify_publications.html', context)


@login_required
@csrf_exempt
def approve_publication(request, publication_id):
    """Approve a publication"""
    if not check_hod_permission(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        publication = get_object_or_404(Publication, id=publication_id)
        publication.is_verified = True
        publication.verified_by = request.user
        publication.save()
        
        # Notify faculty
        Notification.objects.create(
            recipient=publication.faculty.user,
            sender=request.user,
            title="Publication Verified",
            message=f"Your publication '{publication.title[:50]}...' has been verified by HOD.",
            notification_type='INFO'
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# =============================================================================
# LAB ISSUES MANAGEMENT
# =============================================================================

@login_required
def view_lab_issues(request):
    """View and manage lab issues"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    issues = Lab_Issue_Log.objects.select_related('reported_by', 'assigned_to').order_by('-reported_at')
    
    context = {
        'issues': issues,
        'page_title': 'Lab Issues'
    }
    return render(request, 'hod_template/view_lab_issues.html', context)


# =============================================================================
# ANNOUNCEMENTS
# =============================================================================

@login_required
def add_announcement(request):
    """Add department announcement"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = AnnouncementForm(request.POST or None, request.FILES or None)
    context = {'form': form, 'page_title': 'Add Announcement'}
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                announcement = form.save(commit=False)
                announcement.posted_by = request.user
                announcement.save()
                messages.success(request, "Announcement posted successfully!")
                return redirect(reverse('manage_announcement'))
            except Exception as e:
                messages.error(request, f"Could not post: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields correctly")
    
    return render(request, 'hod_template/add_announcement.html', context)


@login_required
def manage_announcement(request):
    """Manage announcements"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    announcements = Announcement.objects.order_by('-created_at')
    context = {
        'announcements': announcements,
        'page_title': 'Manage Announcements'
    }
    return render(request, 'hod_template/manage_announcement.html', context)


@login_required
def delete_announcement(request, announcement_id):
    """Delete an announcement"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    try:
        announcement = get_object_or_404(Announcement, id=announcement_id)
        announcement.delete()
        messages.success(request, "Announcement deleted successfully!")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_announcement'))


# =============================================================================
# NOTIFICATIONS
# =============================================================================

@login_required
def send_notification_page(request):
    """Page to send notifications"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    users = Account_User.objects.filter(is_active=True).exclude(id=request.user.id)
    context = {
        'users': users,
        'page_title': 'Send Notifications'
    }
    return render(request, "hod_template/staff_notification.html", context)


@login_required
@csrf_exempt
def send_notification(request):
    """Send notification to user"""
    if not check_hod_permission(request.user):
        return HttpResponse(False)
    
    user_id = request.POST.get('id')
    message = request.POST.get('message')
    title = request.POST.get('title', 'Notification from HOD')
    
    try:
        recipient = get_object_or_404(Account_User, id=user_id)
        
        Notification.objects.create(
            recipient=recipient,
            sender=request.user,
            title=title,
            message=message,
            notification_type='INFO'
        )
        
        # Send FCM notification if token exists
        if recipient.fcm_token:
            url = "https://fcm.googleapis.com/fcm/send"
            body = {
                'notification': {
                    'title': title,
                    'body': message,
                    'icon': static('dist/img/AdminLTELogo.png')
                },
                'to': recipient.fcm_token
            }
            headers = {
                'Authorization': 'key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB',
                'Content-Type': 'application/json'
            }
            requests.post(url, data=json.dumps(body), headers=headers)
        
        return HttpResponse(True)
    except Exception as e:
        return HttpResponse(False)


# =============================================================================
# PROFILE
# =============================================================================

@login_required
def admin_view_profile(request):
    """View/edit HOD profile"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    user = request.user
    form = AccountUserForm(request.POST or None, request.FILES or None, instance=user)
    
    context = {
        'form': form,
        'page_title': 'View/Edit Profile'
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                user = form.save(commit=False)
                password = form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                
                if 'profile_pic' in request.FILES:
                    fs = FileSystemStorage()
                    filename = fs.save(request.FILES['profile_pic'].name, request.FILES['profile_pic'])
                    user.profile_pic = fs.url(filename)
                
                user.save()
                messages.success(request, "Profile updated successfully!")
                return redirect(reverse('admin_view_profile'))
            except Exception as e:
                messages.error(request, f"Could not update: {str(e)}")
        else:
            messages.error(request, "Please fill the form correctly")
    
    return render(request, "hod_template/admin_view_profile.html", context)


# =============================================================================
# ATTENDANCE VIEW
# =============================================================================

@login_required
def admin_view_attendance(request):
    """View attendance reports"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    assignments = Course_Assignment.objects.select_related('course', 'faculty').filter(is_active=True)
    academic_years = AcademicYear.objects.all()
    
    context = {
        'assignments': assignments,
        'academic_years': academic_years,
        'page_title': 'View Attendance'
    }
    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def get_admin_attendance(request):
    """API to get attendance data"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    assignment_id = request.POST.get('assignment')
    date = request.POST.get('date')
    
    try:
        assignment = get_object_or_404(Course_Assignment, id=assignment_id)
        attendance_records = Attendance.objects.filter(assignment=assignment)
        
        if date:
            attendance_records = attendance_records.filter(date=date)
        
        json_data = []
        for record in attendance_records:
            data = {
                "status": record.status,
                "name": record.student.user.full_name,
                "register_no": record.student.register_no,
                "date": str(record.date),
                "period": record.period
            }
            json_data.append(data)
        
        return JsonResponse(json.dumps(json_data), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# =============================================================================
# QUESTION PAPER MANAGEMENT
# =============================================================================

@login_required
def assign_question_paper(request):
    """Assign question paper setting task to faculty"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    form = QuestionPaperAssignmentForm(request.POST or None)
    
    # Get recent assignments
    recent_assignments = QuestionPaperAssignment.objects.all().order_by('-created_at')[:10]
    
    context = {
        'form': form,
        'recent_assignments': recent_assignments,
        'page_title': 'Assign Question Paper Task'
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                qp_assignment = form.save(commit=False)
                qp_assignment.assigned_by = request.user
                qp_assignment.save()
                
                # Create notification for assigned faculty
                Notification.objects.create(
                    recipient=qp_assignment.assigned_faculty.user,
                    title='Question Paper Assignment',
                    message=f'You have been assigned to set {qp_assignment.get_exam_type_display()} question paper for {qp_assignment.course.course_code} - {qp_assignment.course.title}. Deadline: {qp_assignment.deadline}',
                    notification_type='ANNOUNCEMENT',
                    link=reverse('staff_view_qp_assignments')
                )
                
                messages.success(request, f"Question paper task assigned to {qp_assignment.assigned_faculty.user.full_name} successfully!")
                return redirect(reverse('manage_qp_assignments'))
            except Exception as e:
                messages.error(request, f"Could not assign: {str(e)}")
        else:
            messages.error(request, "Please correct the errors in the form")
    
    return render(request, 'hod_template/assign_question_paper.html', context)


@login_required
def manage_qp_assignments(request):
    """View and manage all question paper assignments"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    # Filter options
    status_filter = request.GET.get('status', '')
    exam_type_filter = request.GET.get('exam_type', '')
    
    assignments = QuestionPaperAssignment.objects.all().select_related(
        'course', 'assigned_faculty__user', 'academic_year', 'semester'
    ).prefetch_related('structured_qp').order_by('-created_at')
    
    if status_filter:
        assignments = assignments.filter(status=status_filter)
    if exam_type_filter:
        assignments = assignments.filter(exam_type=exam_type_filter)
    
    # Statistics
    stats = {
        'total': QuestionPaperAssignment.objects.count(),
        'assigned': QuestionPaperAssignment.objects.filter(status='ASSIGNED').count(),
        'submitted': QuestionPaperAssignment.objects.filter(status='SUBMITTED').count(),
        'approved': QuestionPaperAssignment.objects.filter(status='APPROVED').count(),
        'pending_review': QuestionPaperAssignment.objects.filter(status__in=['SUBMITTED', 'UNDER_REVIEW']).count(),
    }
    
    context = {
        'assignments': assignments,
        'stats': stats,
        'status_choices': QuestionPaperAssignment.STATUS_CHOICES,
        'exam_type_choices': QuestionPaperAssignment.EXAM_TYPE_CHOICES,
        'current_status': status_filter,
        'current_exam_type': exam_type_filter,
        'page_title': 'Manage Question Paper Assignments'
    }
    return render(request, 'hod_template/manage_qp_assignments.html', context)


@login_required
def review_question_paper(request, qp_id):
    """Review submitted question paper"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    qp_assignment = get_object_or_404(QuestionPaperAssignment, id=qp_id)
    form = QuestionPaperReviewForm(request.POST or None, instance=qp_assignment)
    
    context = {
        'qp_assignment': qp_assignment,
        'form': form,
        'page_title': f'Review QP - {qp_assignment.course.course_code}'
    }
    
    if request.method == 'POST':
        if form.is_valid():
            try:
                qp = form.save(commit=False)
                qp.reviewed_by = request.user
                qp.reviewed_at = datetime.now()
                qp.save()
                
                # Notify faculty about review result
                status_text = qp.get_status_display()
                Notification.objects.create(
                    recipient=qp.assigned_faculty.user,
                    title=f'Question Paper Review - {status_text}',
                    message=f'Your question paper for {qp.course.course_code} has been reviewed. Status: {status_text}. Comments: {qp.review_comments or "None"}',
                    notification_type='INFO',
                    link=reverse('staff_view_qp_assignments')
                )
                
                messages.success(request, f"Review submitted. Status: {status_text}")
                return redirect(reverse('manage_qp_assignments'))
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            messages.error(request, "Please correct the errors")
    
    return render(request, 'hod_template/review_question_paper.html', context)


@login_required
def delete_qp_assignment(request, qp_id):
    """Delete a question paper assignment"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    try:
        qp_assignment = get_object_or_404(QuestionPaperAssignment, id=qp_id)
        qp_assignment.delete()
        messages.success(request, "Question paper assignment deleted successfully")
    except Exception as e:
        messages.error(request, f"Could not delete: {str(e)}")
    
    return redirect(reverse('manage_qp_assignments'))


@csrf_exempt
@login_required
def get_faculty_for_course(request):
    """AJAX: Get faculty who teach a specific course"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    course_id = request.POST.get('course_id')
    
    try:
        # Get faculty who have course assignments for this course
        course_assignments = Course_Assignment.objects.filter(
            course_id=course_id, 
            is_active=True
        ).select_related('faculty__user')
        
        faculty_list = []
        seen_ids = set()
        
        for ca in course_assignments:
            if ca.faculty.id not in seen_ids:
                faculty_list.append({
                    'id': ca.faculty.id,
                    'name': ca.faculty.user.full_name,
                    'staff_id': ca.faculty.staff_id
                })
                seen_ids.add(ca.faculty.id)
        
        # If no specific faculty found, return all active faculty
        if not faculty_list:
            all_faculty = Faculty_Profile.objects.filter(user__is_active=True).select_related('user')
            for f in all_faculty:
                faculty_list.append({
                    'id': f.id,
                    'name': f.user.full_name,
                    'staff_id': f.staff_id
                })
        
        return JsonResponse({'faculty': faculty_list})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# =============================================================================
# UTILITY VIEWS
# =============================================================================

@csrf_exempt
def check_email_availability(request):
    """Check if email is available"""
    email = request.POST.get("email")
    try:
        exists = Account_User.objects.filter(email=email).exists()
        return HttpResponse(exists)
    except:
        return HttpResponse(False)


# Aliases for backward compatibility
add_staff = add_faculty
manage_staff = manage_faculty
edit_staff = edit_faculty
delete_staff = delete_faculty
add_session = add_academic_year
manage_session = manage_academic_year

student_feedback_message = view_feedbacks
staff_feedback_message = view_feedbacks
view_staff_leave = view_leave_requests
view_student_leave = view_leave_requests
admin_notify_staff = send_notification_page
admin_notify_student = send_notification_page
send_student_notification = send_notification
send_staff_notification = send_notification


# =============================================================================
# TIMETABLE MANAGEMENT
# =============================================================================

@login_required
def manage_timetables(request):
    """List all timetables with filtering options"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    from django.db.models import Count
    timetables = Timetable.objects.all().select_related(
        'academic_year', 'semester', 'regulation', 'created_by',
        'program_batch', 'program_batch__program'
    ).annotate(
        entry_count=Count('entries')
    ).order_by('-academic_year', 'year', 'batch')
    
    # Filtering
    year_filter = request.GET.get('year')
    batch_filter = request.GET.get('batch')
    academic_year_filter = request.GET.get('academic_year')
    
    if year_filter:
        timetables = timetables.filter(year=year_filter)
    if batch_filter:
        timetables = timetables.filter(batch=batch_filter)
    if academic_year_filter:
        timetables = timetables.filter(academic_year_id=academic_year_filter)
    
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    # Get batch choices from database
    current_year = AcademicYear.get_current()
    if current_year:
        batch_choices = list(ProgramBatch.objects.filter(
            academic_year=current_year
        ).values_list('batch_name', 'batch_display').distinct())
    else:
        batch_choices = []
    
    context = {
        'timetables': timetables,
        'academic_years': academic_years,
        'year_choices': Timetable.YEAR_CHOICES,
        'batch_choices': batch_choices,
        'page_title': 'Manage Timetables'
    }
    return render(request, 'hod_template/manage_timetables.html', context)


@login_required
def add_timetable(request):
    """Redirect to the Create Timetable Wizard"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    return redirect('create_timetable_wizard')


@login_required
def edit_timetable(request, timetable_id):
    """Edit timetable - main grid view for entering schedule"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    timetable = get_object_or_404(
        Timetable.objects.select_related('program_batch', 'program_batch__program'),
        id=timetable_id
    )
    
    # Ensure time slots exist, create defaults if not
    if TimeSlot.objects.count() == 0:
        create_default_time_slots()
    
    time_slots = TimeSlot.objects.all().order_by('slot_number')
    days = TimetableEntry.DAY_CHOICES
    
    # Get existing entries
    entries = TimetableEntry.objects.filter(timetable=timetable).select_related(
        'course', 'faculty__user', 'time_slot', 'lab_room'
    )
    
    # Create lookup dictionary for entries
    entry_lookup = {}
    for entry in entries:
        key = f"{entry.day}_{entry.time_slot.slot_number}"
        entry_lookup[key] = entry
    
    # Get courses for this year's semester
    course_semesters = []
    if timetable.year == 1:
        course_semesters = [1, 2]
    elif timetable.year == 2:
        course_semesters = [3, 4]
    elif timetable.year == 3:
        course_semesters = [5, 6]
    elif timetable.year == 4:
        course_semesters = [7, 8]
    
    courses = Course.objects.filter(semester__in=course_semesters).order_by('course_code')
    faculty_list = Faculty_Profile.objects.filter(user__is_active=True).select_related('user').order_by('user__full_name')
    
    context = {
        'timetable': timetable,
        'time_slots': time_slots,
        'days': days,
        'entry_lookup': entry_lookup,
        'courses': courses,
        'faculty_list': faculty_list,
        'page_title': f'Edit Timetable - {timetable.program_batch.program.code + " " if timetable.program_batch and timetable.program_batch.program else ""}Year {timetable.year} Batch {timetable.batch_display}'
    }
    return render(request, 'hod_template/edit_timetable.html', context)


@login_required
@csrf_exempt
def save_timetable_entry(request):
    """AJAX endpoint to save a single timetable entry"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            timetable_id = data.get('timetable_id')
            day = data.get('day')
            slot_number = data.get('slot_number')
            course_code = data.get('course_code')
            faculty_id = data.get('faculty_id')
            special_note = data.get('special_note', '')
            
            timetable = get_object_or_404(Timetable, id=timetable_id)
            time_slot = get_object_or_404(TimeSlot, slot_number=slot_number)
            
            # Get or create entry
            entry, created = TimetableEntry.objects.get_or_create(
                timetable=timetable,
                day=day,
                time_slot=time_slot,
                defaults={'special_note': special_note}
            )
            
            # Update entry
            if course_code:
                entry.course = Course.objects.filter(course_code=course_code).first()
            else:
                entry.course = None
            
            if faculty_id:
                entry.faculty = Faculty_Profile.objects.filter(id=faculty_id).first()
            else:
                entry.faculty = None
            
            entry.special_note = special_note
            entry.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Entry saved',
                'entry_id': entry.id,
                'display_text': entry.display_text,
                'faculty_name': entry.faculty_name
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
@csrf_exempt
def delete_timetable_entry(request):
    """AJAX endpoint to delete a timetable entry"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            timetable_id = data.get('timetable_id')
            day = data.get('day')
            slot_number = data.get('slot_number')
            
            TimetableEntry.objects.filter(
                timetable_id=timetable_id,
                day=day,
                time_slot__slot_number=slot_number
            ).delete()
            
            return JsonResponse({'success': True, 'message': 'Entry deleted'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def delete_timetable(request, timetable_id):
    """Delete a timetable"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    timetable = get_object_or_404(Timetable, id=timetable_id)
    timetable.delete()
    messages.success(request, "Timetable deleted successfully.")
    return redirect('manage_timetables')


@login_required
def view_timetable(request, timetable_id):
    """View a timetable (read-only)"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    timetable = get_object_or_404(
        Timetable.objects.select_related('program_batch', 'program_batch__program', 'academic_year', 'semester'),
        id=timetable_id
    )
    
    time_slots = TimeSlot.objects.all().order_by('slot_number')
    days = TimetableEntry.DAY_CHOICES
    
    entries = TimetableEntry.objects.filter(timetable=timetable).select_related(
        'course', 'faculty__user', 'lab_assistant__user', 'time_slot'
    )

    assignment_qs = Course_Assignment.objects.filter(
        academic_year=timetable.academic_year,
        semester=timetable.semester,
        is_active=True,
    ).select_related('course', 'faculty__user', 'lab_assistant__user', 'lab_main_faculty__user')
    if timetable.program_batch_id:
        assignment_qs = assignment_qs.filter(batch=timetable.program_batch)
    else:
        assignment_qs = assignment_qs.filter(
            batch__batch_name=timetable.batch,
            batch__year_of_study=timetable.year,
        )
    assignment_by_course_id = {a.course_id: a for a in assignment_qs}

    def _lit_split_faculty_display(entry):
        if not entry.course or entry.course.course_type != 'LIT':
            return None

        assignment = assignment_by_course_id.get(entry.course_id)
        if not assignment or not assignment.separate_lab_theory_staff:
            return None

        theory_name = assignment.faculty.user.full_name if assignment.faculty and assignment.faculty.user else 'TBD'
        lab_main = assignment.effective_lab_main_faculty
        lab_main_name = lab_main.user.full_name if lab_main and lab_main.user else 'TBD'
        label = f"{theory_name}/{lab_main_name}(lab main)"

        if assignment.lab_assistant and assignment.lab_assistant.user:
            label = f"{label} & {assignment.lab_assistant.user.full_name}"

        return label

    # Compute IOC course codes relevant to this timetable so UI can render "IOC" label.
    quick_ioc_qs = Course_Assignment.objects.filter(
        semester=timetable.semester,
        academic_year=timetable.academic_year,
        special_note__icontains='Quick IOC',
        is_active=True,
    )
    if timetable.program_batch_id:
        quick_ioc_qs = quick_ioc_qs.filter(batch=timetable.program_batch)
    else:
        quick_ioc_qs = quick_ioc_qs.filter(batch__batch_name=timetable.batch, batch__year_of_study=timetable.year)

    ioc_course_codes = set(quick_ioc_qs.values_list('course__course_code', flat=True))
    
    # ── Resolve placeholder courses to actual offering codes ──
    # Collect unique placeholder course codes that need resolution
    placeholder_codes_to_resolve = set()
    for entry in entries:
        if entry.course and getattr(entry.course, 'is_placeholder', False):
            if entry.special_note == 'Same-time':
                # Legacy entries: need to resolve from offerings
                placeholder_codes_to_resolve.add(entry.course.course_code)

    # Batch-resolve all placeholder → actual course codes
    resolved_codes_map = {}  # placeholder_code -> "CS23401/CS23301/..."
    current_program = timetable.program_batch.program if timetable.program_batch else None
    if current_program:
        ioc_course_codes.update(
            RegulationCoursePlan.objects.filter(
                semester=timetable.semester.semester_number,
                branch=current_program.code,
                program_type=current_program.level,
            ).filter(
                Q(category__code__in=['IOC', 'EEC']) |
                Q(course__placeholder_type__in=['IOC', 'EEC'])
            ).values_list('course__course_code', flat=True)
        )

    for pc in placeholder_codes_to_resolve:
        plan_filters = {
            'course__course_code': pc,
            'semester': timetable.semester.semester_number,
        }
        if current_program:
            plan_filters['branch'] = current_program.code
            plan_filters['program_type'] = current_program.level
        if timetable.regulation:
            plan_filters['regulation'] = timetable.regulation

        rcp_ids = RegulationCoursePlan.objects.filter(**plan_filters).values_list('id', flat=True)
        offerings = ElectiveCourseOffering.objects.filter(
            regulation_course_plan_id__in=rcp_ids,
            semester=timetable.semester,
            is_active=True,
        ).select_related('actual_course').order_by('actual_course__course_code')
        codes = [o.actual_course.course_code for o in offerings]
        resolved_codes_map[pc] = '/'.join(codes) if codes else pc

    # Build lookup for converting course-code lists in notes to title lists.
    # Example: "G1: CS23016/CS23045 LAB1" -> "G1: Devops / Image Processing LAB1"
    code_token_pattern = re.compile(r'\b[A-Z]{2,}\d[A-Z0-9]*\b')
    all_note_codes = set()
    for entry in entries:
        if entry.special_note:
            all_note_codes.update(code_token_pattern.findall(entry.special_note))
    for codes_text in resolved_codes_map.values():
        all_note_codes.update(code_token_pattern.findall(codes_text))

    code_title_lookup = dict(
        Course.objects.filter(course_code__in=all_note_codes).values_list('course_code', 'title')
    )

    def _render_title_from_code_list(text):
        if not text:
            return text

        value = text.strip()
        prefix = ''
        if ':' in value:
            maybe_prefix, rest = value.split(':', 1)
            maybe_prefix = maybe_prefix.strip()
            if maybe_prefix.upper().startswith('G') and maybe_prefix[1:].isdigit():
                prefix = f"{maybe_prefix}: "
                value = rest.strip()

        lab_suffix = ''
        lab_idx = value.upper().find(' LAB')
        if lab_idx != -1:
            lab_suffix = value[lab_idx:]
            value = value[:lab_idx].strip()

        parts = [p.strip() for p in value.split('/') if p.strip()]
        if not parts:
            return text

        title_parts = [code_title_lookup.get(p, p) for p in parts]
        rendered = ' / '.join(title_parts)
        if lab_suffix:
            rendered = f"{rendered}{lab_suffix}"
        return f"{prefix}{rendered}" if prefix else rendered

    entry_lookup = {}
    for entry in entries:
        if entry.course:
            is_ioc = (
                entry.course.course_code in ioc_course_codes or
                (entry.course.is_placeholder and entry.course.placeholder_type in ['IOC', 'EEC'])
            )
            entry.display_code = 'IOC' if is_ioc else entry.course.course_code
        else:
            entry.display_code = None

        # Set resolved_display for template rendering
        if entry.course and getattr(entry.course, 'is_placeholder', False):
            if entry.special_note == 'Same-time':
                # Legacy: resolve from offerings
                entry.resolved_display = resolved_codes_map.get(
                    entry.course.course_code, entry.course.course_code
                )
            elif entry.special_note and entry.special_note != 'Same-time':
                # New format: special_note already contains resolved codes
                entry.resolved_display = entry.special_note
            else:
                entry.resolved_display = entry.course.course_code
        else:
            entry.resolved_display = None

        entry.faculty_display_name = _lit_split_faculty_display(entry)
        key = f"{entry.day}_{entry.time_slot.slot_number}"
        entry_lookup[key] = entry

    # Build course-details legend directly from timetable entries so placeholder
    # rows (IOC/SDC/AC/etc.) are visible exactly as scheduled.
    course_rows = {}

    def _collect_entry_faculty_labels(entry):
        labels = []
        split_display = getattr(entry, 'faculty_display_name', None)
        if split_display:
            labels.append(split_display)
        elif entry.faculty and entry.faculty.user:
            labels.append(entry.faculty.user.full_name)

        needs_lab_faculty = bool(
            entry.course and (
                entry.course.course_type in ['L', 'LIT']
                or (entry.course.practical_hours or 0) > 0
                or entry.is_lab
            )
        )
        if needs_lab_faculty and entry.lab_assistant and entry.lab_assistant.user:
            labels.append(entry.lab_assistant.user.full_name)

        if not labels:
            labels.append('TBD')
        return labels

    for entry in entries:
        if not entry.course or entry.is_blocked:
            continue

        course = entry.course
        is_placeholder = bool(course.is_placeholder)
        course_code = course.course_code
        course_title = course.title

        if is_placeholder and entry.special_note == 'Same-time' and entry.resolved_display:
            rendered_title = _render_title_from_code_list(entry.resolved_display)
            if rendered_title:
                course_title = rendered_title

        if is_placeholder:
            category = course.placeholder_type or 'Placeholder'
        else:
            category = course.course_type or 'T'

        row_key = (course_code, course_title, category, is_placeholder)
        row = course_rows.setdefault(row_key, {
            'course_code': course_code,
            'course_title': course_title,
            'faculty_names': set(),
            'category': category,
            'is_placeholder': is_placeholder,
        })

        for name in _collect_entry_faculty_labels(entry):
            row['faculty_names'].add(name)

    course_details = []
    for row in course_rows.values():
        faculty_names = sorted([n for n in row['faculty_names'] if n])
        if not faculty_names:
            faculty_names = ['TBD']

        course_details.append({
            'course_code': row['course_code'],
            'course_title': row['course_title'],
            'faculty_name': ' / '.join(faculty_names),
            'category': row['category'],
            'is_placeholder': row['is_placeholder'],
        })

    course_details.sort(key=lambda x: (x['course_code'], x['course_title']))
    
    program_label = ''
    if timetable.program_batch and timetable.program_batch.program:
        program_label = f'{timetable.program_batch.program.code} '
    
    context = {
        'timetable': timetable,
        'time_slots': time_slots,
        'days': days,
        'entry_lookup': entry_lookup,
        'course_details': course_details,
        'page_title': f'Timetable - {program_label}Year {timetable.year} Batch {timetable.batch_display}'
    }
    return render(request, 'hod_template/view_timetable.html', context)


@login_required
def manage_time_slots(request):
    """Manage time slots configuration"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    time_slots = TimeSlot.objects.all().order_by('slot_number')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_defaults':
            create_default_time_slots()
            messages.success(request, "Default time slots created successfully.")
            return redirect('manage_time_slots')
        
        elif action == 'add':
            form = TimeSlotForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Time slot added successfully.")
                return redirect('manage_time_slots')
            else:
                messages.error(request, "Error adding time slot.")
        
        elif action == 'delete':
            slot_id = request.POST.get('slot_id')
            TimeSlot.objects.filter(id=slot_id).delete()
            messages.success(request, "Time slot deleted.")
            return redirect('manage_time_slots')
    
    form = TimeSlotForm()
    
    context = {
        'time_slots': time_slots,
        'form': form,
        'page_title': 'Manage Time Slots'
    }
    return render(request, 'hod_template/manage_time_slots.html', context)


def create_default_time_slots():
    """Create default time slots based on the provided schedule"""
    default_slots = [
        (1, '08:30', '09:20', False),
        (2, '09:25', '10:15', False),
        (3, '10:30', '11:20', False),
        (4, '11:25', '12:15', False),
        # Lunch break (slot 5 could be implicit or we skip)
        (5, '13:10', '14:00', False),
        (6, '14:05', '14:55', False),
        (7, '15:00', '15:50', False),
        (8, '15:55', '16:45', False),
    ]
    
    for slot_num, start, end, is_break in default_slots:
        TimeSlot.objects.get_or_create(
            slot_number=slot_num,
            defaults={
                'start_time': start,
                'end_time': end,
                'is_break': is_break
            }
        )


@login_required
@csrf_exempt
def get_courses_for_semester(request):
    """AJAX endpoint to get courses for a specific year/semester"""
    if request.method == 'GET':
        year = request.GET.get('year')
        
        try:
            year = int(year)
            # Map year to semesters
            course_semesters = []
            if year == 1:
                course_semesters = [1, 2]
            elif year == 2:
                course_semesters = [3, 4]
            elif year == 3:
                course_semesters = [5, 6]
            elif year == 4:
                course_semesters = [7, 8]
            
            courses = Course.objects.filter(semester__in=course_semesters).order_by('course_code')
            course_list = [{'code': c.course_code, 'title': c.title} for c in courses]
            
            return JsonResponse({'courses': course_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
@csrf_exempt  
def get_all_faculty(request):
    """AJAX endpoint to get all active faculty with search support"""
    if request.method == 'GET':
        search = request.GET.get('search', '')
        
        faculty_qs = Faculty_Profile.objects.filter(user__is_active=True).select_related('user')
        
        if search:
            faculty_qs = faculty_qs.filter(
                Q(user__full_name__icontains=search) | 
                Q(staff_id__icontains=search)
            )
        
        faculty_list = [{
            'id': f.id,
            'name': f.user.full_name,
            'staff_id': f.staff_id,
            'designation': f.get_designation_display()
        } for f in faculty_qs.order_by('user__full_name')[:50]]
        
        return JsonResponse({'faculty': faculty_list})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


# =============================================================================
# BULK STUDENT UPLOAD
# =============================================================================

@login_required
def bulk_upload_students(request):
    """Bulk upload students via CSV file"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Get current academic year and configured batches
    current_year = AcademicYear.get_current()
    
    configured_batches = {}  # {program_code: [batch_names]}
    programs_without_batches = []
    
    if current_year:
        # Get all programs and their configured batches for Year 1
        for program in Program.objects.all():
            batches = ProgramBatch.objects.filter(
                academic_year=current_year,
                program=program,
                year_of_study=1,
                is_active=True
            ).values_list('batch_name', flat=True)
            if batches:
                configured_batches[program.code] = list(batches)
            else:
                programs_without_batches.append(program.code)
    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        
        if not csv_file:
            messages.error(request, "Please upload a CSV file")
            return redirect('bulk_upload_students')
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "Please upload a valid CSV file")
            return redirect('bulk_upload_students')
        
        try:
            # Read CSV file (handle BOM from Excel)
            file_content = csv_file.read()
            # Try UTF-8 with BOM first, then regular UTF-8
            try:
                decoded_file = file_content.decode('utf-8-sig')  # Handles BOM
            except:
                decoded_file = file_content.decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            success_count = 0
            error_count = 0
            errors = []
            created_students = []
            
            for row_num, row in enumerate(reader, start=2):  # Start from 2 (header is row 1)
                try:
                    # Validate required fields
                    register_no = row.get('register_no', '').strip()
                    
                    # Handle Excel scientific notation (e.g., 1.24E+09 -> 1240000000)
                    if 'E' in register_no.upper() or 'e' in register_no:
                        try:
                            register_no = str(int(float(register_no)))
                        except ValueError:
                            pass
                    
                    full_name = row.get('full_name', '').strip()
                    email = row.get('email', '').strip().lower()
                    gender = row.get('gender', '').strip().upper()
                    batch = row.get('batch', '').strip().upper()
                    branch = row.get('branch', '').strip().upper()
                    program = row.get('program', 'UG').strip().upper()
                    entry_type = row.get('entry_type', 'REGULAR').strip().upper()
                    
                    # Optional fields
                    phone = row.get('phone', '').strip()
                    parent_name = row.get('parent_name', '').strip()
                    parent_phone = row.get('parent_phone', '').strip()
                    address = row.get('address', '').strip()
                    
                    # Admission year - optional, defaults based on academic year context
                    from datetime import datetime
                    admission_year_str = row.get('admission_year', '').strip()
                    if admission_year_str:
                        try:
                            admission_year = int(admission_year_str)
                            # Validate reasonable year range (2015 to current+1)
                            current_cal_year = datetime.now().year
                            if admission_year < 2015 or admission_year > current_cal_year + 1:
                                errors.append(f"Row {row_num}: admission_year must be between 2015 and {current_cal_year + 1}")
                                error_count += 1
                                continue
                        except ValueError:
                            errors.append(f"Row {row_num}: admission_year must be a valid year (got '{admission_year_str}')")
                            error_count += 1
                            continue
                    else:
                        # Default admission_year based on academic year context
                        # Academic year starts in Aug, so:
                        # - If Jan-Jul (before new batch joins): current 1st years are from previous year
                        # - If Aug-Dec (after new batch joins): current 1st years are from this year
                        today = datetime.now()
                        if today.month >= 8:  # Aug-Dec: new batch has joined
                            admission_year = today.year
                        else:  # Jan-Jul: 1st years are from previous year's Aug
                            admission_year = today.year - 1
                    
                    # Current semester - optional, if provided use it, else calculate from database
                    current_sem_str = row.get('current_sem', '').strip()
                    if current_sem_str:
                        try:
                            current_sem = int(current_sem_str)
                            # Validate semester range
                            if current_sem < 1 or current_sem > 8:
                                errors.append(f"Row {row_num}: current_sem must be between 1 and 8 (got {current_sem})")
                                error_count += 1
                                continue
                            # Validate lateral entry starts from sem 3
                            if entry_type == 'LATERAL' and current_sem < 3:
                                errors.append(f"Row {row_num}: LATERAL entry students must be in semester 3 or higher (got {current_sem})")
                                error_count += 1
                                continue
                        except ValueError:
                            errors.append(f"Row {row_num}: current_sem must be a valid number (got '{current_sem_str}')")
                            error_count += 1
                            continue
                    else:
                        # Auto-calculate semester from database config
                        # Check if there's a current semester configured in database
                        db_current_sem = Semester.get_current()
                        
                        if db_current_sem:
                            # Use database semester config to determine odd/even
                            is_odd_semester = db_current_sem.semester_number % 2 == 1
                            years_passed = datetime.now().year - admission_year
                            
                            # Adjust for academic year (odd sem starts in Aug of admission year)
                            # If in odd sem: student in year N is in sem (N*2 - 1)
                            # If in even sem: student in year N is in sem (N*2)
                            if is_odd_semester:
                                semester_offset = years_passed * 2 + 1
                            else:
                                # Even semester: we're in the next calendar year
                                semester_offset = years_passed * 2
                        else:
                            # Fallback: use month-based calculation if no semester configured
                            today = datetime.now()
                            current_month = today.month
                            years_passed = today.year - admission_year
                            
                            if current_month >= 8:  # Aug onwards = Odd semester
                                semester_offset = years_passed * 2 + 1
                            elif current_month <= 5:  # Jan-May = Even semester
                                semester_offset = years_passed * 2
                            else:  # June-July = Summer break
                                semester_offset = years_passed * 2
                        
                        # For lateral entry, add 2 (they skip sem 1 & 2)
                        if entry_type == 'LATERAL':
                            semester_offset += 2
                        
                        # Clamp to valid range (1-8 for B.E.)
                        current_sem = max(1, min(semester_offset, 8))
                    
                    # Validate entry_type
                    if entry_type not in ['REGULAR', 'LATERAL']:
                        errors.append(f"Row {row_num}: entry_type must be REGULAR or LATERAL (got '{entry_type}')")
                        error_count += 1
                        continue
                    
                    # Validation
                    if not all([register_no, full_name, email, gender, batch, branch]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        error_count += 1
                        continue
                    
                    if len(register_no) != 10 or not register_no.isdigit():
                        errors.append(f"Row {row_num}: Register number must be 10 digits (got '{register_no}' with length {len(register_no)})")
                        error_count += 1
                        continue
                    
                    if gender not in ['M', 'F', 'O']:
                        errors.append(f"Row {row_num}: Gender must be M, F, or O")
                        error_count += 1
                        continue
                    
                    # Validate branch exists in database
                    # Note: We validate branch exists as a Program, the regulation mapping is handled separately
                    all_branches = list(Program.objects.values_list('code', flat=True).distinct())
                    if branch not in all_branches:
                        errors.append(f"Row {row_num}: Branch '{branch}' not found. Valid options: {', '.join(all_branches)}")
                        error_count += 1
                        continue
                    
                    # Calculate year of study from current_sem for batch validation
                    year_of_study = (current_sem + 1) // 2
                    
                    # Get batches for this branch and year of study
                    branch_batches_for_year = ProgramBatch.objects.filter(
                        academic_year=current_year,
                        program__code=branch,
                        year_of_study=year_of_study,
                        is_active=True
                    ).values_list('batch_name', flat=True)
                    valid_batches = list(branch_batches_for_year)
                    
                    # Check if batches are configured for this branch and year
                    if not valid_batches:
                        # Fallback: check if ANY batches exist for this branch
                        any_batches = ProgramBatch.objects.filter(
                            academic_year=current_year,
                            program__code=branch,
                            is_active=True
                        ).values_list('batch_name', 'year_of_study')
                        if any_batches:
                            available = [f"{b[0]} (Year {b[1]})" for b in any_batches]
                            errors.append(f"Row {row_num}: No batches configured for {branch} Year {year_of_study}. Available: {', '.join(available)}")
                        else:
                            errors.append(f"Row {row_num}: No batches configured for {branch}. Please configure batches in 'Manage Batches' first.")
                        error_count += 1
                        continue
                    
                    # Validate batch against configured batches for this branch and year
                    if batch not in valid_batches:
                        errors.append(f"Row {row_num}: Batch '{batch}' not valid for {branch} Year {year_of_study}. Configured batches: {', '.join(valid_batches)}")
                        error_count += 1
                        continue
                    
                    # Map program
                    program_map = {'B.E': 'UG', 'BE': 'UG', 'UG': 'UG', 'M.E': 'PG', 'ME': 'PG', 'PG': 'PG', 'PH.D': 'PHD', 'PHD': 'PHD'}
                    program_type = program_map.get(program, 'UG')
                    
                    # Check duplicates
                    if Account_User.objects.filter(email=email).exists():
                        errors.append(f"Row {row_num}: Email {email} already exists")
                        error_count += 1
                        continue
                    
                    if Student_Profile.objects.filter(register_no=register_no).exists():
                        errors.append(f"Row {row_num}: Register number {register_no} already exists")
                        error_count += 1
                        continue
                    
                    # Use transaction to rollback if any step fails
                    with transaction.atomic():
                        # Create user with unusable password (forces password setup)
                        user = Account_User.objects.create(
                            email=email,
                            full_name=full_name,
                            gender=gender,
                            phone=phone or None,
                            address=address or None,
                            role='STUDENT',
                            is_active=True
                        )
                        user.set_unusable_password()  # User must set password via email
                        user.save()
                        
                        # Update student profile (auto-created by signal)
                        student = user.student_profile
                        student.register_no = register_no
                        student.batch_label = batch
                        student.branch = branch
                        student.program_type = program_type
                        student.entry_type = entry_type
                        student.admission_year = admission_year
                        student.current_sem = current_sem
                        student.parent_name = parent_name or None
                        student.parent_phone = parent_phone or None
                        
                        # Auto-assign regulation using ProgramRegulation mapping
                        # REGULAR: joins as 1st year → follows regulation of their admission year
                        # LATERAL: joins as 2nd year → follows regulation of batch that started 1 year earlier
                        regulation_lookup_year = admission_year if entry_type == 'REGULAR' else (admission_year - 1)
                        
                        # Use ProgramRegulation to find the correct regulation
                        regulation = ProgramRegulation.get_regulation_for_student(
                            program_code=branch,
                            program_level=program_type,
                            admission_year=regulation_lookup_year
                        )
                        
                        # Fallback to old method if no ProgramRegulation mapping exists
                        if not regulation:
                            regulation = Regulation.objects.filter(
                                year__lte=regulation_lookup_year
                            ).order_by('-year').first()
                        
                        if regulation:
                            student.regulation = regulation
                        
                        student.save()
                        
                        created_students.append({
                            'user': user,
                            'email': email,
                            'name': full_name,
                            'college_email': student.college_email,
                            'register_no': register_no
                        })
                        success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            # Send OTP emails to college email for first-time login
            email_success = 0
            email_failed = 0
            for student_data in created_students:
                try:
                    send_first_login_notification(student_data)
                    email_success += 1
                except Exception as e:
                    email_failed += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f"Successfully created {success_count} students. Login instructions sent to college emails: {email_success}")
            if error_count > 0:
                messages.warning(request, f"Failed to create {error_count} students. See details below.")
            
            context = {
                'page_title': 'Bulk Upload Results',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:20],  # Show first 20 errors
                'total_errors': len(errors),
                'email_success': email_success,
                'email_failed': email_failed,
            }
            return render(request, 'hod_template/bulk_upload_results.html', context)
            
        except Exception as e:
            messages.error(request, f"Error processing CSV file: {str(e)}")
            return redirect('bulk_upload_students')
    
    # Get available regulations for display
    all_regulations = Regulation.objects.all().order_by('-year')
    
    context = {
        'page_title': 'Bulk Upload Students',
        'current_year': current_year,
        'configured_batches': configured_batches,
        'programs_without_batches': programs_without_batches,
        'all_regulations': all_regulations,
    }
    return render(request, 'hod_template/bulk_upload_students.html', context)


@login_required
def download_student_template(request):
    """Download CSV template for bulk student upload"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_upload_template.csv"'
    
    # Get configured programs and batches for sample data
    programs = list(Program.objects.values_list('code', flat=True)[:2])
    current_year = AcademicYear.get_current()
    batch_examples = ['A', 'B']  # Default examples
    if current_year:
        batches = ProgramBatch.objects.filter(
            academic_year=current_year,
            year_of_study=1,
            is_active=True
        ).values_list('batch_name', flat=True).distinct()[:2]
        if batches:
            batch_examples = list(batches)
    
    # Use actual program codes or placeholders
    program1 = programs[0] if len(programs) > 0 else 'CSE'
    program2 = programs[1] if len(programs) > 1 else 'IT'
    batch1 = batch_examples[0] if batch_examples else 'A'
    batch2 = batch_examples[1] if len(batch_examples) > 1 else 'B'
    
    writer = csv.writer(response)
    # Header row
    writer.writerow([
        'register_no', 'full_name', 'email', 'gender', 'batch', 'branch', 
        'program', 'entry_type', 'admission_year', 'current_sem', 'phone', 'parent_name', 
        'parent_phone', 'address'
    ])
    # Sample data rows - showing different scenarios
    from datetime import datetime
    curr_year = datetime.now().year
    writer.writerow([
        '2023105001', 'Senior (Explicit Sem)', 'senior@student.edu', 'M', batch1, program1,
        'B.E', 'REGULAR', 2023, 6, '9876543210', 'Mr. Doe', '9876543211', 'Chennai'
    ])
    writer.writerow([
        '2024105002', 'Junior (Auto Sem)', 'junior@student.edu', 'F', batch2, program2,
        'B.E', 'REGULAR', 2024, '', '9876543212', 'Mrs. Smith', '9876543213', 'Coimbatore'  # Empty = auto-calculate
    ])
    writer.writerow([
        '2026105003', 'New Student', 'new@student.edu', 'M', batch1, program1,
        'B.E', 'REGULAR', '', '', '9876543214', '', '', ''  # Both empty = current year, auto-calc sem
    ])
    writer.writerow([
        '2025105004', 'Lateral Entry', 'lateral@student.edu', 'F', batch2, program2,
        'B.E', 'LATERAL', 2025, 4, '', '', '', ''  # Lateral must be sem 3+
    ])
    
    return response


def send_first_login_notification(student_data):
    """
    Send first-time login notification to student's college email.
    The college email is auto-generated: <register_no>@student.annauniv.edu
    """
    try:
        subject = 'Welcome to CSE Department ERP - First Time Login Instructions'
        message = f"""
Dear {student_data['name']},

Your account has been created in the CSE Department ERP System.

Register Number: {student_data['register_no']}
Personal Email (for login): {student_data['email']}
College Email: {student_data['college_email']}

To set your password, please follow these steps:

1. Visit the ERP portal and click on "First Time Login"
2. Enter your 10-digit Register Number: {student_data['register_no']}
3. An OTP will be sent to THIS college email ({student_data['college_email']})
4. Enter the OTP to verify your identity
5. Set your password

After setting your password, you can login using:
- Email: {student_data['email']}
- Password: (the password you set)

If you have any issues, please contact the CSE Department office.

Regards,
CSE Department
College of Engineering Guindy
Anna University
        """
        
        send_mail(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [student_data['college_email']],
            fail_silently=False,
        )
        return True
    except Exception as e:
        print(f"Email error for {student_data['register_no']}: {e}")
        return False


def send_password_setup_email(request, user):
    """Send email to user to set up their password.
    Returns a tuple: (success: bool, error_message: str)
    """
    try:
        # Generate password reset token
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Build password reset URL
        reset_url = request.build_absolute_uri(
            reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
        )
        
        subject = 'Welcome to CSE Department ERP - Set Your Password'
        message = f"""
Dear {user.full_name},

Your account has been created in the CSE Department ERP System.

Email: {user.email}

Please click the link below to set your password:
{reset_url}

This link will expire in 24 hours.

If you did not request this, please ignore this email.

Regards,
CSE Department
College of Engineering Guindy
        """
        
        send_mail(
            subject,
            message,
            settings.EMAIL_HOST_USER,
            [user.email],
            fail_silently=False,
        )
        return True, ""
    except Exception as e:
        print(f"Email error: {e}")
        return False, str(e)


@login_required
def resend_password_email(request, student_id):
    """Resend password setup email to a student"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    try:
        student = get_object_or_404(Student_Profile, id=student_id)
        user = student.user
        
        if user.has_usable_password():
            return JsonResponse({'error': 'User has already set their password'}, status=400)
        
        email_sent, email_error = send_password_setup_email(request, user)
        if not email_sent:
            return JsonResponse({'error': f'Failed to send email: {email_error}'}, status=400)
        return JsonResponse({'success': True, 'message': f'Password setup email sent to {user.email}'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# =============================================================================
# SEMESTER PROMOTION MANAGEMENT
# =============================================================================

from .models import SemesterPromotion, PromotionSchedule, check_and_promote_students, promote_students_manually, create_promotion_schedules_for_semester
from datetime import timedelta
from django.utils import timezone


@login_required
def manage_promotions(request):
    """View and manage student semester promotions"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Get pending promotions
    today = timezone.now().date()
    pending_schedules = PromotionSchedule.objects.filter(
        executed=False
    ).select_related('semester', 'semester__academic_year').order_by('scheduled_date')
    
    # Get recent promotions
    recent_promotions = SemesterPromotion.objects.select_related(
        'student', 'student__user', 'academic_year', 'promoted_by'
    ).order_by('-promoted_at')[:50]
    
    # Get students by semester for manual promotion
    semesters = range(1, 9)  # Semesters 1-8
    students_by_sem = {}
    for sem in semesters:
        students_by_sem[sem] = Student_Profile.objects.filter(
            current_sem=sem, status='ACTIVE'
        ).count()
    
    # Check for overdue promotions
    overdue_count = pending_schedules.filter(scheduled_date__lt=today).count()

    # Build semester end dates dict {sem_number: 'YYYY-MM-DD'} from most recent academic year
    import json as _json
    sem_end_dates = {}
    try:
        seen_sems = set()
        for sem in Semester.objects.order_by('-academic_year__year', 'semester_number'):
            if sem.semester_number not in seen_sems:
                sem_end_dates[sem.semester_number] = sem.end_date.isoformat()
                seen_sems.add(sem.semester_number)
    except Exception:
        pass

    context = {
        'page_title': 'Semester Promotions',
        'pending_schedules': pending_schedules,
        'recent_promotions': recent_promotions,
        'students_by_sem': students_by_sem,
        'overdue_count': overdue_count,
        'today': today,
        'sem_end_dates_json': _json.dumps(sem_end_dates),
    }
    
    return render(request, 'hod_template/manage_promotions.html', context)


@login_required
@csrf_exempt
def run_auto_promotion(request):
    """Manually trigger the auto-promotion check"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    results = check_and_promote_students(promoted_by=request.user)
    
    return JsonResponse({
        'success': True,
        'total_promoted': results['total_promoted'],
        'semesters_processed': results['semesters_processed'],
        'errors': results['errors']
    })


@login_required
@csrf_exempt
def manual_promote_students(request):
    """Manually promote selected students"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        student_ids = data.get('student_ids', [])
        to_semester = int(data.get('to_semester', 0))
        
        if not student_ids:
            return JsonResponse({'error': 'No students selected'}, status=400)
        
        if to_semester < 1 or to_semester > 8:
            return JsonResponse({'error': 'Invalid target semester'}, status=400)

        from_semester = to_semester - 1

        # Enforce sequential-only promotion
        if from_semester < 1:
            return JsonResponse({'error': 'Cannot promote to Semester 1'}, status=400)

        # Date gate: use program-specific end date if available, else semester default
        try:
            sem_obj = Semester.objects.filter(
                semester_number=from_semester
            ).order_by('-academic_year__year').first()
            if sem_obj:
                today = timezone.now().date()
                blocked_programs = set()
                for stu in Student_Profile.objects.filter(id__in=student_ids, current_sem=from_semester):
                    prog = Program.objects.filter(code=stu.branch, level=stu.program_type).first()
                    psd = ProgramSemesterDate.objects.filter(semester=sem_obj, program=prog).first() if prog else None
                    effective_end = (psd.end_date if psd and psd.end_date else sem_obj.end_date)
                    if today < effective_end:
                        blocked_programs.add(
                            "{} {} (ends {})".format(stu.program_type, stu.branch, effective_end)
                        )
                if blocked_programs:
                    return JsonResponse({
                        'error': 'Semester {} has not ended yet for: {}. Promotion not allowed.'.format(
                            from_semester, ', '.join(sorted(blocked_programs))
                        )
                    }, status=400)
        except Exception:
            pass  # If no semester record, skip date gate

        students = Student_Profile.objects.filter(id__in=student_ids, current_sem=from_semester)
        
        # Get current academic year
        academic_year = AcademicYear.objects.order_by('-year').first()
        
        results = promote_students_manually(
            students=students,
            to_semester=to_semester,
            promoted_by=request.user,
            academic_year=academic_year
        )
        
        return JsonResponse({
            'success': True,
            'promoted': results['success'],
            'errors': results['errors']
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def create_promotion_schedule(request):
    """Create a promotion schedule for a semester"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        semester_id = data.get('semester_id')
        from_semester = int(data.get('from_semester', 0))
        scheduled_date = data.get('scheduled_date')
        
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Create or update schedule
        schedule, created = PromotionSchedule.objects.update_or_create(
            semester=semester,
            target_semester_number=from_semester,
            defaults={
                'scheduled_date': scheduled_date,
                'executed': False
            }
        )
        
        return JsonResponse({
            'success': True,
            'created': created,
            'schedule_id': schedule.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_students_for_promotion(request):
    """Get list of students in a specific semester for promotion"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Access Denied'}, status=403)
    
    semester = request.GET.get('semester')
    if not semester:
        return JsonResponse({'error': 'Semester required'}, status=400)
    
    students = Student_Profile.objects.filter(
        current_sem=int(semester),
        status='ACTIVE'
    ).select_related('user').order_by('register_no')
    
    data = [{
        'id': s.id,
        'name': s.user.get_full_name(),
        'register_no': s.register_no,
        'current_sem': s.current_sem,
        'year_of_study': s.year_of_study,
        'batch': s.batch_label
    } for s in students]
    
    return JsonResponse({'students': data})


# =============================================================================
# STRUCTURED QUESTION PAPER - HOD REVIEW
# =============================================================================

@login_required
def hod_review_structured_qps(request):
    """List all structured question papers for HOD review"""
    from main_app.models import StructuredQuestionPaper
    
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Get all structured QPs - show only the latest revision per course/faculty
    qps = StructuredQuestionPaper.objects.all().select_related(
        'faculty__user', 'course', 'academic_year', 'semester', 'regulation'
    ).order_by('-revision_number', '-submitted_at', '-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        qps = qps.filter(status=status_filter)
    
    context = {
        'qps': qps,
        'status_filter': status_filter,
        'page_title': 'Review Structured Question Papers'
    }
    return render(request, "hod_template/review_structured_qps.html", context)


@login_required
def hod_review_structured_qp_detail(request, qp_id):
    """Detailed review of a structured question paper"""
    from main_app.models import StructuredQuestionPaper
    
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id)
    
    # Update status if viewing for first time
    if qp.status == 'SUBMITTED':
        qp.status = 'UNDER_REVIEW'
        qp.save()
    
    # Calculate distribution and validation
    distribution = qp.calculate_marks_distribution()
    validation_result = qp.validate_distribution()
    validation_errors = validation_result['errors']
    validation_suggestions = validation_result['suggestions']
    
    # Repetition detection
    repetitions = qp.check_repetitions()
    
    # Get questions by part
    part_a_questions = qp.get_part_a_questions()
    part_b_questions = qp.get_part_b_questions()
    part_c_questions = qp.get_part_c_questions()
    
    # Group Part B by OR pairs
    part_b_pairs = {}
    for q in part_b_questions:
        if q.or_pair_number not in part_b_pairs:
            part_b_pairs[q.or_pair_number] = []
        part_b_pairs[q.or_pair_number].append(q)
    
    context = {
        'qp': qp,
        'part_a_questions': part_a_questions,
        'part_b_pairs': sorted(part_b_pairs.items()),
        'part_c_questions': part_c_questions,
        'distribution': distribution,
        'validation_errors': validation_errors,
        'validation_suggestions': validation_suggestions,
        'repetitions': repetitions,
        'can_approve': len(validation_errors) == 0,
        'page_title': f'Review QP - {qp.course.course_code}'
    }
    return render(request, "hod_template/review_structured_qp_detail.html", context)


@login_required
def hod_approve_structured_qp(request, qp_id):
    """Approve a structured question paper"""
    from main_app.models import StructuredQuestionPaper
    from django.utils import timezone
    
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id)
    
    # Validate can approve
    if qp.status == 'APPROVED':
        messages.warning(request, "Question paper already approved.")
        return redirect('hod_review_structured_qp_detail', qp_id=qp.id)
    
    # Check validation
    validation_result = qp.validate_distribution()
    validation_errors = validation_result['errors']
    if validation_errors:
        messages.error(request, "Cannot approve: " + "; ".join(validation_errors))
        return redirect('hod_review_structured_qp_detail', qp_id=qp.id)
    
    if request.method == 'POST':
        comments = request.POST.get('hod_comments', '')
        release_datetime_str = request.POST.get('release_datetime', '')
        
        # Parse release datetime
        release_dt = None
        if release_datetime_str:
            try:
                from datetime import datetime
                release_dt = datetime.strptime(release_datetime_str, '%Y-%m-%dT%H:%M')
                release_dt = timezone.make_aware(release_dt) if timezone.is_naive(release_dt) else release_dt
            except (ValueError, TypeError):
                messages.error(request, "Invalid release date/time format.")
                return redirect('hod_review_structured_qp_detail', qp_id=qp.id)
        
        qp.status = 'APPROVED'
        qp.hod_comments = comments
        qp.reviewed_by = request.user
        qp.reviewed_at = timezone.now()
        qp.release_datetime = release_dt
        # Auto-approve answer key along with QP
        if qp.answer_key_document and qp.answer_key_status == 'SUBMITTED':
            qp.answer_key_status = 'APPROVED'
        qp.save()
        
        # Save approved questions to QuestionBank for repetition detection
        from main_app.models import QuestionBank
        for q in qp.questions.all():
            QuestionBank.objects.get_or_create(
                course=qp.course,
                question_text=q.question_text,
                source_qp=qp,
                defaults={
                    'part': q.part,
                    'marks': q.marks,
                    'bloom_level': q.bloom_level or '',
                    'course_outcome': q.course_outcome or '',
                    'exam_session': qp.exam_month_year or '',
                    'question_image': q.question_image if q.question_image else None,
                }
            )
        
        # Update assignment if linked
        if qp.qp_assignment:
            qp.qp_assignment.status = 'APPROVED'
            qp.qp_assignment.hod_comments = comments
            qp.qp_assignment.reviewed_at = timezone.now()
            qp.qp_assignment.save()
        
        # Notify faculty (mention answer key requirement)
        release_info = f' It will be released to students on {release_dt.strftime("%d %b %Y at %I:%M %p")}.' if release_dt else ''
        Notification.objects.create(
            recipient=qp.faculty.user,
            title='Question Paper Approved',
            message=f'Your structured question paper for {qp.course.course_code} has been approved by HOD.{release_info} Please submit the answer key.',
            notification_type='INFO'
        )
        
        messages.success(request, f"Question paper for {qp.course.course_code} approved successfully!")
        return redirect('hod_review_structured_qps')
    
    context = {
        'qp': qp,
        'page_title': f'Approve QP - {qp.course.course_code}'
    }
    return render(request, "hod_template/approve_structured_qp.html", context)


@login_required
def hod_reject_structured_qp(request, qp_id):
    """Reject a structured question paper with feedback"""
    from main_app.models import StructuredQuestionPaper
    from django.utils import timezone
    
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id)
    
    if qp.status == 'APPROVED':
        messages.error(request, "Cannot reject an approved question paper.")
        return redirect('hod_review_structured_qp_detail', qp_id=qp.id)
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not rejection_reason:
            messages.error(request, "Rejection reason is required.")
            return redirect('hod_review_structured_qp_detail', qp_id=qp.id)
        
        qp.status = 'REJECTED'
        qp.hod_comments = rejection_reason
        qp.reviewed_by = request.user
        qp.reviewed_at = timezone.now()
        qp.save()
        
        # Update assignment if linked
        if qp.qp_assignment:
            qp.qp_assignment.status = 'REJECTED'
            qp.qp_assignment.hod_comments = rejection_reason
            qp.qp_assignment.reviewed_at = timezone.now()
            qp.qp_assignment.save()
        
        # Notify faculty
        Notification.objects.create(
            recipient=qp.faculty.user,
            title='Question Paper Rejected',
            message=f'Your structured question paper for {qp.course.course_code} requires revision. Reason: {rejection_reason[:100]}',
            notification_type='WARNING'
        )
        
        messages.warning(request, f"Question paper rejected. Faculty has been notified.")
        return redirect('hod_review_structured_qps')
    
    context = {
        'qp': qp,
        'page_title': f'Reject QP - {qp.course.course_code}'
    }
    return render(request, "hod_template/reject_structured_qp.html", context)


@login_required
def hod_download_structured_qp(request, qp_id):
    """Download structured question paper document"""
    from main_app.models import StructuredQuestionPaper
    from django.http import FileResponse, Http404
    
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id)

    base_field = qp.generated_document if qp.generated_document else qp.uploaded_document
    if not base_field:
        raise Http404("Document not available")

    checklist_exists = bool(qp.submission_checklist or qp.auto_distribution_checklist)
    filename = base_field.name.split('/')[-1]
    is_docx = filename.lower().endswith('.docx')

    if checklist_exists and is_docx:
        from docx import Document

        with base_field.open('rb') as fp:
            doc = Document(fp)

        append_checklist_to_document(doc, qp)

        output = io.BytesIO()
        doc.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = f'attachment; filename="qp_{qp.id}_complete.docx"'
        return response

    response = FileResponse(base_field.open('rb'))
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# EXAM SCHEDULE MANAGEMENT
# =============================================================================

@login_required
def manage_exam_schedules(request):
    """View and manage all exam schedules"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    # Update status of all schedules based on current time
    schedules = ExamSchedule.objects.select_related(
        'structured_qp', 'structured_qp__course', 'structured_qp__faculty',
        'semester', 'scheduled_by'
    ).order_by('-exam_date', '-start_time')
    
    # Update statuses
    for schedule in schedules:
        schedule.update_status()
    
    # Filter options
    status_filter = request.GET.get('status', '')
    course_filter = request.GET.get('course', '')
    
    if status_filter:
        schedules = schedules.filter(status=status_filter)
    if course_filter:
        schedules = schedules.filter(structured_qp__course__id=course_filter)
    
    # Get courses for filter dropdown
    courses = Course.objects.filter(
        id__in=ExamSchedule.objects.values_list('structured_qp__course', flat=True)
    ).distinct()
    
    context = {
        'schedules': schedules,
        'courses': courses,
        'status_filter': status_filter,
        'course_filter': course_filter,
        'page_title': 'Manage Exam Schedules'
    }
    return render(request, 'hod_template/manage_exam_schedules.html', context)


@login_required
def schedule_exam(request):
    """HOD schedules an exam for an approved question paper"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    if request.method == 'POST':
        form = ExamScheduleForm(request.POST)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.scheduled_by = request.user
            schedule.save()
            
            # Notify relevant students
            course = schedule.structured_qp.course
            messages.success(
                request, 
                f"Exam scheduled for {course.course_code} - {course.title} on {schedule.exam_date}"
            )
            return redirect('manage_exam_schedules')
    else:
        form = ExamScheduleForm()
    
    # Get approved QPs that don't have schedules yet
    approved_qps = StructuredQuestionPaper.objects.filter(
        status='APPROVED'
    ).exclude(
        exam_schedule__isnull=False
    ).select_related('course', 'faculty')
    
    context = {
        'form': form,
        'approved_qps': approved_qps,
        'page_title': 'Schedule Exam'
    }
    return render(request, 'hod_template/schedule_exam.html', context)


@login_required
def edit_exam_schedule(request, schedule_id):
    """Edit an existing exam schedule (only if exam hasn't ended)"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    schedule = get_object_or_404(ExamSchedule, id=schedule_id)
    
    # Check if editable
    if schedule.is_exam_ended:
        messages.error(request, "Cannot edit schedule after exam has ended.")
        return redirect('manage_exam_schedules')
    
    if request.method == 'POST':
        form = ExamScheduleEditForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, "Exam schedule updated successfully.")
            return redirect('manage_exam_schedules')
    else:
        form = ExamScheduleEditForm(instance=schedule)
    
    context = {
        'form': form,
        'schedule': schedule,
        'page_title': f'Edit Exam Schedule - {schedule.structured_qp.course.course_code}'
    }
    return render(request, 'hod_template/edit_exam_schedule.html', context)


@login_required
def delete_exam_schedule(request, schedule_id):
    """Delete an exam schedule (only if exam hasn't started)"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    schedule = get_object_or_404(ExamSchedule, id=schedule_id)
    
    # Only allow deletion if exam hasn't started
    if schedule.is_exam_started:
        messages.error(request, "Cannot delete schedule after exam has started.")
        return redirect('manage_exam_schedules')
    
    course_code = schedule.structured_qp.course.course_code
    schedule.delete()
    messages.success(request, f"Exam schedule for {course_code} deleted successfully.")
    return redirect('manage_exam_schedules')


@login_required
def view_exam_schedule_detail(request, schedule_id):
    """View detailed information about an exam schedule"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    schedule = get_object_or_404(
        ExamSchedule.objects.select_related(
            'structured_qp', 'structured_qp__course', 'structured_qp__faculty',
            'semester', 'scheduled_by'
        ),
        id=schedule_id
    )
    
    # Update status
    schedule.update_status()
    
    # Get questions for this QP
    questions = schedule.structured_qp.questions.all().order_by('part', 'question_number')
    
    context = {
        'schedule': schedule,
        'questions': questions,
        'page_title': f'Exam Schedule - {schedule.structured_qp.course.course_code}'
    }
    return render(request, 'hod_template/exam_schedule_detail.html', context)


@login_required
def mark_exam_completed(request, schedule_id):
    """Manually mark an exam as completed"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    schedule = get_object_or_404(ExamSchedule, id=schedule_id)
    
    if request.method == 'POST':
        schedule.status = 'COMPLETED'
        schedule.save()
        messages.success(request, "Exam marked as completed. QP and answers will now be visible to students.")
        return redirect('manage_exam_schedules')
    
    return redirect('view_exam_schedule_detail', schedule_id=schedule_id)


@login_required
def cancel_exam_schedule(request, schedule_id):
    """Cancel an exam schedule"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')
    
    schedule = get_object_or_404(ExamSchedule, id=schedule_id)
    
    if schedule.is_exam_ended:
        messages.error(request, "Cannot cancel an exam that has already ended.")
        return redirect('manage_exam_schedules')
    
    if request.method == 'POST':
        schedule.status = 'CANCELLED'
        schedule.save()
        messages.warning(request, f"Exam for {schedule.structured_qp.course.course_code} has been cancelled.")
        return redirect('manage_exam_schedules')
    
    return redirect('view_exam_schedule_detail', schedule_id=schedule_id)



# =============================================================================
# TIMETABLE CREATION WIZARD
# =============================================================================

@login_required
def create_timetable_wizard(request):
    """Main wizard page for creating timetables (3-step process)"""
    if not check_hod_permission(request.user):
        return redirect('/')
    
    current_year = AcademicYear.get_current()
    current_semester = Semester.get_current()
    programs = Program.objects.all().order_by('level', 'code')
    labs = LabRoom.objects.filter(is_active=True).prefetch_related(
        'restrictions', 'restrictions__program', 'restrictions__course'
    )
    time_slots = TimeSlot.objects.all().order_by('slot_number')
    # All lab courses for restriction dropdown (global, not per-program)
    # Note: Course PK is course_code, not id
    all_lab_courses = list(Course.objects.filter(
        course_type__in=['L', 'LIT']
    ).order_by('course_code').values('course_code', 'title'))
    
    context = {
        'page_title': 'Create Timetable - Wizard',
        'current_year': current_year,
        'current_semester': current_semester,
        'programs': programs,
        'labs': labs,
        'time_slots': time_slots,
        'days': TimetableEntry.DAY_CHOICES,
        'year_choices': [(1, '1st Year'), (2, '2nd Year'), (3, '3rd Year'), (4, '4th Year')],
        'all_lab_courses_json': json.dumps(all_lab_courses),
    }
    return render(request, 'hod_template/create_timetable_wizard.html', context)


@login_required
def api_get_batches_for_program(request):
    """AJAX: Given program_id + year_of_study, return available batches"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    program_id = request.GET.get('program_id')
    year_of_study = request.GET.get('year_of_study')
    
    if not program_id or not year_of_study:
        return JsonResponse({'error': 'program_id and year_of_study required'}, status=400)
    
    current_year = AcademicYear.get_current()
    if not current_year:
        return JsonResponse({'error': 'No active academic year found'}, status=400)
    
    batches = _get_cohort_batches(current_year, int(program_id), int(year_of_study))
    
    data = [{
        'id': b.id,
        'batch_name': b.batch_name,
        'batch_display': b.batch_display or b.batch_name,
        'capacity': b.capacity,
    } for b in batches]
    
    # Also get semester info
    try:
        year_int = int(year_of_study)
        # Year 1 -> Sems 1,2; Year 2 -> Sems 3,4; etc.
        sem_start = (year_int - 1) * 2 + 1
        sem_end = year_int * 2
        semesters = Semester.objects.filter(
            academic_year=current_year,
            semester_number__gte=sem_start,
            semester_number__lte=sem_end
        ).order_by('semester_number')
        semester_data = [{
            'id': s.id,
            'semester_number': s.semester_number,
            'semester_type': s.semester_type,
        } for s in semesters]
    except Exception:
        semester_data = []
    
    return JsonResponse({
        'batches': data,
        'semesters': semester_data,
        'academic_year': str(current_year) if current_year else '',
        'academic_year_id': current_year.id if current_year else None,
    })


@login_required
def api_get_courses_for_program_year(request):
    """AJAX: Given program_id + year_of_study, return courses from Course_Assignment"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    program_id = request.GET.get('program_id')
    year_of_study = request.GET.get('year_of_study')
    
    if not program_id or not year_of_study:
        return JsonResponse({'error': 'program_id and year_of_study required'}, status=400)
    
    year_of_study = int(year_of_study)
    
    current_year = AcademicYear.get_current()
    if not current_year:
        return JsonResponse({'error': 'No active academic year found'}, status=400)
    
    # Determine the correct semester for this year of study
    current_semester = Semester.get_current_for_year(year_of_study)
    
    # Get active course assignments for this program+year's batches
    # using centralized cohort mapping.
    batches = _get_cohort_batches(current_year, int(program_id), year_of_study)
    
    assignment_filter = {
        'academic_year': current_year,
        'batch__in': batches,
        'is_active': True,
    }
    if current_semester:
        assignment_filter['semester'] = current_semester
    
    assignments = Course_Assignment.objects.filter(
        **assignment_filter
    ).exclude(
        special_note__icontains='Skip Timetable'
    ).select_related('course', 'faculty', 'faculty__user', 'batch')
    
    # Legacy batch_label-only records are intentionally ignored for wizard scheduling,
    # because they are not visible/manageable in normal assignment flows and can cause
    # hidden stale courses to be scheduled after deletion.
    all_assignments = list(assignments)
    
    # Build unique course list
    courses_seen = set()
    courses_data = []
    assignments_data = []
    
    for a in all_assignments:
        if a.course.course_code not in courses_seen:
            courses_seen.add(a.course.course_code)
            courses_data.append({
                'course_code': a.course.course_code,
                'title': a.course.title,
                'course_type': a.course.course_type,
                'credits': a.course.credits,
                'is_lab': a.course.course_type in ['L', 'LIT'],
                'ltp': a.course.ltp_display,
            })
        
        assignments_data.append({
            'id': a.id,
            'course_code': a.course.course_code,
            'course_title': a.course.title,
            'faculty_id': a.faculty.id,
            'faculty_name': a.faculty.user.full_name,
            'batch_name': a.batch.batch_name if a.batch else a.batch_label,
            'batch_id': a.batch.id if a.batch else None,
        })
    
    # ── Also include elective placeholder courses (PEC/OEC) from RegulationCoursePlan ──
    # These appear in the wizard as "PEC-03", "PEC-04" etc. (not the resolved subject names).
    # The timetable engine schedules the SLOT (PEC-03) and blocks ALL offering faculty.
    try:
        program = Program.objects.get(id=program_id)
        if program.regulation_id and current_semester:
            sem_number = current_semester.semester_number
            # Get elective plan entries for this regulation+semester+branch
            elective_plans = RegulationCoursePlan.objects.filter(
                regulation_id=program.regulation_id,
                semester=sem_number,
                is_elective=True,
                branch=program.code,
            ).select_related('course')

            sorted_batches = list(batches.order_by('batch_name'))

            for plan in elective_plans:
                placeholder = plan.course
                # Skip OEC and AC (blocked-slot) placeholders — not auto-scheduled
                if placeholder.course_code.startswith('OEC') or placeholder.course_code.startswith('AC'):
                    continue

                # Skip if already present via Course_Assignment
                if placeholder.course_code in courses_seen:
                    continue

                # Check if at least one offering is configured for this placeholder
                has_offering = ElectiveCourseOffering.objects.filter(
                    regulation_course_plan=plan,
                    semester=current_semester,
                    is_active=True,
                ).exists()

                if not has_offering:
                    continue

                courses_seen.add(placeholder.course_code)

                # Determine course_type from the offerings (placeholder has type=None)
                # Use the first offering's actual course type as a hint
                first_offering = ElectiveCourseOffering.objects.filter(
                    regulation_course_plan=plan,
                    semester=current_semester,
                    is_active=True,
                ).select_related('actual_course').first()
                
                inferred_type = 'T'  # default
                inferred_ltp = f'{placeholder.credits}-0-0'
                if first_offering and first_offering.actual_course:
                    inferred_type = first_offering.actual_course.course_type or 'T'
                    inferred_ltp = first_offering.actual_course.ltp_display

                courses_data.append({
                    'course_code': placeholder.course_code,
                    'title': placeholder.title,
                    'course_type': inferred_type,
                    'credits': placeholder.credits,
                    'is_lab': inferred_type in ['L', 'LIT'],
                    'ltp': inferred_ltp,
                    'is_placeholder': True,
                })

                # Gather ALL faculty from offerings under this placeholder
                offerings = ElectiveCourseOffering.objects.filter(
                    regulation_course_plan=plan,
                    semester=current_semester,
                    is_active=True,
                ).select_related('actual_course').order_by('actual_course__course_code')
                offering_codes = [o.actual_course.course_code for o in offerings]
                if offering_codes:
                    courses_data[-1]['offering_codes'] = offering_codes
                    courses_data[-1]['offering_display'] = '/'.join(offering_codes)
                all_offering_faculty = ElectiveOfferingFacultyAssignment.objects.filter(
                    offering__in=offerings,
                    is_active=True,
                ).select_related('faculty__user')

                # Create one assignment per batch using offering faculty
                # (one faculty per batch, or same faculty for all batches if single offering batch)
                faculty_list = list(all_offering_faculty)

                if faculty_list:
                    # For display: assign to each batch
                    for pb in sorted_batches:
                        # Pick the first faculty (all teach under this PEC slot)
                        fa = faculty_list[0]
                        assignments_data.append({
                            'id': f'elective_{plan.pk}_{pb.id}',
                            'course_code': placeholder.course_code,
                            'course_title': placeholder.title,
                            'faculty_id': fa.faculty.id,
                            'faculty_name': fa.faculty.user.full_name,
                            'batch_name': pb.batch_name,
                            'batch_id': pb.id,
                            'is_elective': True,
                        })
    except Program.DoesNotExist:
        pass
    
    return JsonResponse({
        'courses': courses_data,
        'assignments': assignments_data,
    })


@login_required
@csrf_exempt
def api_save_reservation(request):
    """AJAX: Save or update a FixedSlotReservation"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = json.loads(request.body)
        config_id = data.get('config_id')
        batch_id = data['batch_id']
        day = data['day']
        slot_number = data['slot_number']
        is_blocked = data.get('is_blocked', False)
        block_reason = data.get('block_reason', '')
        course_code = data.get('course_code')
        faculty_id = data.get('faculty_id')
        special_note = data.get('special_note', '')
        apply_to_all = data.get('apply_to_all_batches', False)
        
        time_slot = get_object_or_404(TimeSlot, slot_number=slot_number)
        batch = get_object_or_404(ProgramBatch, id=batch_id)
        
        # For blocked slots, course is optional
        course = None
        faculty = None
        if not is_blocked:
            if not course_code:
                return JsonResponse({'error': 'course_code required for non-blocked slots'}, status=400)
            course = get_object_or_404(Course, course_code=course_code)
            faculty = Faculty_Profile.objects.filter(id=faculty_id).first() if faculty_id else None
        
        # Get or create config
        if config_id:
            config = get_object_or_404(TimetableConfig, id=config_id)
        else:
            # Create new config
            academic_year_id = data.get('academic_year_id')
            semester_id = data.get('semester_id')
            program_id = data.get('program_id')
            year_of_study = data.get('year_of_study')
            
            config, _ = TimetableConfig.objects.get_or_create(
                academic_year_id=academic_year_id,
                semester_id=semester_id,
                program_id=program_id,
                year_of_study=year_of_study,
                is_generated=False,
                defaults={'created_by': request.user}
            )
        
        reservations_created = []
        
        if apply_to_all:
            # Apply COURSE to all batches, but look up each batch's own assigned faculty
            all_batches = _get_cohort_batches(
                config.academic_year,
                config.program_id,
                config.year_of_study,
            )

            if not all_batches.exists():
                return JsonResponse({
                    'error': f'No batches found for {config.program.code} Year {config.year_of_study} in current/cohort mapping.'
                }, status=400)

            for b in all_batches:
                batch_faculty = faculty  # default to the selected faculty
                batch_faculty_name = faculty.user.full_name if faculty else ''
                
                if not is_blocked and course:
                    # Look up this batch's assigned faculty for this course
                    # Don't filter by semester — config semester may differ from assignment semester
                    assignment = Course_Assignment.objects.filter(
                        course=course,
                        batch=b,
                        academic_year=config.academic_year,
                        is_active=True,
                    ).select_related('faculty__user').first()
                    if assignment:
                        batch_faculty = assignment.faculty
                        batch_faculty_name = assignment.faculty.user.full_name
                
                defaults = {
                    'course': course,
                    'faculty': batch_faculty,
                    'special_note': special_note or None,
                    'apply_to_all_batches': True,
                    'is_blocked': is_blocked,
                    'block_reason': block_reason or None,
                }
                res, _ = FixedSlotReservation.objects.update_or_create(
                    config=config,
                    batch=b,
                    day=day,
                    time_slot=time_slot,
                    defaults=defaults,
                )
                reservations_created.append({
                    'id': res.id,
                    'batch_id': b.id,
                    'batch_name': b.batch_name,
                    'faculty_id': batch_faculty.id if batch_faculty else None,
                    'faculty_name': batch_faculty_name,
                })
        else:
            defaults = {
                'course': course,
                'faculty': faculty,
                'special_note': special_note or None,
                'apply_to_all_batches': False,
                'is_blocked': is_blocked,
                'block_reason': block_reason or None,
            }
            res, _ = FixedSlotReservation.objects.update_or_create(
                config=config,
                batch=batch,
                day=day,
                time_slot=time_slot,
                defaults=defaults,
            )
            reservations_created.append({
                'id': res.id,
                'batch_id': batch.id,
                'batch_name': batch.batch_name,
                'faculty_id': faculty.id if faculty else None,
                'faculty_name': faculty.user.full_name if faculty else '',
            })

        if not reservations_created:
            return JsonResponse({'error': 'No reservations were created for the selected context.'}, status=400)
        
        return JsonResponse({
            'success': True,
            'config_id': config.id,
            'reservations': reservations_created,
            'course_code': course.course_code if course else '',
            'is_blocked': is_blocked,
            'block_reason': block_reason,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@csrf_exempt
def api_delete_reservation(request):
    """AJAX: Delete a FixedSlotReservation"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = json.loads(request.body)
        config_id = data['config_id']
        batch_id = data['batch_id']
        day = data['day']
        slot_number = data['slot_number']
        
        time_slot = get_object_or_404(TimeSlot, slot_number=slot_number)
        
        # Check if this was an "apply to all" reservation
        reservation = FixedSlotReservation.objects.filter(
            config_id=config_id,
            batch_id=batch_id,
            day=day,
            time_slot=time_slot
        ).first()
        
        deleted_batches = []
        if reservation and reservation.apply_to_all_batches:
            # Delete from all batches
            all_deletions = FixedSlotReservation.objects.filter(
                config_id=config_id,
                day=day,
                time_slot=time_slot,
                apply_to_all_batches=True
            )
            for d in all_deletions:
                deleted_batches.append(d.batch_id)
            all_deletions.delete()
        elif reservation:
            deleted_batches.append(reservation.batch_id)
            reservation.delete()
        
        return JsonResponse({'success': True, 'deleted_batches': deleted_batches})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_get_reservations(request):
    """AJAX: Get all reservations for a config"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    config_id = request.GET.get('config_id')
    if not config_id:
        return JsonResponse({'reservations': []})
    
    reservations = FixedSlotReservation.objects.filter(
        config_id=config_id
    ).select_related('batch', 'course', 'faculty', 'faculty__user', 'time_slot')
    
    data = [{
        'id': r.id,
        'batch_id': r.batch.id,
        'batch_name': r.batch.batch_name,
        'day': r.day,
        'slot_number': r.time_slot.slot_number,
        'course_code': r.course.course_code if r.course else '',
        'course_title': r.course.title if r.course else '',
        'faculty_id': r.faculty.id if r.faculty else None,
        'faculty_name': r.faculty.user.full_name if r.faculty else '',
        'special_note': r.special_note or '',
        'apply_to_all_batches': r.apply_to_all_batches,
        'is_blocked': r.is_blocked,
        'block_reason': r.block_reason or '',
    } for r in reservations]
    
    return JsonResponse({'reservations': data})


@login_required
@csrf_exempt
def api_save_lab_config(request):
    """AJAX: Save lab room and restriction configuration (full sync)"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = json.loads(request.body)
        labs_data = data.get('labs', [])
        config_id = data.get('config_id')
        selected_lab_ids = data.get('selected_lab_ids', [])
        full_sync = bool(data.get('full_sync', False))
        
        # Track which existing lab IDs are still present
        kept_lab_ids = set()
        
        for lab_data in labs_data:
            lab_id = lab_data.get('id')
            room_code = str(lab_data.get('room_code', '')).strip()
            lab_name = str(lab_data.get('name', '')).strip()
            capacity = int(lab_data.get('capacity', 60) or 60)
            
            if lab_id:
                lab = LabRoom.objects.get(id=int(lab_id))
                new_code = room_code or lab.room_code
                # Check for room_code conflict with a different active lab
                conflict = LabRoom.objects.filter(
                    room_code=new_code, is_active=True
                ).exclude(id=lab.id).first()
                if conflict:
                    return JsonResponse({
                        'error': f'Room code "{new_code}" is already used by another lab.'
                    }, status=400)
                # Remove any inactive lab with this code to avoid unique constraint
                LabRoom.objects.filter(
                    room_code=new_code, is_active=False
                ).exclude(id=lab.id).delete()
                lab.room_code = new_code
                lab.name = lab_name or lab.name
                lab.capacity = capacity
                lab.is_active = True
                lab.save()
                kept_lab_ids.add(lab.id)
            else:
                if not room_code or not lab_name:
                    continue  # skip empty rows
                lab, _created = LabRoom.objects.update_or_create(
                    room_code=room_code,
                    defaults={
                        'name': lab_name,
                        'capacity': capacity,
                        'is_active': True,
                    },
                )
                kept_lab_ids.add(lab.id)
            
            # Update restrictions — clear and recreate
            LabRestriction.objects.filter(lab=lab).delete()
            
            for restriction in lab_data.get('restrictions', []):
                prog_id = restriction.get('program_id')
                year = restriction.get('year_of_study')
                course_id = restriction.get('course_id')
                course_code = restriction.get('course_code')
                
                # Clean types — JS may send empty strings
                prog_id = int(prog_id) if prog_id else None
                year = int(year) if year else None
                # course_id is actually course_code (Course PK is a string)
                # Accept either course_id or course_code from JS
                course_pk = course_id or course_code or None
                if course_pk:
                    course_pk = str(course_pk).strip()
                    if not Course.objects.filter(course_code=course_pk).exists():
                        course_pk = None
                
                if prog_id or year or course_pk:  # Only create non-empty restrictions
                    LabRestriction.objects.create(
                        lab=lab,
                        program_id=prog_id,
                        year_of_study=year,
                        course_id=course_pk,
                    )
        
        # Deactivate labs that were removed from the UI only for explicit full-sync saves.
        if full_sync and kept_lab_ids:
            LabRoom.objects.filter(is_active=True).exclude(id__in=kept_lab_ids).update(is_active=False)
        
        # Refresh labs data and return updated IDs
        updated_labs = []
        for lab in LabRoom.objects.filter(id__in=kept_lab_ids, is_active=True):
            updated_labs.append({'id': lab.id, 'room_code': lab.room_code})
        
        # Save selected labs for config if config_id provided
        if config_id and selected_lab_ids:
            config = TimetableConfig.objects.get(id=int(config_id))
            TimetableConfigLab.objects.filter(config=config).delete()
            for lid in selected_lab_ids:
                if int(lid) in kept_lab_ids:
                    TimetableConfigLab.objects.create(config=config, lab_id=int(lid))
        
        return JsonResponse({'success': True, 'labs_saved': len(labs_data), 'labs': updated_labs})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@csrf_exempt
def generate_timetables_from_config(request):
    """
    POST: Create actual Timetable + TimetableEntry records for all batches
    from the wizard's configuration (reserved slots) PLUS auto-fill remaining
    slots using the timetable engine.
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = json.loads(request.body)
        config_id = data.get('config_id')
        effective_from = data.get('effective_from')
        generation_preferences = data.get('generation_preferences') or {}
        if not isinstance(generation_preferences, dict):
            generation_preferences = {}
        
        if not config_id:
            return JsonResponse({'error': 'config_id required'}, status=400)
        
        config = get_object_or_404(TimetableConfig, id=config_id)
        
        if config.is_generated:
            return JsonResponse({'error': 'Timetables already generated from this config'}, status=400)
        
        from datetime import date as date_type
        eff_date = datetime.strptime(effective_from, '%Y-%m-%d').date() if effective_from else date_type.today()
        
        # Use the auto-fill engine
        from main_app.utils.timetable_engine import TimetableEngine
        engine = TimetableEngine(config, generation_preferences=generation_preferences)
        result = engine.generate(effective_date=eff_date)
        
        if not result['success']:
            return JsonResponse({'error': result.get('error', 'Generation failed')}, status=400)
        
        return JsonResponse({
            'success': True,
            'timetables': result['timetables'],
            'warnings': result['warnings'],
            'redirect_url': reverse('manage_timetables'),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_preview_generation(request):
    """AJAX GET: Preview what auto-generation would produce without saving."""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    config_id = request.GET.get('config_id')
    if not config_id:
        return JsonResponse({'error': 'config_id required'}, status=400)
    
    try:
        config = TimetableConfig.objects.get(id=config_id)
        from main_app.utils.timetable_engine import TimetableEngine
        engine = TimetableEngine(config)
        preview = engine.preview()
        return JsonResponse({'success': True, 'preview': preview})
    except TimetableConfig.DoesNotExist:
        return JsonResponse({'error': 'Config not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_get_labs_for_config(request):
    """AJAX GET: Return labs selected for a config, or all active labs."""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    config_id = request.GET.get('config_id')
    
    all_labs = LabRoom.objects.filter(is_active=True).prefetch_related(
        'restrictions', 'restrictions__program', 'restrictions__course'
    )
    
    selected_ids = set()
    if config_id:
        selected_ids = set(
            TimetableConfigLab.objects.filter(config_id=config_id).values_list('lab_id', flat=True)
        )
    
    labs_data = []
    for lab in all_labs:
        restrictions = []
        for r in lab.restrictions.all():
            restrictions.append({
                'program_id': r.program_id,
                'program_code': r.program.code if r.program else None,
                'year_of_study': r.year_of_study,
                'course_id': r.course_id,
                'course_code': r.course.course_code if r.course else None,
            })
        labs_data.append({
            'id': lab.id,
            'room_code': lab.room_code,
            'name': lab.name,
            'capacity': lab.capacity,
            'is_selected': lab.id in selected_ids if selected_ids else True,
            'restrictions': restrictions,
        })
    
    return JsonResponse({'labs': labs_data})


def _semester_timing_status(semester, today):
    """Return timing status for planning priority."""
    if today < semester.start_date:
        return 'UPCOMING'
    if semester.start_date <= today <= semester.end_date:
        return 'CURRENT'
    return 'COMPLETED'


def _semester_numbers_for_type(semester_type):
    return [1, 3, 5, 7] if semester_type == 'ODD' else [2, 4, 6, 8]


def _get_cohort_batches(current_year, program_id, year_of_study):
    """
    Resolve ProgramBatch rows for a visible year_of_study.

    Primary: current academic year + requested year_of_study
    Fallback: admission academic year + year_of_study=1 (cohort mapping)
    """
    year_int = int(year_of_study)

    qs = ProgramBatch.objects.filter(
        academic_year=current_year,
        program_id=program_id,
        year_of_study=year_int,
        is_active=True,
    ).order_by('batch_name')

    if qs.exists() or year_int <= 1:
        return qs

    try:
        start_year = int(current_year.year.split('-')[0])
        admission_start_year = start_year - (year_int - 1)
        admission_year_label = f"{admission_start_year}-{str(admission_start_year + 1)[-2:]}"
        admission_year = AcademicYear.objects.filter(year=admission_year_label).first()
    except (ValueError, IndexError):
        admission_year = None

    if not admission_year:
        return ProgramBatch.objects.none()

    return ProgramBatch.objects.filter(
        academic_year=admission_year,
        program_id=program_id,
        year_of_study=1,
        is_active=True,
    ).order_by('batch_name')


def _get_program_year_combos(current_year):
    """Return merged program/year combos from direct batches + assignments."""
    combo_map = {}

    direct_combos = ProgramBatch.objects.filter(
        academic_year=current_year,
        is_active=True,
    ).values('program__id', 'program__code', 'program__name', 'year_of_study').annotate(
        batch_count=Count('id')
    )

    for combo in direct_combos:
        key = (combo['program__id'], combo['year_of_study'])
        combo_map[key] = {
            'program__id': combo['program__id'],
            'program__code': combo['program__code'],
            'program__name': combo['program__name'],
            'year_of_study': combo['year_of_study'],
            'batch_names': list(
                ProgramBatch.objects.filter(
                    academic_year=current_year,
                    program_id=combo['program__id'],
                    year_of_study=combo['year_of_study'],
                    is_active=True,
                ).values_list('batch_name', flat=True).order_by('batch_name')
            ),
        }

    assignment_combos = Course_Assignment.objects.filter(
        academic_year=current_year,
        is_active=True,
        batch__isnull=False,
        semester__isnull=False,
    ).select_related('batch__program', 'semester')

    for assignment in assignment_combos:
        program = assignment.batch.program
        year = assignment.semester.year_of_study
        key = (program.id, year)
        if key not in combo_map:
            combo_map[key] = {
                'program__id': program.id,
                'program__code': program.code,
                'program__name': program.name,
                'year_of_study': year,
                'batch_names': [],
            }
        batch_name = assignment.batch.batch_name
        if batch_name and batch_name not in combo_map[key]['batch_names']:
            combo_map[key]['batch_names'].append(batch_name)

    return sorted(
        combo_map.values(),
        key=lambda c: (c['program__code'], c['year_of_study'])
    )


def _select_target_semester_for_program_year(current_year, program_id, year_of_study, require_assignments=False):
    """
    Pick the semester to generate timetable for a specific program/year in the current academic year.

    Priority:
    1) UPCOMING semester (earliest sem number) with assignments
    2) CURRENT semester (earliest sem number) with assignments
    3) COMPLETED semester (earliest sem number) with assignments

    If require_assignments=False, the same timing priority is used without assignment filtering.
    """
    from django.utils import timezone

    sem_start = (year_of_study - 1) * 2 + 1
    sem_end = year_of_study * 2
    semesters = list(
        Semester.objects.filter(
            academic_year=current_year,
            semester_number__in=[sem_start, sem_end],
        ).order_by('semester_number')
    )
    if not semesters:
        return None

    batch_qs = _get_cohort_batches(current_year, program_id, year_of_study)

    assignment_counts = {
        row['semester_id']: row['c']
        for row in Course_Assignment.objects.filter(
            academic_year=current_year,
            batch__in=batch_qs,
            batch__is_active=True,
            is_active=True,
        ).values('semester_id').annotate(c=Count('id'))
    }

    if require_assignments:
        semesters = [s for s in semesters if assignment_counts.get(s.id, 0) > 0]
        if not semesters:
            return None

    today = timezone.now().date()
    status_rank = {'UPCOMING': 0, 'CURRENT': 1, 'COMPLETED': 2}

    return sorted(
        semesters,
        key=lambda s: (
            status_rank.get(_semester_timing_status(s, today), 3),
            s.semester_number,
        ),
    )[0]


@login_required
def api_get_all_program_year_status(request):
    """
    AJAX GET: Return all program+year combinations that have active batches
    in the current academic year, along with their config/reservation status.
    Used by the multi-program wizard Step 2.
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    current_year = AcademicYear.get_current()
    current_semester = Semester.get_current()

    if not current_year:
        return JsonResponse({'error': 'No current academic year configured'}, status=400)

    combos = _get_program_year_combos(current_year)

    results = []
    for combo in combos:
        prog_id = combo['program__id']
        year = combo['year_of_study']

        year_semester = _select_target_semester_for_program_year(
            current_year=current_year,
            program_id=prog_id,
            year_of_study=year,
            require_assignments=False,
        )
        if not year_semester:
            continue

        # Check if a TimetableConfig exists
        config = TimetableConfig.objects.filter(
            academic_year=current_year,
            semester=year_semester,
            program_id=prog_id,
            year_of_study=year,
        ).first()

        # Count reservations if config exists
        reserved = 0
        blocked = 0
        config_id = None
        is_generated = False
        same_time_count = 0
        if config:
            config_id = config.id
            is_generated = config.is_generated
            reserved = config.reservations.filter(is_blocked=False).count()
            blocked = config.reservations.filter(is_blocked=True).count()
            same_time_count = config.same_time_constraints.count()

        # Check if timetable already exists
        existing_timetables = Timetable.objects.filter(
            academic_year=current_year,
            semester=year_semester,
            year=year,
            program_batch__program_id=prog_id,
            is_active=True,
        ).count()

        # Get batch names
        batch_names = list(
            sorted(combo.get('batch_names', []))
        )

        results.append({
            'program_id': prog_id,
            'program_code': combo['program__code'],
            'program_name': combo['program__name'],
            'year_of_study': year,
            'batch_count': len(batch_names),
            'batch_names': batch_names,
            'config_id': config_id,
            'semester_id': year_semester.id,
            'semester_number': year_semester.semester_number,
            'semester_type': year_semester.semester_type,
            'reserved_count': reserved,
            'blocked_count': blocked,
            'same_time_count': same_time_count,
            'is_generated': is_generated,
            'existing_timetables': existing_timetables,
        })

    banner_semester = None
    if results:
        banner_semester = Semester.objects.filter(id=results[0]['semester_id']).first()
    if not banner_semester:
        banner_semester = current_semester

    return JsonResponse({
        'success': True,
        'academic_year': str(current_year),
        'academic_year_id': current_year.id,
        'semester_id': banner_semester.id if banner_semester else None,
        'semester_number': banner_semester.semester_number if banner_semester else None,
        'semester_type': banner_semester.semester_type if banner_semester else None,
        'combos': results,
    })


@login_required
@csrf_exempt
def api_delete_timetables(request):
    """
    POST: Delete all active timetables for a given academic year + semester + program + year,
    and optionally reset the TimetableConfig.
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        scope = data.get('scope', 'program_year')  # 'program_year' or 'all'

        current_year = AcademicYear.get_current()
        current_semester = Semester.get_current()

        if scope == 'all':
            # Delete ALL active timetables for current academic year (any semester)
            deleted_count = 0
            timetables = Timetable.objects.filter(
                academic_year=current_year,
                is_active=True,
            )
            for tt in timetables:
                tt.entries.all().delete()
                tt.delete()
                deleted_count += 1

            # Reset all configs for this academic year
            TimetableConfig.objects.filter(
                academic_year=current_year,
            ).update(is_generated=False)

            return JsonResponse({'success': True, 'deleted': deleted_count})

        else:
            # Delete for specific program + year
            program_id = data.get('program_id')
            year_of_study = data.get('year_of_study')
            if not program_id or not year_of_study:
                return JsonResponse({'error': 'program_id and year_of_study required'}, status=400)

            # Compute the correct semester for this year_of_study
            sem_type = current_semester.semester_type
            yr = int(year_of_study)
            target_sem_num = yr * 2 if sem_type == 'EVEN' else (yr - 1) * 2 + 1
            year_semester = Semester.objects.filter(
                academic_year=current_year, semester_number=target_sem_num
            ).first() or current_semester

            deleted_count = 0
            timetables = Timetable.objects.filter(
                academic_year=current_year,
                year=yr,
                program_batch__program_id=int(program_id),
                is_active=True,
            )
            for tt in timetables:
                tt.entries.all().delete()
                tt.delete()
                deleted_count += 1

            # Reset config
            TimetableConfig.objects.filter(
                academic_year=current_year,
                semester=year_semester,
                program_id=int(program_id),
                year_of_study=yr,
            ).update(is_generated=False)

            return JsonResponse({'success': True, 'deleted': deleted_count})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@csrf_exempt
def api_generate_all_timetables(request):
    """
    POST: Generate timetables for ALL program+year configs at once.
    This ensures cross-program faculty conflict detection.
    Uses the correct semester for each year_of_study (Year 1→Sem 2, Year 2→Sem 4, etc.)
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        generation_preferences = data.get('generation_preferences') or {}
        if not isinstance(generation_preferences, dict):
            generation_preferences = {}

        current_year = AcademicYear.get_current()

        if not current_year:
            return JsonResponse({'error': 'No current academic year configured'}, status=400)

        # Find all program+year combos, including cohort-mapped years
        combos_with_batches = _get_program_year_combos(current_year)

        all_configs = []
        for combo in combos_with_batches:
            yr = combo['year_of_study']
            pid = combo['program__id']
            semester = _select_target_semester_for_program_year(
                current_year=current_year,
                program_id=pid,
                year_of_study=yr,
                require_assignments=True,
            )

            if not semester:
                continue  # Skip if no semester found

            # Get or create a single config for this program+year+semester
            config = TimetableConfig.objects.filter(
                academic_year=current_year,
                semester=semester,
                program_id=pid,
                year_of_study=yr,
            ).first()

            if not config:
                config = TimetableConfig.objects.create(
                    academic_year=current_year,
                    semester=semester,
                    program_id=pid,
                    year_of_study=yr,
                    created_by=request.user,
                )

            all_configs.append(config)

        if not all_configs:
            return JsonResponse({'error': 'No program/year/semester combinations found'}, status=400)

        from main_app.utils.timetable_engine import TimetableEngine
        result = TimetableEngine.generate_all(
            all_configs,
            created_by=request.user,
            generation_preferences=generation_preferences,
        )

        if not result['success']:
            return JsonResponse({'error': result.get('error', 'Generation failed')}, status=400)

        return JsonResponse({
            'success': True,
            'timetables': result['timetables'],
            'warnings': result['warnings'],
            'redirect_url': reverse('manage_timetables'),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_get_me_lab_assist_mappings(request):
    """
    GET: Return M.E class rows, available B.E lab courses, and saved mappings.

    Params:
      - academic_year_id (optional)
      - semester_type ('ODD'|'EVEN', optional)
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    academic_year_id = request.GET.get('academic_year_id')
    semester_type = (request.GET.get('semester_type') or '').upper().strip()

    if academic_year_id:
        current_year = AcademicYear.objects.filter(id=academic_year_id).first()
    else:
        current_year = AcademicYear.get_current()

    if not current_year:
        return JsonResponse({'error': 'No active academic year found'}, status=400)

    if semester_type not in {'ODD', 'EVEN'}:
        current_semester = Semester.get_current()
        semester_type = current_semester.semester_type if current_semester else 'ODD'

    sem_numbers = _semester_numbers_for_type(semester_type)

    combo_rows = _get_program_year_combos(current_year)
    program_ids = [row['program__id'] for row in combo_rows]
    programs = {
        p.id: p for p in Program.objects.filter(id__in=program_ids)
    }

    me_classes = []
    for row in combo_rows:
        program = programs.get(row['program__id'])
        if not program or program.degree != 'ME':
            continue

        me_classes.append({
            'program_id': program.id,
            'program_code': program.code,
            'program_name': program.name,
            'year_of_study': row['year_of_study'],
            'label': f"{program.get_degree_display()} {program.code} - Year {row['year_of_study']}",
            'batch_count': len(row.get('batch_names', [])),
        })

    me_classes.sort(key=lambda item: (item['program_code'], item['year_of_study']))

    be_lab_assignments = Course_Assignment.objects.filter(
        academic_year=current_year,
        semester__semester_number__in=sem_numbers,
        batch__program__degree='BE',
        batch__is_active=True,
        is_active=True,
        course__course_type__in=['L', 'LIT'],
    ).select_related('course', 'batch__program', 'semester')

    be_course_map = {}
    for assignment in be_lab_assignments:
        course = assignment.course
        if not course:
            continue

        code = course.course_code
        if code not in be_course_map:
            be_course_map[code] = {
                'course_code': code,
                'title': course.title,
                'contexts': set(),
            }

        context_label = (
            f"{assignment.batch.program.code} Y{assignment.semester.year_of_study} "
            f"(Sem {assignment.semester.semester_number})"
        )
        be_course_map[code]['contexts'].add(context_label)

    be_lab_courses = []
    for code in sorted(be_course_map.keys()):
        info = be_course_map[code]
        be_lab_courses.append({
            'course_code': code,
            'title': info['title'],
            'contexts': sorted(info['contexts']),
        })

    saved_rows = MELabAssistConstraint.objects.filter(
        academic_year=current_year,
        semester_type=semester_type,
        me_program__degree='ME',
    ).select_related('be_lab_course')

    mappings = {}
    for row in saved_rows:
        key = f"{row.me_program_id}_{row.me_year_of_study}"
        mappings.setdefault(key, []).append(row.be_lab_course.course_code)

    return JsonResponse({
        'success': True,
        'academic_year_id': current_year.id,
        'academic_year': str(current_year),
        'semester_type': semester_type,
        'me_classes': me_classes,
        'be_lab_courses': be_lab_courses,
        'mappings': mappings,
    })


@login_required
@csrf_exempt
def api_save_me_lab_assist_mappings(request):
    """
    POST: Replace all M.E -> B.E lab-assist mappings for academic year + semester type.

    Body:
    {
      academic_year_id,
      semester_type,
      mappings: [
        { me_program_id, me_year_of_study, be_lab_course_codes: ['CSXXXX', ...] }
      ]
    }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)

        academic_year_id = data.get('academic_year_id')
        semester_type = (data.get('semester_type') or '').upper().strip()
        mapping_rows = data.get('mappings', [])

        if academic_year_id:
            current_year = AcademicYear.objects.filter(id=academic_year_id).first()
        else:
            current_year = AcademicYear.get_current()

        if not current_year:
            return JsonResponse({'error': 'No active academic year found'}, status=400)

        if semester_type not in {'ODD', 'EVEN'}:
            current_semester = Semester.get_current()
            semester_type = current_semester.semester_type if current_semester else 'ODD'

        sem_numbers = _semester_numbers_for_type(semester_type)

        me_programs = {
            p.id: p for p in Program.objects.filter(degree='ME')
        }

        valid_be_lab_codes = set(
            Course_Assignment.objects.filter(
                academic_year=current_year,
                semester__semester_number__in=sem_numbers,
                batch__program__degree='BE',
                batch__is_active=True,
                is_active=True,
                course__course_type__in=['L', 'LIT'],
            ).values_list('course__course_code', flat=True)
        )

        create_rows = []
        seen_pairs = set()
        for row in mapping_rows:
            me_program_id = int(row.get('me_program_id')) if row.get('me_program_id') else None
            me_year_of_study = int(row.get('me_year_of_study')) if row.get('me_year_of_study') else None
            selected_codes = row.get('be_lab_course_codes', []) or []

            if not me_program_id or not me_year_of_study:
                continue

            me_program = me_programs.get(me_program_id)
            if not me_program:
                continue

            for course_code in sorted(set(selected_codes)):
                if course_code not in valid_be_lab_codes:
                    continue

                unique_key = (me_program_id, me_year_of_study, course_code)
                if unique_key in seen_pairs:
                    continue
                seen_pairs.add(unique_key)

                create_rows.append(MELabAssistConstraint(
                    academic_year=current_year,
                    semester_type=semester_type,
                    me_program_id=me_program_id,
                    me_year_of_study=me_year_of_study,
                    be_lab_course_id=course_code,
                    created_by=request.user,
                ))

        with transaction.atomic():
            MELabAssistConstraint.objects.filter(
                academic_year=current_year,
                semester_type=semester_type,
            ).delete()

            if create_rows:
                MELabAssistConstraint.objects.bulk_create(create_rows)

        return JsonResponse({
            'success': True,
            'saved_count': len(create_rows),
            'semester_type': semester_type,
            'academic_year_id': current_year.id,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


# ──────────────────────────────────────────────────
# Timetable Wizard — Same-Time Constraint APIs
# ──────────────────────────────────────────────────

@csrf_exempt
def api_toggle_same_time_constraint(request):
    """
    POST: Toggle a course as same-time-for-all-batches within a config.
    If the constraint already exists, delete it (un-toggle); otherwise create it.
    Body: { config_id, academic_year_id, semester_id, program_id, year_of_study, course_code }
    Returns: { success, enabled (bool), config_id, constraints: [...] }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        course_code = data.get('course_code')
        if not course_code:
            return JsonResponse({'error': 'course_code required'}, status=400)

        course = Course.objects.get(course_code=course_code)

        # Get or create config
        config_id = data.get('config_id')
        if config_id:
            config = TimetableConfig.objects.get(id=config_id)
        else:
            academic_year = AcademicYear.objects.get(id=data['academic_year_id'])
            semester = Semester.objects.get(id=data['semester_id'])
            program = Program.objects.get(id=data['program_id'])
            year_of_study = int(data['year_of_study'])
            config, _ = TimetableConfig.objects.get_or_create(
                academic_year=academic_year,
                semester=semester,
                program=program,
                year_of_study=year_of_study,
                defaults={'created_by': request.user},
            )

        # Toggle
        existing = SameTimeConstraint.objects.filter(config=config, course=course)
        if existing.exists():
            existing.delete()
            enabled = False
        else:
            SameTimeConstraint.objects.create(config=config, course=course)
            enabled = True

        # Return current list
        constraints = list(
            SameTimeConstraint.objects.filter(config=config)
            .values_list('course__course_code', flat=True)
        )

        return JsonResponse({
            'success': True,
            'enabled': enabled,
            'config_id': config.id,
            'constraints': constraints,
        })

    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


def api_get_same_time_constraints(request):
    """
    GET: Return all same-time constraints for a given config_id.
    Params: config_id
    Returns: { success, constraints: [course_code, ...] }
    """
    config_id = request.GET.get('config_id')
    if not config_id:
        return JsonResponse({'success': True, 'constraints': []})

    constraints = list(
        SameTimeConstraint.objects.filter(config_id=config_id)
        .values_list('course__course_code', flat=True)
    )
    return JsonResponse({'success': True, 'constraints': constraints})


def _resolve_placeholder_offerings(config, placeholder_code):
    """Return (placeholder_course, offerings_qs) for a config+placeholder code."""
    placeholder_course = Course.objects.filter(
        course_code=placeholder_code,
        is_placeholder=True,
    ).first()
    if not placeholder_course:
        return None, ElectiveCourseOffering.objects.none()

    plan = RegulationCoursePlan.objects.filter(
        regulation=config.program.regulation,
        semester=config.semester.semester_number,
        branch=config.program.code,
        course=placeholder_course,
        is_elective=True,
    ).first()
    if not plan:
        return placeholder_course, ElectiveCourseOffering.objects.none()

    offerings = ElectiveCourseOffering.objects.filter(
        regulation_course_plan=plan,
        semester=config.semester,
        is_active=True,
    ).select_related('actual_course').order_by('actual_course__course_code')
    return placeholder_course, offerings


def api_get_pec_combination_rules(request):
    """
    GET: Fetch PEC combination matrix context + existing rules.
    Params: config_id, slot_a_code, slot_b_code
    """
    try:
        config_id = request.GET.get('config_id')
        slot_a_code = request.GET.get('slot_a_code')
        slot_b_code = request.GET.get('slot_b_code')

        if not config_id or not slot_a_code or not slot_b_code:
            return JsonResponse({'error': 'config_id, slot_a_code, slot_b_code required'}, status=400)

        config = TimetableConfig.objects.select_related('program', 'semester').get(id=config_id)

        slot_a_course, slot_a_offerings_qs = _resolve_placeholder_offerings(config, slot_a_code)
        slot_b_course, slot_b_offerings_qs = _resolve_placeholder_offerings(config, slot_b_code)

        if not slot_a_course or not slot_b_course:
            return JsonResponse({'error': 'Invalid placeholder course code(s)'}, status=404)

        rule_rows = PECCourseCombinationRule.objects.filter(
            config=config,
            slot_a_course=slot_a_course,
            slot_b_course=slot_b_course,
        ).select_related('offering_a__actual_course', 'offering_b__actual_course')

        rules = []
        for row in rule_rows:
            rules.append({
                'id': row.id,
                'offering_a_id': row.offering_a_id,
                'offering_a_code': row.offering_a.actual_course.course_code,
                'offering_b_id': row.offering_b_id,
                'offering_b_code': row.offering_b.actual_course.course_code,
                'can_overlap': row.can_overlap,
            })

        return JsonResponse({
            'success': True,
            'slot_a': {
                'course_code': slot_a_course.course_code,
                'title': slot_a_course.title,
                'offerings': [
                    {
                        'id': o.id,
                        'course_code': o.actual_course.course_code,
                        'title': o.actual_course.title,
                        'ltp': o.actual_course.ltp_display,
                    }
                    for o in slot_a_offerings_qs
                ],
            },
            'slot_b': {
                'course_code': slot_b_course.course_code,
                'title': slot_b_course.title,
                'offerings': [
                    {
                        'id': o.id,
                        'course_code': o.actual_course.course_code,
                        'title': o.actual_course.title,
                        'ltp': o.actual_course.ltp_display,
                    }
                    for o in slot_b_offerings_qs
                ],
            },
            'rules': rules,
        })
    except TimetableConfig.DoesNotExist:
        return JsonResponse({'error': 'TimetableConfig not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_save_pec_combination_rules(request):
    """
    POST: Replace PEC combination rules for one slot pair in a config.
    Body: {config_id, slot_a_code, slot_b_code, rules:[{offering_a_id, offering_b_id, can_overlap}]}
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        config_id = data.get('config_id')
        slot_a_code = data.get('slot_a_code')
        slot_b_code = data.get('slot_b_code')
        rules = data.get('rules', [])

        if not config_id or not slot_a_code or not slot_b_code:
            return JsonResponse({'error': 'config_id, slot_a_code, slot_b_code required'}, status=400)

        if slot_a_code == slot_b_code:
            return JsonResponse({'error': 'slot_a_code and slot_b_code must be different'}, status=400)

        # Canonical slot ordering avoids duplicate mirrored records.
        if slot_a_code > slot_b_code:
            slot_a_code, slot_b_code = slot_b_code, slot_a_code
            rules = [
                {
                    'offering_a_id': row.get('offering_b_id'),
                    'offering_b_id': row.get('offering_a_id'),
                    'can_overlap': row.get('can_overlap', True),
                }
                for row in rules
            ]

        config = TimetableConfig.objects.select_related('program', 'semester').get(id=config_id)
        slot_a_course, slot_a_offerings_qs = _resolve_placeholder_offerings(config, slot_a_code)
        slot_b_course, slot_b_offerings_qs = _resolve_placeholder_offerings(config, slot_b_code)

        if not slot_a_course or not slot_b_course:
            return JsonResponse({'error': 'Invalid placeholder course code(s)'}, status=404)

        valid_a_ids = set(slot_a_offerings_qs.values_list('id', flat=True))
        valid_b_ids = set(slot_b_offerings_qs.values_list('id', flat=True))

        with transaction.atomic():
            PECCourseCombinationRule.objects.filter(
                config=config,
            ).filter(
                Q(slot_a_course=slot_a_course, slot_b_course=slot_b_course)
                | Q(slot_a_course=slot_b_course, slot_b_course=slot_a_course)
            ).delete()

            create_rows = []
            for row in rules:
                offering_a_id = int(row.get('offering_a_id'))
                offering_b_id = int(row.get('offering_b_id'))
                can_overlap = bool(row.get('can_overlap', True))

                if offering_a_id not in valid_a_ids or offering_b_id not in valid_b_ids:
                    return JsonResponse({'error': 'Invalid offering mapping for selected slots'}, status=400)

                create_rows.append(PECCourseCombinationRule(
                    config=config,
                    slot_a_course=slot_a_course,
                    slot_b_course=slot_b_course,
                    offering_a_id=offering_a_id,
                    offering_b_id=offering_b_id,
                    can_overlap=can_overlap,
                ))

            if create_rows:
                PECCourseCombinationRule.objects.bulk_create(create_rows)

        return JsonResponse({'success': True, 'saved_count': len(create_rows)})
    except TimetableConfig.DoesNotExist:
        return JsonResponse({'error': 'TimetableConfig not found'}, status=404)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid rules payload'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


def api_validate_pec_combination_rules(request):
    """
    GET: Validate whether complete matrix rules exist for a slot pair.
    Params: config_id, slot_a_code, slot_b_code
    """
    try:
        config_id = request.GET.get('config_id')
        slot_a_code = request.GET.get('slot_a_code')
        slot_b_code = request.GET.get('slot_b_code')

        if not config_id or not slot_a_code or not slot_b_code:
            return JsonResponse({'error': 'config_id, slot_a_code, slot_b_code required'}, status=400)

        config = TimetableConfig.objects.select_related('program', 'semester').get(id=config_id)
        slot_a_course, slot_a_offerings_qs = _resolve_placeholder_offerings(config, slot_a_code)
        slot_b_course, slot_b_offerings_qs = _resolve_placeholder_offerings(config, slot_b_code)

        if not slot_a_course or not slot_b_course:
            return JsonResponse({'error': 'Invalid placeholder course code(s)'}, status=404)

        expected_pairs = slot_a_offerings_qs.count() * slot_b_offerings_qs.count()
        configured_pairs = PECCourseCombinationRule.objects.filter(
            config=config,
        ).filter(
            Q(slot_a_course=slot_a_course, slot_b_course=slot_b_course)
            | Q(slot_a_course=slot_b_course, slot_b_course=slot_a_course)
        ).count()

        return JsonResponse({
            'success': True,
            'is_complete': expected_pairs == configured_pairs,
            'expected_pairs': expected_pairs,
            'configured_pairs': configured_pairs,
        })
    except TimetableConfig.DoesNotExist:
        return JsonResponse({'error': 'TimetableConfig not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


# ──────────────────────────────────────────────────
# Timetable Wizard — Faculty Time-Block APIs
# ──────────────────────────────────────────────────

def api_get_faculty_time_blocks(request):
    """
    GET: Return all faculty time-blocks for the current academic year & semester.
    Optional query param: faculty_id — filter to a single faculty.
    Returns: { success, blocks: [{id, faculty_id, faculty_name, day, slot_number, reason}, ...] }
    """
    try:
        semester_id = request.GET.get('semester_id')
        academic_year_id = request.GET.get('academic_year_id')

        if semester_id:
            current_semester = Semester.objects.filter(id=semester_id).select_related('academic_year').first()
            if not current_semester:
                return JsonResponse({'error': 'Semester not found'}, status=404)
            current_year = current_semester.academic_year
            if academic_year_id and str(current_year.id) != str(academic_year_id):
                return JsonResponse({'error': 'semester_id and academic_year_id mismatch'}, status=400)
        else:
            current_semester = Semester.get_current()
            if not current_semester:
                return JsonResponse({'success': True, 'blocks': []})
            current_year = current_semester.academic_year

        target_sem_numbers = [1, 3, 5, 7] if current_semester.semester_type == 'ODD' else [2, 4, 6, 8]

        qs = FacultyTimeBlock.objects.filter(
            academic_year=current_year,
            semester__semester_number__in=target_sem_numbers,
        ).select_related('faculty__user', 'time_slot')

        faculty_id = request.GET.get('faculty_id')
        if faculty_id:
            qs = qs.filter(faculty_id=faculty_id)

        blocks = []
        for b in qs:
            blocks.append({
                'id': b.id,
                'faculty_id': b.faculty_id,
                'faculty_name': b.faculty.user.full_name if hasattr(b.faculty.user, 'full_name') else str(b.faculty),
                'staff_id': b.faculty.staff_id,
                'day': b.day,
                'slot_number': b.time_slot.slot_number,
                'reason': b.reason,
            })

        return JsonResponse({'success': True, 'blocks': blocks})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_toggle_faculty_time_block(request):
    """
    POST: Toggle a faculty time-block on/off.
    Body: { faculty_id, day, slot_number, reason (optional) }
    Returns: { success, enabled, blocks: [...] } — returns ALL blocks for that faculty.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        faculty_id = int(data['faculty_id'])
        day = data['day']
        slot_number = int(data['slot_number'])
        reason = data.get('reason', '')

        semester_id = data.get('semester_id')
        academic_year_id = data.get('academic_year_id')

        if semester_id:
            current_semester = Semester.objects.filter(id=semester_id).select_related('academic_year').first()
            if not current_semester:
                return JsonResponse({'error': 'Semester not found'}, status=404)
            current_year = current_semester.academic_year
            if academic_year_id and str(current_year.id) != str(academic_year_id):
                return JsonResponse({'error': 'semester_id and academic_year_id mismatch'}, status=400)
        else:
            current_semester = Semester.get_current()
            if not current_semester:
                return JsonResponse({'error': 'No current semester found'}, status=400)
            current_year = current_semester.academic_year

        faculty = Faculty_Profile.objects.get(id=faculty_id)
        time_slot = TimeSlot.objects.get(slot_number=slot_number)

        target_sem_numbers = [1, 3, 5, 7] if current_semester.semester_type == 'ODD' else [2, 4, 6, 8]

        existing = FacultyTimeBlock.objects.filter(
            academic_year=current_year,
            semester__semester_number__in=target_sem_numbers,
            faculty=faculty,
            day=day,
            time_slot=time_slot,
        )
        if existing.exists():
            existing.delete()
            enabled = False
        else:
            FacultyTimeBlock.objects.create(
                academic_year=current_year,
                semester=current_semester,
                faculty=faculty,
                day=day,
                time_slot=time_slot,
                reason=reason,
                created_by=request.user,
            )
            enabled = True

        # Return updated blocks for this faculty
        blocks = list(
            FacultyTimeBlock.objects.filter(
                academic_year=current_year,
                semester__semester_number__in=target_sem_numbers,
                faculty=faculty,
            ).select_related('time_slot').values_list('day', 'time_slot__slot_number')
        )
        block_list = [{'day': d, 'slot_number': s} for d, s in blocks]

        return JsonResponse({
            'success': True,
            'enabled': enabled,
            'blocks': block_list,
        })

    except Faculty_Profile.DoesNotExist:
        return JsonResponse({'error': 'Faculty not found'}, status=404)
    except TimeSlot.DoesNotExist:
        return JsonResponse({'error': 'Time slot not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_bulk_save_faculty_time_blocks(request):
    """
    POST: Save a full set of blocks for one faculty at once (replaces existing).
    Body: { faculty_id, blocks: [{day, slot_number, reason},...] }
    Returns: { success, saved_count }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        faculty_id = int(data['faculty_id'])
        blocks_data = data.get('blocks', [])

        semester_id = data.get('semester_id')
        academic_year_id = data.get('academic_year_id')

        if semester_id:
            current_semester = Semester.objects.filter(id=semester_id).select_related('academic_year').first()
            if not current_semester:
                return JsonResponse({'error': 'Semester not found'}, status=404)
            current_year = current_semester.academic_year
            if academic_year_id and str(current_year.id) != str(academic_year_id):
                return JsonResponse({'error': 'semester_id and academic_year_id mismatch'}, status=400)
        else:
            current_semester = Semester.get_current()
            if not current_semester:
                return JsonResponse({'error': 'No current semester found'}, status=400)
            current_year = current_semester.academic_year

        faculty = Faculty_Profile.objects.get(id=faculty_id)

        target_sem_numbers = [1, 3, 5, 7] if current_semester.semester_type == 'ODD' else [2, 4, 6, 8]

        with transaction.atomic():
            # Delete existing blocks for this faculty for the same semester type (ODD/EVEN)
            FacultyTimeBlock.objects.filter(
                academic_year=current_year,
                semester__semester_number__in=target_sem_numbers,
                faculty=faculty,
            ).delete()

            # Create new blocks
            for b in blocks_data:
                ts = TimeSlot.objects.get(slot_number=int(b['slot_number']))
                FacultyTimeBlock.objects.create(
                    academic_year=current_year,
                    semester=current_semester,
                    faculty=faculty,
                    day=b['day'],
                    time_slot=ts,
                    reason=b.get('reason', ''),
                    created_by=request.user,
                )

        return JsonResponse({
            'success': True,
            'saved_count': len(blocks_data),
        })

    except Faculty_Profile.DoesNotExist:
        return JsonResponse({'error': 'Faculty not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@csrf_exempt
def api_save_pec_arrangement(request):
    """
    Save the PEC scheduling-group configuration ONLY.
    Groups define scheduling constraints — NOT PEC slot identity.
    
    - Same group: courses MAY run in parallel (algo's choice)
    - Different groups: courses MUST run at different time slots
    
    This does NOT create/modify ElectiveCourseOffering records.
    Faculty assignment and PEC slot mapping (for transcripts) are handled
    separately via the "Manage" button on each PEC placeholder slot.
    
    Expects JSON:
    {
        "semester_id": 24,
        "branch": "CSE",
        "program_type": "UG",
        "plan_ids": [22, 23],
        "groups": [
            [{"code":"CS8091","title":"IP","credits":3,"type":"T","batch_count":1,"capacity":30}, ...],
            [{"code":"CS8093","title":"PP","credits":3,"type":"T","batch_count":1,"capacity":30}],
            [{"code":"CS8094","title":"SPM","credits":3,"type":"T","batch_count":1,"capacity":30}],
        ]
    }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    import json as _json
    
    try:
        data = _json.loads(request.body)
        semester_id = data.get('semester_id')
        branch = data.get('branch', '')
        program_type = data.get('program_type', 'UG')
        plan_ids = data.get('plan_ids', [])
        groups = data.get('groups', [])
        
        if not semester_id or not branch:
            return JsonResponse({'error': 'semester_id and branch required'}, status=400)
        
        semester = Semester.objects.get(id=semester_id)
        
        # Validate plan_ids are PEC placeholders
        plans = RegulationCoursePlan.objects.filter(
            id__in=plan_ids
        ).select_related('course')
        plan_map = {p.id: p for p in plans}
        
        for pid in plan_ids:
            if pid not in plan_map:
                return JsonResponse({'error': f'Invalid plan_id: {pid}'}, status=400)
            if not plan_map[pid].course.is_placeholder or plan_map[pid].course.placeholder_type != 'PEC':
                return JsonResponse({'error': f'Plan {pid} is not a PEC placeholder'}, status=400)
        
        min_groups = len(plan_ids)
        # Previously we required at least one group per PEC slot. Allow fewer groups
        # and distribute groups across available PEC slots in round-robin fashion.
        
        # Find regulation from first plan
        regulation = plans.first().regulation if plans.exists() else None
        
        # Save/update the PECGroupConfig (scheduling groups only)
        config, created = PECGroupConfig.objects.update_or_create(
            semester=semester,
            branch=branch,
            program_type=program_type,
            defaults={
                'regulation': regulation,
                'groups': groups,
                'min_groups': min_groups,
            }
        )

        result = {
            'success': True,
            'config_id': config.id,
            'group_count': len(groups),
        }

        # Optional: push offerings into PEC slots
        push = data.get('push_offerings') or data.get('push') or False
        if push:
            created_count = 0
            updated_count = 0
            removed_count = 0
            # Map group index -> plan id (round-robin across available plan_ids)
            for gi, group in enumerate(groups):
                if len(plan_ids) == 0:
                    break
                target_plan_id = plan_ids[gi % len(plan_ids)]
                plan = plan_map.get(target_plan_id)
                if not plan:
                    continue
                # For each course object in this group, create offering if missing
                for course_obj in group:
                    code = course_obj.get('code') or course_obj.get('course_code')
                    if not code:
                        continue
                    try:
                        course = Course.objects.get(course_code=code)
                    except Course.DoesNotExist:
                        continue

                    # Check existing offering for this plan + course + semester
                    existing = ElectiveCourseOffering.objects.filter(
                        regulation_course_plan=plan,
                        actual_course=course,
                        semester=semester
                    ).first()

                    batch_count = int(course_obj.get('batch_count', 1) or 1)
                    capacity = int(course_obj.get('capacity', course.credits or 30) or 30)

                    if existing:
                        # Update counts if different
                        changed = False
                        if existing.batch_count != batch_count:
                            existing.batch_count = batch_count
                            changed = True
                        if existing.capacity_per_batch != capacity:
                            existing.capacity_per_batch = capacity
                            changed = True
                        if changed:
                            existing.save()
                            updated_count += 1
                    else:
                        ElectiveCourseOffering.objects.create(
                            regulation_course_plan=plan,
                            semester=semester,
                            actual_course=course,
                            batch_count=batch_count,
                            capacity_per_batch=capacity,
                            elective_vertical=plan.elective_vertical
                        )
                        created_count += 1

            result['created'] = created_count
            result['updated'] = updated_count
            result['removed'] = removed_count

        return JsonResponse(result)
    
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'Semester not found'}, status=404)
    except _json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


# ==========================================================================
# PG COURSE CLUBBING APIs
# ==========================================================================

def _get_program_batches_for_offering(offering, current_year):
    return list(
        ProgramBatch.objects.filter(
            academic_year=current_year,
            program__code=offering.regulation_course_plan.branch,
            program__level=offering.regulation_course_plan.program_type,
            year_of_study=offering.semester.year_of_study,
            is_active=True,
        ).select_related('program').order_by('batch_name')
    )


def _resolve_assignment_effective_course_for_pg_clubbing(assignment, mapped_course_cache, quick_ioc_cache):
    """
    Resolve the effective real course used for PG clubbing detection.

    IOC/EEC placeholders can represent a mapped real catalog course. For clubbing,
    we should group by that real course code (when available), not by the placeholder code.
    """
    course = assignment.course
    if not course or not course.is_placeholder:
        return course

    note = assignment.special_note or ''
    mapped_match = re.search(r'Mapped\s+Course\s*:\s*([A-Za-z0-9\-]+)', note, re.IGNORECASE)
    if mapped_match:
        mapped_code = mapped_match.group(1).strip()
        if mapped_code:
            mapped_course = mapped_course_cache.get(mapped_code)
            if mapped_course is None and mapped_code not in mapped_course_cache:
                mapped_course = Course.objects.filter(course_code=mapped_code).first()
                mapped_course_cache[mapped_code] = mapped_course
            if mapped_course:
                return mapped_course

    # Quick IOC flow creates a real-course assignment + synced placeholder assignment.
    # If this row is the placeholder one, infer the real course from the quick IOC row.
    if course.placeholder_type in ['IOC', 'EEC']:
        cache_key = (
            assignment.academic_year_id,
            assignment.semester_id,
            assignment.batch_id,
            assignment.faculty_id,
        )
        if cache_key not in quick_ioc_cache:
            quick_ioc_cache[cache_key] = (
                Course_Assignment.objects
                .filter(
                    academic_year_id=assignment.academic_year_id,
                    semester_id=assignment.semester_id,
                    batch_id=assignment.batch_id,
                    faculty_id=assignment.faculty_id,
                    special_note__icontains='Quick IOC',
                    is_active=True,
                )
                .exclude(course__is_placeholder=True)
                .select_related('course')
                .first()
            )
        quick_ioc_assignment = quick_ioc_cache.get(cache_key)
        if quick_ioc_assignment and quick_ioc_assignment.course:
            return quick_ioc_assignment.course

    return course


def _collect_pg_clubbing_candidates(current_year, year_filter=None):
    """
    Collect PG course+faculty combos that appear in 2+ PG programs.
    
    Supports clubbing across course categories (IOC, PEC, PCC, ECC, etc.) 
    as long as they reference the same actual course code and faculty.
    
    For example, if CS23001-Data Mining is:
      - Assigned to ME CSE as an IOC
      - Assigned to ME SE as a PEC
      Both taught by the same faculty → they can be clubbed together.
    """
    from collections import defaultdict

    candidate_groups = defaultdict(lambda: {
        'course': None,
        'faculty': None,
        'semester': None,
        'programs': {},
        'category_types': set(),  # Track what categories this course appears as
    })
    mapped_course_cache = {}
    quick_ioc_cache = {}

    def add_candidate(course, faculty, semester, program, batch_id, category_type=None):
        if not faculty or not program or not semester:
            return
        key = (course.course_code, faculty.id, semester.id)
        group = candidate_groups[key]
        group['course'] = course
        group['faculty'] = faculty
        group['semester'] = semester
        if category_type:
            group['category_types'].add(category_type)
            
        pdata = group['programs'].setdefault(program.id, {
            'program_id': program.id,
            'program_code': program.code,
            'program_name': str(program),
            'batch_ids': [],
        })
        if batch_id and batch_id not in pdata['batch_ids']:
            pdata['batch_ids'].append(batch_id)

    # ── Collect from regular course assignments ──
    assignments = (
        Course_Assignment.objects
        .filter(
            academic_year=current_year,
            batch__program__level='PG',
            batch__is_active=True,
            is_active=True,
        )
        .select_related('course', 'faculty', 'faculty__user', 'batch', 'batch__program', 'semester')
    )
    if year_filter:
        year_value = int(year_filter)
        assignments = assignments.filter(
            Q(semester__year_of_study=year_value) |
            Q(batch__year_of_study=year_value)
        )

    for assignment in assignments:
        effective_course = _resolve_assignment_effective_course_for_pg_clubbing(
            assignment,
            mapped_course_cache,
            quick_ioc_cache,
        )
        if not effective_course:
            continue

        add_candidate(
            effective_course,
            assignment.faculty,
            assignment.semester,
            assignment.batch.program if assignment.batch_id else None,
            assignment.batch_id,
            category_type='Core' if not assignment.course.is_placeholder else assignment.course.placeholder_type,
        )

    # ── Collect from elective offerings (including IOC/EEC) ──
    # This allows clubbing of the SAME REAL COURSE across different categories
    # (e.g., CS23001 as IOC in one program and PEC in another)
    offering_qs = ElectiveOfferingFacultyAssignment.objects.filter(
        offering__semester__academic_year=current_year,
        offering__regulation_course_plan__program_type='PG',
        offering__is_active=True,
        is_active=True,
        faculty_id__isnull=False,
    ).select_related(
        'faculty__user',
        'offering__actual_course',
        'offering__semester',
        'offering__regulation_course_plan',
    )
    if year_filter:
        offering_qs = offering_qs.filter(offering__semester__year_of_study=int(year_filter))

    for fa in offering_qs:
        offering = fa.offering
        batches = _get_program_batches_for_offering(offering, current_year)
        target_batch = None
        if batches and 1 <= fa.batch_number <= len(batches):
            target_batch = batches[fa.batch_number - 1]
        elif batches:
            target_batch = batches[0]

        program = target_batch.program if target_batch else Program.objects.filter(
            code=offering.regulation_course_plan.branch,
            level=offering.regulation_course_plan.program_type,
        ).first()

        # Track the category from the regulation course plan
        category_display = offering.regulation_course_plan.category.code if offering.regulation_course_plan.category else 'Elective'

        add_candidate(
            offering.actual_course,
            fa.faculty,
            offering.semester,
            program,
            target_batch.id if target_batch else None,
            category_type=category_display,
        )

    return candidate_groups

@login_required
def api_get_pg_clubbing_candidates(request):
    """
    GET: Auto-detect clubbing candidates for PG programs.
    Finds course+faculty combos that appear in 2+ PG programs for the
    current academic year / semester.

    Optional query params:
        year_of_study  – filter for specific year (default: all PG years)

    Returns {
      success, candidates: [
        { course_code, course_title, faculty_id, faculty_name,
          programs: [{program_id, program_code, program_name, batch_ids: [..]}],
          is_clubbed, group_id }
      ]
    }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    current_year = AcademicYear.get_current()
    if not current_year:
        return JsonResponse({'error': 'No active academic year'}, status=400)

    year_filter = request.GET.get('year_of_study')

    groups = _collect_pg_clubbing_candidates(current_year, year_filter=year_filter)

    candidates = []
    for (course_code, faculty_id, semester_id), payload in groups.items():
        program_map = payload['programs']
        if len(program_map) < 2:
            continue  # Only clubbable if 2+ programs share this course+faculty

        # Check if already clubbed
        existing_group = ClubbedCourseGroup.objects.filter(
            academic_year=current_year,
            semester_id=semester_id,
            course__course_code=course_code,
            faculty_id=faculty_id,
            is_active=True,
        ).first()

        course = payload['course']
        faculty = payload['faculty']
        semester = payload['semester']
        category_types = sorted(payload['category_types'])  # e.g., ['IOC', 'PEC']
        
        candidates.append({
            'course_code': course_code,
            'course_title': course.title,
            'faculty_id': faculty_id,
            'faculty_name': faculty.user.full_name,
            'semester_id': semester_id,
            'semester_label': str(semester),
            'programs': list(program_map.values()),
            'is_clubbed': existing_group is not None,
            'group_id': existing_group.id if existing_group else None,
            'category_types': category_types,  # e.g., ['IOC', 'PEC'] for cross-category clubbing
            'is_ioc_related': any(cat in category_types for cat in ['IOC', 'EEC']),  # Flag if any IOC/EEC
        })

    candidates.sort(key=lambda c: c['course_code'])
    return JsonResponse({'success': True, 'candidates': candidates})


@csrf_exempt
@login_required
def api_toggle_pg_clubbing(request):
    """
    POST: Toggle clubbing for a PG course+faculty combo.
    Body: { course_code, faculty_id, semester_id }

    If not clubbed → creates ClubbedCourseGroup + ClubbedCourseMembers for
    every PG ProgramBatch that has a matching Course_Assignment.
    If already clubbed → deactivates the group.

    Returns { success, is_clubbed, group_id }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body)
        course_code = data['course_code']
        faculty_id = int(data['faculty_id'])
        semester_id = int(data['semester_id'])

        current_year = AcademicYear.get_current()
        if not current_year:
            return JsonResponse({'error': 'No active academic year'}, status=400)

        course = Course.objects.get(course_code=course_code)
        faculty = Faculty_Profile.objects.get(id=faculty_id)
        semester = Semester.objects.get(id=semester_id)

        # Check if already clubbed
        existing = ClubbedCourseGroup.objects.filter(
            academic_year=current_year,
            semester=semester,
            course=course,
            faculty=faculty,
            is_active=True,
        ).first()

        if existing:
            # Un-club: deactivate group + remove members
            existing.is_active = False
            existing.save()
            existing.members.all().delete()
            return JsonResponse({
                'success': True,
                'is_clubbed': False,
                'group_id': None,
            })

        candidate_groups = _collect_pg_clubbing_candidates(current_year)
        payload = candidate_groups.get((course_code, faculty_id, semester_id))
        if not payload or len(payload['programs']) < 2:
            return JsonResponse({
                'error': 'Need at least 2 program-batches to club',
            }, status=400)

        batch_ids = sorted({batch_id for program in payload['programs'].values() for batch_id in program['batch_ids']})

        # Re-activate or create group
        group, created = ClubbedCourseGroup.objects.get_or_create(
            academic_year=current_year,
            semester=semester,
            course=course,
            faculty=faculty,
            defaults={'is_active': True},
        )
        if not created:
            group.is_active = True
            group.save()
            group.members.all().delete()  # rebuild

        # Add members
        for bid in batch_ids:
            ClubbedCourseMember.objects.create(group=group, program_batch_id=bid)

        return JsonResponse({
            'success': True,
            'is_clubbed': True,
            'group_id': group.id,
        })

    except (Course.DoesNotExist, Faculty_Profile.DoesNotExist, Semester.DoesNotExist):
        return JsonResponse({'error': 'Course/Faculty/Semester not found'}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_get_clubbed_groups(request):
    """
    GET: List all active clubbed course groups for the current academic year.
    Returns { success, groups: [ { id, course_code, course_title,
              faculty_name, programs: [..], semester_label } ] }
    """
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    current_year = AcademicYear.get_current()
    if not current_year:
        return JsonResponse({'success': True, 'groups': []})

    groups = (
        ClubbedCourseGroup.objects
        .filter(academic_year=current_year, is_active=True)
        .select_related('course', 'faculty', 'faculty__user', 'semester')
        .prefetch_related('members', 'members__program_batch', 'members__program_batch__program')
    )

    result = []
    for g in groups:
        programs = []
        for m in g.members.select_related('program_batch__program'):
            programs.append({
                'program_code': m.program_batch.program.code,
                'program_name': str(m.program_batch.program),
                'batch_id': m.program_batch_id,
            })
        result.append({
            'id': g.id,
            'course_code': g.course.course_code,
            'course_title': g.course.title,
            'faculty_id': g.faculty_id,
            'faculty_name': g.faculty.user.full_name,
            'semester_id': g.semester_id,
            'semester_label': str(g.semester),
            'programs': programs,
        })

    return JsonResponse({'success': True, 'groups': result})


# =============================================================================
# HOD ANSWER KEY REVIEW
# =============================================================================

@login_required
def hod_review_answer_keys(request):
    """List all submitted answer keys pending HOD review"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')

    pending_keys = StructuredQuestionPaper.objects.filter(
        answer_key_status='SUBMITTED'
    ).select_related('course', 'faculty__user', 'regulation').order_by('-answer_key_submitted_at')

    reviewed_keys = StructuredQuestionPaper.objects.filter(
        answer_key_status__in=['APPROVED', 'REJECTED']
    ).select_related('course', 'faculty__user', 'regulation').order_by('-answer_key_submitted_at')[:20]

    context = {
        'pending_keys': pending_keys,
        'reviewed_keys': reviewed_keys,
        'page_title': 'Review Answer Keys'
    }
    return render(request, "hod_template/review_answer_keys.html", context)


@login_required
def hod_approve_answer_key(request, qp_id):
    """Approve a submitted answer key"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')

    if request.method != 'POST':
        return redirect('hod_review_answer_keys')

    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id, answer_key_status='SUBMITTED')
    qp.answer_key_status = 'APPROVED'
    qp.answer_key_comments = request.POST.get('comments', '')
    qp.save()

    # Notify faculty
    try:
        from .models import NotificationRecipient
        notification = Notification.objects.create(
            type='answer_key_approved',
            message=f"Your answer key for {qp.course.course_code} - {qp.course.title} "
                    f"({qp.exam_month_year}) has been approved.",
            created_by=request.user
        )
        NotificationRecipient.objects.create(
            notification=notification, recipient=qp.faculty.user
        )
    except Exception:
        pass

    messages.success(request, f"Answer key for {qp.course.course_code} approved.")
    return redirect('hod_review_answer_keys')


@login_required
def hod_reject_answer_key(request, qp_id):
    """Reject a submitted answer key with comments"""
    if not check_hod_permission(request.user):
        messages.error(request, "Access Denied. HOD privileges required.")
        return redirect('/')

    if request.method != 'POST':
        return redirect('hod_review_answer_keys')

    qp = get_object_or_404(StructuredQuestionPaper, id=qp_id, answer_key_status='SUBMITTED')
    comments = request.POST.get('comments', '')
    if not comments.strip():
        messages.error(request, "Please provide a reason for rejection.")
        return redirect('hod_review_answer_keys')

    qp.answer_key_status = 'REJECTED'
    qp.answer_key_comments = comments
    qp.save()

    # Notify faculty
    try:
        from .models import NotificationRecipient
        notification = Notification.objects.create(
            type='answer_key_rejected',
            message=f"Your answer key for {qp.course.course_code} - {qp.course.title} "
                    f"({qp.exam_month_year}) has been rejected. Reason: {comments}",
            created_by=request.user
        )
        NotificationRecipient.objects.create(
            notification=notification, recipient=qp.faculty.user
        )
    except Exception:
        pass

    messages.success(request, f"Answer key for {qp.course.course_code} rejected.")
    return redirect('hod_review_answer_keys')
